#!/usr/bin/env python3
"""
Edit Report Analysis
====================

Compiles all evaluator edit reports into a 5-tab Excel workbook for
cross-evaluator and cross-model analysis.

Tabs:
  1. Model Overview      — which model performs best overall?
  2. Task Breakdown      — which fields/tasks need the most editing?
  3. Evaluator Behavior  — are some evaluators stricter? more likely to add terms?
  4. Collection Difficulty — which collections are hardest for LLMs?
  5. Evaluator Notes     — all archivist notes in one place

Usage:
    python edit-report-analysis.py
    python edit-report-analysis.py --evaluations-dir /path/to/evaluations
    python edit-report-analysis.py --output-dir /path/to/output

Expects edit_report_*.json files at:
    {evaluations_dir}/{EvaluatorName}/{EvaluatorName}_changes/edit_report_*.json

Output: edit_report_analysis_{YYYY-MM-DD}.xlsx
"""

import os
import sys
import re
import json
import argparse
import statistics
from datetime import date
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

TEXT_FIELDS = [
    "title", "description", "format_media", "genre", "medium",
    "support", "sheet_info", "date_on_drawing", "content_warning",
]
LIST_FIELDS = ["contributors", "named_entities", "geographic_entities"]

HEADER_FILL      = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT      = Font(color="FFFFFF", bold=True)
TITLE_FONT       = Font(bold=True, size=13)
CENTER           = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_WRAP         = Alignment(vertical="top", wrap_text=True)
EXPLANATION_FILL = PatternFill("solid", fgColor="EAF2FF")
EXPLANATION_FONT = Font(italic=True, color="444444")

# Alternating row fills for the glossary tab (one shade per tab group)
_GLOSS_FILLS = [
    PatternFill("solid", fgColor="C9DEF0"),  # medium blue
    PatternFill("solid", fgColor="FFF2CC"),  # light amber
]

GLOSSARY = [
    # ── Model Overview ────────────────────────────────────────────────────────
    {"Tab": "Model Overview", "Metric": "Edit Rate (%)", "Direction": "Lower = better",
     "Definition": "Pooled: (total records with ≥1 edit / total records reviewed across all reports) × 100. Every record counts equally."},
    {"Tab": "Model Overview", "Metric": "Text % Changed", "Direction": "Lower = better",
     "Definition": "Pooled: total Levenshtein edit distance across all text fields and all records ÷ max(total original, total edited) character volume. Every record counts equally."},
    {"Tab": "Model Overview", "Metric": "Chars Added / Record", "Direction": "Lower = better",
     "Definition": "Pooled: total characters inserted across all text fields ÷ total records reviewed. Every record counts equally."},
    {"Tab": "Model Overview", "Metric": "Chars Deleted / Record", "Direction": "Lower = better",
     "Definition": "Pooled: total characters removed across all text fields ÷ total records reviewed. Every record counts equally."},
    {"Tab": "Model Overview", "Metric": "Edit Effort Ratio", "Direction": "Lower = better",
     "Definition": "Pooled: (total chars added + total chars deleted) ÷ max(total original, total edited) character volume. Measures raw rewrite volume regardless of net change direction. Every record counts equally."},
    {"Tab": "Model Overview", "Metric": "List Item Retention (%)", "Direction": "Higher = better",
     "Definition": "Pooled: (total AI list items kept or corrected by archivist / total AI list items across all records) × 100. Every record counts equally."},
    {"Tab": "Model Overview", "Metric": "Subject Acceptance (%)", "Direction": "Higher = better",
     "Definition": "Pooled: (total AI subject terms approved / total AI subject terms suggested across all records) × 100. Every record counts equally."},
    {"Tab": "Model Overview", "Metric": "Heading Approval (%)", "Direction": "Higher = better",
     "Definition": "Pooled: (total AI LCSH/FAST headings approved / total AI headings suggested across all records) × 100. A term can be right while its full heading form still needs adjustment."},
    {"Tab": "Model Overview", "Metric": "Archivist Addition Rate (%)", "Direction": "Lower = better",
     "Definition": "Pooled: (records where archivist selected ≥1 vocab term not suggested by AI / total records reviewed) × 100. Every record counts equally."},
    {"Tab": "Model Overview", "Metric": "Avg Custom Terms Added", "Direction": "Lower = better",
     "Definition": "Pooled: total vocab terms added from search results / total records reviewed. Every record counts equally."},
    {"Tab": "Model Overview", "Metric": "Avg Custom Subjects Added", "Direction": "Lower = better",
     "Definition": "Pooled: total subject terms written in manually / total records reviewed. Every record counts equally."},
    {"Tab": "Model Overview", "Metric": "Collections", "Direction": "—",
     "Definition": "Count of distinct archival collections in this model's reports."},
    {"Tab": "Model Overview", "Metric": "Evaluators", "Direction": "—",
     "Definition": "Count of distinct evaluators who reviewed this model."},
    # ── Task Breakdown ────────────────────────────────────────────────────────
    {"Tab": "Task Breakdown", "Metric": "Field", "Direction": "—",
     "Definition": "Metadata field name (e.g., title, description, contributors, named_entities)."},
    {"Tab": "Task Breakdown", "Metric": "Field Type", "Direction": "—",
     "Definition": "text = single block of free text; list = discrete items. Each type uses different quality metrics."},
    {"Tab": "Task Breakdown", "Metric": "[Model name columns]", "Direction": "Lower = better",
     "Definition": "Pooled: (total records where this field was edited / total records reviewed for this field) × 100, for that model. Every record counts equally."},
    {"Tab": "Task Breakdown", "Metric": "Avg Edit Rate (%)", "Direction": "Lower = better",
     "Definition": "Pooled edit rate for this field averaged across models. Used to rank fields most-to-least edited."},
    {"Tab": "Task Breakdown", "Metric": "Avg % Changed", "Direction": "Lower = better",
     "Definition": "Text only. Pooled: total Levenshtein edit distance ÷ max(total original, total edited) character volume for this field across all records. Every record counts equally."},
    {"Tab": "Task Breakdown", "Metric": "Avg Sim When Edited (%)", "Direction": "Higher = better",
     "Definition": "Text only. Avg fuzzy similarity between AI output and archivist's final text (0–100%), on edited records only. Report-averaged."},
    {"Tab": "Task Breakdown", "Metric": "Avg Token Sort Ratio", "Direction": "Higher = better",
     "Definition": "Text only. Word-order-insensitive similarity; 100 = same words reordered. Report-averaged on edited records only."},
    {"Tab": "Task Breakdown", "Metric": "Avg Chars Added / Record", "Direction": "Lower = better",
     "Definition": "Text only. Pooled: total characters inserted for this field ÷ total records reviewed. Every record counts equally."},
    {"Tab": "Task Breakdown", "Metric": "Avg Chars Deleted / Record", "Direction": "Lower = better",
     "Definition": "Text only. Pooled: total characters removed for this field ÷ total records reviewed. Every record counts equally."},
    {"Tab": "Task Breakdown", "Metric": "Avg Edit Effort Ratio", "Direction": "Lower = better",
     "Definition": "Text only. Pooled: (total chars added + total chars deleted) ÷ max(total original, total edited) character volume for this field. Every record counts equally."},
    {"Tab": "Task Breakdown", "Metric": "Avg Retention Rate (%)", "Direction": "Higher = better",
     "Definition": "List only. Pooled: (total AI items kept or corrected / total original AI items across all records) × 100. Every record counts equally."},
    {"Tab": "Task Breakdown", "Metric": "Avg Items Removed Rate (%)", "Direction": "Lower = better",
     "Definition": "List only. Pooled: (total items removed / total original AI items across all records) × 100. Every record counts equally."},
    {"Tab": "Task Breakdown", "Metric": "Avg Items Added Rate (%)", "Direction": "Lower = better",
     "Definition": "List only. Pooled: (total items added by archivist / total original AI items across all records) × 100. Every record counts equally."},
    {"Tab": "Task Breakdown", "Metric": "subjects — rejection rate", "Direction": "Lower = better",
     "Definition": "Pooled: total subjects rejected ÷ total subjects suggested across all records. Reads consistently alongside other fields (lower = better)."},
    # ── Evaluator Behavior ────────────────────────────────────────────────────
    {"Tab": "Evaluator Behavior", "Metric": "Records Reviewed", "Direction": "—",
     "Definition": "Total records reviewed by this evaluator across all collections and models."},
    {"Tab": "Evaluator Behavior", "Metric": "Edit Rate (%)", "Direction": "—",
     "Definition": "Pooled: (total records with ≥1 edit / total records reviewed) × 100. Reflects strictness; also influenced by model/collection assignment. Every record counts equally."},
    {"Tab": "Evaluator Behavior", "Metric": "Records Unchanged (%)", "Direction": "Higher = less strict",
     "Definition": "Pooled: (total records with no edits / total records reviewed) × 100. Every record counts equally."},
    {"Tab": "Evaluator Behavior", "Metric": "Text % Changed", "Direction": "Lower = better",
     "Definition": "Pooled: total Levenshtein edit distance ÷ max(total original, total edited) character volume across all this evaluator's records. Distinguishes heavy rewriters from light editors."},
    {"Tab": "Evaluator Behavior", "Metric": "Subject Acceptance (%)", "Direction": "Higher = less strict",
     "Definition": "Pooled: (total AI subject terms approved / total AI subject terms suggested) × 100. Every record counts equally."},
    {"Tab": "Evaluator Behavior", "Metric": "Heading Approval (%)", "Direction": "Higher = less strict",
     "Definition": "Pooled: (total AI headings approved / total AI headings suggested) × 100. Every record counts equally."},
    {"Tab": "Evaluator Behavior", "Metric": "Archivist Addition Rate (%)", "Direction": "—",
     "Definition": "Pooled: (records where ≥1 vocab term added from search results / total records reviewed) × 100. Every record counts equally."},
    {"Tab": "Evaluator Behavior", "Metric": "Avg Custom Terms Added", "Direction": "—",
     "Definition": "Pooled: total vocab terms added from search results / total records reviewed. Every record counts equally."},
    {"Tab": "Evaluator Behavior", "Metric": "Avg Custom Subjects Added", "Direction": "—",
     "Definition": "Pooled: total subject terms written in manually / total records reviewed. Every record counts equally."},
    {"Tab": "Evaluator Behavior", "Metric": "Notes Written (%)", "Direction": "—",
     "Definition": "Pooled: (records where evaluator wrote a note / total records reviewed) × 100. Every record counts equally."},
    # ── Collection Difficulty ─────────────────────────────────────────────────
    {"Tab": "Collection Difficulty", "Metric": "Records Reviewed", "Direction": "—",
     "Definition": "Total records reviewed for this collection across all models and evaluators."},
    {"Tab": "Collection Difficulty", "Metric": "Avg Edit Rate (%)", "Direction": "Lower = better",
     "Definition": "Pooled: (total records with ≥1 edit / total records reviewed) × 100 across all models/evaluators for this collection. Every record counts equally."},
    {"Tab": "Collection Difficulty", "Metric": "Avg Text % Changed", "Direction": "Lower = better",
     "Definition": "Pooled: total Levenshtein edit distance ÷ max(total original, total edited) character volume across all records for this collection. Every record counts equally."},
    {"Tab": "Collection Difficulty", "Metric": "Chars Added / Record", "Direction": "Lower = better",
     "Definition": "Pooled: total characters inserted across all text fields ÷ total records reviewed for this collection. Every record counts equally."},
    {"Tab": "Collection Difficulty", "Metric": "Chars Deleted / Record", "Direction": "Lower = better",
     "Definition": "Pooled: total characters removed across all text fields ÷ total records reviewed for this collection. Every record counts equally."},
    {"Tab": "Collection Difficulty", "Metric": "Subject Acceptance (%)", "Direction": "Higher = better",
     "Definition": "Pooled: (total AI subject terms approved / total AI subject terms suggested) × 100 for this collection. Every record counts equally."},
    {"Tab": "Collection Difficulty", "Metric": "Heading Approval (%)", "Direction": "Higher = better",
     "Definition": "Pooled: (total AI headings approved / total AI headings suggested) × 100 for this collection. Every record counts equally."},
    {"Tab": "Collection Difficulty", "Metric": "Best Model", "Direction": "—",
     "Definition": "Model with the lowest avg edit rate for this collection."},
    {"Tab": "Collection Difficulty", "Metric": "Best Model Edit Rate (%)", "Direction": "Lower = better",
     "Definition": "Edit rate of the best-performing model for this collection. High = collection was hard for all models."},
    # ── Evaluator Notes ───────────────────────────────────────────────────────
    {"Tab": "Evaluator Notes", "Metric": "(no metrics)", "Direction": "—",
     "Definition": "Raw archivist notes organized by Collection × Model with one column per evaluator. Qualitative feedback only."},
]

# ---------------------------------------------------------------------------
# Discovery and parsing
# ---------------------------------------------------------------------------

def _find_evaluations_dir(output_folders_root: str) -> str | None:
    """Search output_folders for any directory containing an evaluations/ subfolder."""
    if not os.path.isdir(output_folders_root):
        return None
    entries = sorted(os.scandir(output_folders_root), key=lambda e: e.name, reverse=True)
    for entry in entries:
        if entry.is_dir():
            candidate = os.path.join(entry.path, "evaluations")
            if os.path.isdir(candidate):
                return candidate
    return None


def _extract_collection(image_folder: str, model: str) -> str:
    """
    Extract collection name from image_folder string.
    e.g. "ArchImagesAI_alfred-zucker_gpt-5.1_2026-03-06_Time_17-34-27", model="gpt-5.1"
         → "alfred-zucker"
    """
    prefix = "ArchImagesAI_"
    if image_folder.startswith(prefix):
        remainder = image_folder[len(prefix):]
        remainder = re.sub(r"_\d{4}-\d{2}-\d{2}_Time_[\d-]+$", "", remainder)
        if model and remainder.endswith("_" + model):
            return remainder[: -(len(model) + 1)]
        return remainder
    return image_folder


def parse_report_file(filepath: str) -> dict | None:
    """Parse one edit_report JSON into a flat ParsedReport dict. Returns None on failure."""
    try:
        with open(filepath, encoding="utf-8") as f:
            raw = json.load(f)
    except Exception as e:
        print(f"  Warning: could not read {os.path.basename(filepath)}: {e}")
        return None

    bi    = raw.get("batch_info", {})
    ev    = bi.get("evaluator") or bi.get("archivist", "Unknown")
    models_used = bi.get("models_used", [])
    model = models_used[0] if models_used else "unknown"
    collection = _extract_collection(bi.get("image_folder", ""), model)

    scores = raw.get("average_quality_scores", {})
    tfs    = raw.get("text_fields_summary", {})
    tft    = tfs.get("text_fields_totals", {})
    subj   = raw.get("subjects", {})
    hdgs   = raw.get("subject_headings", {})
    adds   = hdgs.get("archivist_additions", {})

    def _score(key: str) -> float:
        return scores.get(key, {}).get("value", 0.0)

    # Per-field data
    text_fields: dict[str, dict] = {}
    for fname, fd in tfs.get("text_fields", {}).items():
        text_fields[fname] = {
            "records_edited":   fd.get("records_edited", 0),
            "records_reviewed": fd.get("records_reviewed", 0),
            "pct_changed":        fd.get("pct_changed", 0.0),
            "avg_similarity":     fd.get("avg_edited_field_similarity_ratio", 0.0),
            "avg_token_sort":     fd.get("avg_edited_field_token_sort_ratio", 0.0),
            "chars_added":        fd.get("chars_added", 0),
            "chars_deleted":      fd.get("chars_deleted", 0),
            "edit_effort_ratio":  fd.get("edit_effort_ratio", 0.0),
            "original_length":    fd.get("original_length", 0),
            "new_length":         fd.get("new_length", 0),
            "edit_distance":      fd.get("edit_distance", 0),
        }

    list_fields: dict[str, dict] = {}
    for fname, fd in tfs.get("list_fields", {}).items():
        list_fields[fname] = {
            "records_edited":   fd.get("records_edited", 0),
            "records_reviewed": fd.get("records_reviewed", 0),
            "retention_rate":   fd.get("retention_rate", 100.0),
            "original_count":   fd.get("original_count", 0),
            "items_kept":       fd.get("items_kept", 0),
            "items_corrected":  fd.get("items_corrected", 0),
            "items_removed":    fd.get("items_removed", 0),
            "items_added":      fd.get("items_added", 0),
        }

    # Raw counts for per-record pooled aggregation
    records_reviewed_count = bi.get("records_reviewed", 0)
    records_edited_count   = bi.get("records_edited", records_reviewed_count - bi.get("records_reviewed_only", 0))
    records_unchanged_count = bi.get("records_reviewed_only", 0)

    # Text totals: exact values direct from JSON
    text_edit_distance_total = tft.get("edit_distance", 0)
    text_char_volume         = max(tft.get("original_length", 0), tft.get("new_length", 0))

    # Subject/heading raw counts
    subjects_total    = subj.get("total", 0)
    subjects_accepted = subj.get("accepted", 0)
    subjects_rejected = subj.get("rejected", 0)
    _ai_sel           = hdgs.get("ai_selected", {})
    headings_total    = _ai_sel.get("total", 0)
    headings_approved = _ai_sel.get("approved", 0)

    # List field totals: items_kept + items_corrected = retained
    list_original_total = sum(fd.get("original_count", 0) for fd in list_fields.values())
    list_retained_total = sum(
        fd.get("items_kept", 0) + fd.get("items_corrected", 0)
        for fd in list_fields.values()
    )

    # Per-record stats: vocab additions and notes
    records_with_vocab_adds = 0
    records_with_notes = 0
    for rec in raw.get("records", []):
        if rec.get("archivist_notes", "").strip():
            records_with_notes += 1
        # A vocab addition = approved term with "other-" prefix (archivist added from vocabulary list)
        has_vocab_add = any(
            e.get("edit_type") == "term_decision"
            and str(e.get("term_id", "")).startswith("other-")
            and "approved" in str(e.get("new_value", "")).lower()
            for e in rec.get("edits", [])
        )
        if has_vocab_add:
            records_with_vocab_adds += 1


    return {
        "filepath":                   filepath,
        "evaluator":                  ev,
        "collection":                 collection,
        "model":                      model,
        # Counts (exact — used for per-record pooled aggregation)
        "records_reviewed":           records_reviewed_count,
        "records_edited_count":       records_edited_count,
        "records_unchanged_count":    records_unchanged_count,
        "records_with_vocab_adds_count": records_with_vocab_adds,
        "records_with_notes_count":   records_with_notes,
        "text_edit_distance_total":   text_edit_distance_total,
        "text_char_volume":           text_char_volume,
        "text_chars_added_total":     tft.get("chars_added", 0),
        "text_chars_deleted_total":   tft.get("chars_deleted", 0),
        "subjects_total":             subjects_total,
        "subjects_accepted":          subjects_accepted,
        "subjects_rejected":          subjects_rejected,
        "headings_total":             headings_total,
        "headings_approved":          headings_approved,
        "list_original_total":        list_original_total,
        "list_retained_total":        list_retained_total,
        "custom_terms_total":         adds.get("custom_terms", 0),
        "custom_subjects_total":      subj.get("custom_added", 0),
        # Legacy pre-computed rates (kept for reference; not used in aggregation)
        "format_media_approval_rate": _score("format_media_approval_rate"),
        "text_fields":                text_fields,
        "list_fields":                list_fields,
        "records":                    raw.get("records", []),
    }


def discover_all_reports(evaluations_dir: str) -> list[dict]:
    """Walk evaluations_dir/{EV}/{EV}_changes/edit_report_*.json and parse all."""
    reports = []
    if not os.path.isdir(evaluations_dir):
        print(f"Error: evaluations directory not found: {evaluations_dir}")
        return reports

    for entry in sorted(os.scandir(evaluations_dir), key=lambda e: e.name):
        if not entry.is_dir() or entry.name.startswith("."):
            continue
        for sub in sorted(os.scandir(entry.path), key=lambda e: e.name):
            if not sub.is_dir() or not sub.name.endswith("_changes"):
                continue
            for fname in sorted(os.listdir(sub.path)):
                if not fname.startswith("edit_report_") or not fname.endswith(".json"):
                    continue
                report = parse_report_file(os.path.join(sub.path, fname))
                if report:
                    reports.append(report)

    return reports

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _pool_rate(reports: list[dict], num_key: str, den_key: str, decimals: int = 1) -> float | None:
    """Pool numerator and denominator across reports, return rate as a percentage."""
    total_num = sum(r.get(num_key, 0) for r in reports)
    total_den = sum(r.get(den_key, 0) for r in reports)
    if total_den == 0:
        return None
    return round(total_num / total_den * 100, decimals)


def _pool_ratio(reports: list[dict], num_key: str, den_key: str, decimals: int = 4) -> float | None:
    """Pool numerator and denominator across reports, return ratio (not ×100)."""
    total_num = sum(r.get(num_key, 0) for r in reports)
    total_den = sum(r.get(den_key, 0) for r in reports)
    if total_den == 0:
        return None
    return round(total_num / total_den, decimals)

# ---------------------------------------------------------------------------
# Computation functions
# ---------------------------------------------------------------------------

def compute_model_overview(reports: list[dict]) -> pd.DataFrame:
    """Tab 1: One row per model, sorted by edit rate ascending (best first)."""
    by_model: dict[str, list] = defaultdict(list)
    for r in reports:
        by_model[r["model"]].append(r)

    rows = []
    for model, reps in sorted(by_model.items()):
        _total_records  = sum(r["records_reviewed"]         for r in reps)
        _total_added    = sum(r["text_chars_added_total"]   for r in reps)
        _total_deleted  = sum(r["text_chars_deleted_total"] for r in reps)
        _total_effort   = _total_added + _total_deleted
        _total_vol      = sum(r["text_char_volume"]         for r in reps)
        _custom_terms   = sum(r["custom_terms_total"]       for r in reps)
        _custom_subj    = sum(r["custom_subjects_total"]    for r in reps)
        rows.append({
            "Model":                       model,
            "Edit Rate (%)":               _pool_rate(reps, "records_edited_count",       "records_reviewed"),
            "Text % Changed":              round(_pool_ratio(reps, "text_edit_distance_total", "text_char_volume") * 100, 1)
                                           if _total_vol > 0 else None,
            "Chars Added / Record":        round(_total_added   / _total_records, 2) if _total_records > 0 else None,
            "Chars Deleted / Record":      round(_total_deleted / _total_records, 2) if _total_records > 0 else None,
            "Edit Effort Ratio":           round(_total_effort / _total_vol, 4) if _total_vol > 0 else None,
            "List Item Retention (%)":     _pool_rate(reps, "list_retained_total",         "list_original_total"),
            "Subject Acceptance (%)":      _pool_rate(reps, "subjects_accepted",            "subjects_total"),
            "Heading Approval (%)":        _pool_rate(reps, "headings_approved",            "headings_total"),
            "Archivist Addition Rate (%)": _pool_rate(reps, "records_with_vocab_adds_count","records_reviewed"),
            "Avg Custom Terms Added":      round(_custom_terms / _total_records, 2) if _total_records > 0 else None,
            "Avg Custom Subjects Added":   round(_custom_subj  / _total_records, 2) if _total_records > 0 else None,
            "Collections":                 len({r["collection"] for r in reps}),
            "Evaluators":                  len({r["evaluator"] for r in reps}),
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("Edit Rate (%)", ascending=True, na_position="last").reset_index(drop=True)
    return df


def compute_task_breakdown(reports: list[dict]) -> pd.DataFrame:
    """
    Tab 2: One row per field. Per-model edit rates + aggregate quality columns.
    Text fields: Avg % Changed, Avg Similarity When Edited, Avg Token Sort Ratio.
    List fields: Avg Retention Rate, Avg Items Removed Rate, Avg Items Added Rate.
    'subjects' row shows rejection rate (100 - acceptance_rate) for consistency.
    Sorted by avg edit rate descending (most-edited fields first).
    """
    models = sorted({r["model"] for r in reports})

    rows = []

    def _field_text_fds(reps, fname):
        """Yield (report, field_dict) for reports that have this text field with records_reviewed > 0."""
        for r in reps:
            fd = r["text_fields"].get(fname, {})
            if fd.get("records_reviewed", 0) > 0:
                yield r, fd

    def _field_list_fds(reps, fname):
        """Yield (report, field_dict) for reports that have this list field with records_reviewed > 0."""
        for r in reps:
            fd = r["list_fields"].get(fname, {})
            if fd.get("records_reviewed", 0) > 0:
                yield r, fd

    for fname in TEXT_FIELDS:
        row: dict = {"Field": fname, "Field Type": "text"}
        model_rates = []
        for model in models:
            mreps = [r for r in reports if r["model"] == model]
            tot_edited   = sum(fd["records_edited"]   for _, fd in _field_text_fds(mreps, fname))
            tot_reviewed = sum(fd["records_reviewed"] for _, fd in _field_text_fds(mreps, fname))
            if tot_reviewed > 0:
                rate = round(tot_edited / tot_reviewed * 100, 1)
                row[model] = rate
                model_rates.append(tot_edited / tot_reviewed * 100)
            else:
                row[model] = None

        # Aggregate text quality metrics — pooled across all records
        tot_edit_dist  = sum(fd["edit_distance"]   for _, fd in _field_text_fds(reports, fname))
        tot_orig_len   = sum(fd["original_length"] for _, fd in _field_text_fds(reports, fname))
        tot_new_len    = sum(fd["new_length"]       for _, fd in _field_text_fds(reports, fname))
        tot_char_vol   = max(tot_orig_len, tot_new_len)
        tot_reviewed   = sum(fd["records_reviewed"] for _, fd in _field_text_fds(reports, fname))
        tot_ca         = sum(fd["chars_added"]      for _, fd in _field_text_fds(reports, fname))
        tot_cd         = sum(fd["chars_deleted"]    for _, fd in _field_text_fds(reports, fname))

        # Avg similarity/token-sort: still report-averaged (ratio metrics with no natural count denominator)
        all_sim = [
            fd["avg_similarity"]
            for _, fd in _field_text_fds(reports, fname)
            if fd.get("records_edited", 0) > 0 and fd.get("avg_similarity", 0) > 0
        ]
        all_tsort = [
            fd["avg_token_sort"]
            for _, fd in _field_text_fds(reports, fname)
            if fd.get("records_edited", 0) > 0 and fd.get("avg_token_sort", 0) > 0
        ]

        row["Avg Edit Rate (%)"]          = round(statistics.mean(model_rates), 1) if model_rates else None
        row["Avg % Changed"]              = round(tot_edit_dist / tot_char_vol * 100, 1) if tot_char_vol > 0 else None
        row["Avg Sim When Edited (%)"]    = round(statistics.mean(all_sim), 1) if all_sim else None
        row["Avg Token Sort Ratio"]       = round(statistics.mean(all_tsort), 1) if all_tsort else None
        row["Avg Chars Added / Record"]   = round(tot_ca / tot_reviewed, 2) if tot_reviewed > 0 else None
        row["Avg Chars Deleted / Record"] = round(tot_cd / tot_reviewed, 2) if tot_reviewed > 0 else None
        row["Avg Edit Effort Ratio"]      = round((tot_ca + tot_cd) / tot_char_vol, 4) if tot_char_vol > 0 else None
        row["Avg Retention Rate (%)"]     = None
        row["Avg Items Removed Rate (%)"] = None
        row["Avg Items Added Rate (%)"]   = None
        rows.append(row)

    for fname in LIST_FIELDS:
        row = {"Field": fname, "Field Type": "list"}
        model_rates = []
        for model in models:
            mreps = [r for r in reports if r["model"] == model]
            tot_edited   = sum(fd["records_edited"]   for _, fd in _field_list_fds(mreps, fname))
            tot_reviewed = sum(fd["records_reviewed"] for _, fd in _field_list_fds(mreps, fname))
            if tot_reviewed > 0:
                rate = round(tot_edited / tot_reviewed * 100, 1)
                row[model] = rate
                model_rates.append(tot_edited / tot_reviewed * 100)
            else:
                row[model] = None

        # Pooled list quality metrics
        tot_orig     = sum(fd["original_count"]               for _, fd in _field_list_fds(reports, fname))
        tot_retained = sum(fd["items_kept"] + fd["items_corrected"] for _, fd in _field_list_fds(reports, fname))
        tot_removed  = sum(fd["items_removed"]                for _, fd in _field_list_fds(reports, fname))
        tot_added    = sum(fd["items_added"]                  for _, fd in _field_list_fds(reports, fname))

        row["Avg Edit Rate (%)"]          = round(statistics.mean(model_rates), 1) if model_rates else None
        row["Avg % Changed"]              = None
        row["Avg Sim When Edited (%)"]    = None
        row["Avg Token Sort Ratio"]       = None
        row["Avg Chars Added / Record"]   = None
        row["Avg Chars Deleted / Record"] = None
        row["Avg Edit Effort Ratio"]      = None
        row["Avg Retention Rate (%)"]     = round(tot_retained / tot_orig * 100, 1) if tot_orig > 0 else None
        row["Avg Items Removed Rate (%)"] = round(tot_removed  / tot_orig * 100, 1) if tot_orig > 0 else None
        row["Avg Items Added Rate (%)"]   = round(tot_added    / tot_orig * 100, 1) if tot_orig > 0 else None
        rows.append(row)

    # Subjects: rejection rate = 100 - acceptance_rate (lower = better, consistent with other fields)
    subj_row: dict = {"Field": "subjects", "Field Type": "list — rejection rate"}
    model_rates = []
    for model in models:
        mreps   = [r for r in reports if r["model"] == model]
        tot_rej = sum(r["subjects_rejected"] for r in mreps)
        tot_tot = sum(r["subjects_total"]    for r in mreps)
        if tot_tot > 0:
            rate = round(tot_rej / tot_tot * 100, 1)
            subj_row[model] = rate
            model_rates.append(rate)
        else:
            subj_row[model] = None
    _all_rej = sum(r["subjects_rejected"] for r in reports)
    _all_tot = sum(r["subjects_total"]    for r in reports)
    subj_row["Avg Edit Rate (%)"]          = round(_all_rej / _all_tot * 100, 1) if _all_tot > 0 else None
    subj_row["Avg % Changed"]              = None
    subj_row["Avg Sim When Edited (%)"]    = None
    subj_row["Avg Token Sort Ratio"]       = None
    subj_row["Avg Chars Added / Record"]   = None
    subj_row["Avg Chars Deleted / Record"] = None
    subj_row["Avg Edit Effort Ratio"]      = None
    subj_row["Avg Retention Rate (%)"]     = None
    subj_row["Avg Items Removed Rate (%)"] = None
    subj_row["Avg Items Added Rate (%)"]   = None
    rows.append(subj_row)

    df = pd.DataFrame(rows)
    model_cols = [m for m in models if m in df.columns]
    col_order = (
        ["Field", "Field Type"] + model_cols +
        ["Avg Edit Rate (%)", "Avg % Changed", "Avg Sim When Edited (%)",
         "Avg Token Sort Ratio", "Avg Chars Added / Record", "Avg Chars Deleted / Record", "Avg Edit Effort Ratio",
         "Avg Retention Rate (%)",
         "Avg Items Removed Rate (%)", "Avg Items Added Rate (%)"]
    )
    df = df[[c for c in col_order if c in df.columns]]

    if not df.empty:
        df = df.sort_values("Avg Edit Rate (%)", ascending=False, na_position="last").reset_index(drop=True)
    return df


def compute_evaluator_behavior(reports: list[dict]) -> pd.DataFrame:
    """Tab 3: One row per evaluator."""
    by_ev: dict[str, list] = defaultdict(list)
    for r in reports:
        by_ev[r["evaluator"]].append(r)

    rows = []
    for ev, reps in sorted(by_ev.items()):
        _total_records = sum(r["records_reviewed"]       for r in reps)
        _custom_terms  = sum(r["custom_terms_total"]     for r in reps)
        _custom_subj   = sum(r["custom_subjects_total"]  for r in reps)
        _total_vol     = sum(r["text_char_volume"]        for r in reps)
        _total_dist    = sum(r["text_edit_distance_total"] for r in reps)
        rows.append({
            "Evaluator":                   ev,
            "Records Reviewed":            _total_records,
            "Edit Rate (%)":               _pool_rate(reps, "records_edited_count",        "records_reviewed"),
            "Records Unchanged (%)":       _pool_rate(reps, "records_unchanged_count",     "records_reviewed"),
            "Text % Changed":              round(_total_dist / _total_vol * 100, 1) if _total_vol > 0 else None,
            "Subject Acceptance (%)":      _pool_rate(reps, "subjects_accepted",            "subjects_total"),
            "Heading Approval (%)":        _pool_rate(reps, "headings_approved",            "headings_total"),
            "Archivist Addition Rate (%)": _pool_rate(reps, "records_with_vocab_adds_count","records_reviewed"),
            "Avg Custom Terms Added":      round(_custom_terms / _total_records, 2) if _total_records > 0 else None,
            "Avg Custom Subjects Added":   round(_custom_subj  / _total_records, 2) if _total_records > 0 else None,
            "Notes Written (%)":           _pool_rate(reps, "records_with_notes_count",    "records_reviewed"),
        })

    return pd.DataFrame(rows)


def compute_collection_difficulty(reports: list[dict]) -> pd.DataFrame:
    """Tab 4: One row per collection, sorted by avg edit rate descending (hardest first)."""
    collections = sorted({r["collection"] for r in reports})

    rows = []
    for coll in collections:
        creps = [r for r in reports if r["collection"] == coll]

        # Per-model pooled edit rate to find the best model
        model_edit_rates: dict[str, float] = {}
        for model in {r["model"] for r in creps}:
            mreps    = [r for r in creps if r["model"] == model]
            tot_ed   = sum(r["records_edited_count"] for r in mreps)
            tot_rev  = sum(r["records_reviewed"]     for r in mreps)
            if tot_rev > 0:
                model_edit_rates[model] = tot_ed / tot_rev * 100

        best_model = min(model_edit_rates, key=model_edit_rates.get) if model_edit_rates else None

        _total_records = sum(r["records_reviewed"]           for r in creps)
        _total_added   = sum(r["text_chars_added_total"]     for r in creps)
        _total_deleted = sum(r["text_chars_deleted_total"]   for r in creps)
        _total_dist    = sum(r["text_edit_distance_total"]   for r in creps)
        _total_vol     = sum(r["text_char_volume"]           for r in creps)

        rows.append({
            "Collection":                coll,
            "Records Reviewed":          _total_records,
            "Avg Edit Rate (%)":         _pool_rate(creps, "records_edited_count",  "records_reviewed"),
            "Avg Text % Changed":        round(_total_dist / _total_vol * 100, 1) if _total_vol > 0 else None,
            "Chars Added / Record":      round(_total_added   / _total_records, 2) if _total_records > 0 else None,
            "Chars Deleted / Record":    round(_total_deleted / _total_records, 2) if _total_records > 0 else None,
            "Subject Acceptance (%)":    _pool_rate(creps, "subjects_accepted",     "subjects_total"),
            "Heading Approval (%)":      _pool_rate(creps, "headings_approved",     "headings_total"),
            "Best Model":                best_model,
            "Best Model Edit Rate (%)":  round(model_edit_rates[best_model], 1) if best_model else None,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("Avg Edit Rate (%)", ascending=False, na_position="last").reset_index(drop=True)
    return df


def compile_evaluator_notes(reports: list[dict]) -> pd.DataFrame:
    """
    Tab 5: One row per (Collection, Model). One column per evaluator containing
    all their notes for that instance, formatted as 'DrawingID: note' per line.
    """
    evaluators = sorted({r["evaluator"] for r in reports})
    combos     = sorted({(r["collection"], r["model"]) for r in reports})

    # Build: (collection, model, evaluator) → ["DrawingID: note", ...]
    note_map: dict[tuple, list[str]] = defaultdict(list)
    for rep in reports:
        key = (rep["collection"], rep["model"], rep["evaluator"])
        for rec in rep["records"]:
            note = rec.get("archivist_notes", "").strip()
            if note:
                drawing_id = rec.get("id", "")
                note_map[key].append(f"{drawing_id}: {note}" if drawing_id else note)

    rows = []
    for coll, model in combos:
        row: dict = {"Collection": coll, "Model": model}
        has_any = False
        for ev in evaluators:
            notes = note_map.get((coll, model, ev), [])
            row[ev] = "\n".join(notes) if notes else None
            if notes:
                has_any = True
        if has_any:
            rows.append(row)

    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["Collection", "Model"] + evaluators)

# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

class ExcelWriter:
    def __init__(self, output_path: str):
        self.wb = Workbook()
        self.output_path = output_path
        if self.wb.active:
            self.wb.remove(self.wb.active)

    def _write_title(self, ws, title: str, n_cols: int) -> None:
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = TITLE_FONT
        if n_cols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    def _write_df(
        self,
        ws,
        df: pd.DataFrame,
        header_row: int = 2,
        freeze_col: int = 1,
        explanation: dict | None = None,
    ) -> None:
        cols = list(df.columns)
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

        data_start = header_row + 1
        if explanation is not None:
            for col_idx, col_name in enumerate(cols, 1):
                cell = ws.cell(row=header_row + 1, column=col_idx, value=explanation.get(col_name, ""))
                cell.fill = EXPLANATION_FILL
                cell.font = EXPLANATION_FONT
                cell.alignment = TOP_WRAP
            data_start = header_row + 2

        for row_idx, row_data in enumerate(df.itertuples(index=False), data_start):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(
                    row=row_idx, column=col_idx,
                    value=None if pd.isna(value) else value,
                )
                cell.alignment = TOP_WRAP
        ws.freeze_panes = ws.cell(row=data_start, column=freeze_col)

    def _auto_width(self, ws, min_w: int = 10, max_w: int = 40) -> None:
        for col in ws.columns:
            best = min_w
            for cell in col:
                if cell.value:
                    best = min(max(best, len(str(cell.value)) + 2), max_w)
            ws.column_dimensions[get_column_letter(col[0].column)].width = best

    def write_model_overview(self, df: pd.DataFrame) -> None:
        ws = self.wb.create_sheet("Model Overview")
        self._write_title(ws, "Model Overview — Averaged Across All Evaluators and Collections", len(df.columns))
        explanation = {
            "Model":                       "",
            "Edit Rate (%)":               "pooled: (total records with ≥1 edit / total records reviewed) × 100 — every record counts equally",
            "Text % Changed":              "pooled: total Levenshtein edit distance ÷ max(total original, total edited) character volume across all records",
            "Chars Added / Record":        "pooled: total chars inserted ÷ total records reviewed — every record counts equally",
            "Chars Deleted / Record":      "pooled: total chars removed ÷ total records reviewed — every record counts equally",
            "Edit Effort Ratio":           "pooled: (total chars added + total chars deleted) ÷ max(total original, total edited) character volume — every record counts equally",
            "List Item Retention (%)":     "pooled: (total items kept or corrected / total original AI items) × 100 — every record counts equally",
            "Subject Acceptance (%)":      "pooled: (total subjects approved / total subjects suggested) × 100 — every record counts equally",
            "Heading Approval (%)":        "pooled: (total headings approved / total headings suggested) × 100 — every record counts equally",
            "Archivist Addition Rate (%)": "pooled: (records where ≥1 term added from search results / total records reviewed) × 100",
            "Avg Custom Terms Added":      "pooled: total vocabulary terms added from search results / total records reviewed",
            "Avg Custom Subjects Added":   "pooled: total subject terms written in manually / total records reviewed",
            "Collections":                 "count of distinct collections covered by this model's reports",
            "Evaluators":                  "count of distinct evaluators who reviewed this model",
        }
        self._write_df(ws, df, freeze_col=2, explanation=explanation)
        self._auto_width(ws, min_w=12, max_w=40)

    def write_task_breakdown(self, df: pd.DataFrame) -> None:
        ws = self.wb.create_sheet("Task Breakdown")
        self._write_title(ws, "Task Breakdown — Edit Rate (%) by Field and Model  |  'subjects' row shows rejection rate", len(df.columns))
        explanation: dict = {
            "Field":                       "",
            "Field Type":                  "",
            "Avg Edit Rate (%)":           "pooled edit rate for this field averaged across models — every record counts equally",
            "Avg % Changed":               "text only: pooled total edit distance ÷ max(total original, total edited) character volume — every record counts equally",
            "Avg Sim When Edited (%)":     "text only: report-averaged fuzzy similarity on edited records (0–100%)",
            "Avg Token Sort Ratio":        "text only: report-averaged word-order-insensitive similarity on edited records",
            "Avg Chars Added / Record":    "text only: pooled chars inserted ÷ total records reviewed for this field — every record counts equally",
            "Avg Chars Deleted / Record":  "text only: pooled chars removed ÷ total records reviewed for this field — every record counts equally",
            "Avg Edit Effort Ratio":       "text only: pooled (chars added + chars deleted) ÷ max(total original, total edited) character volume — every record counts equally",
            "Avg Retention Rate (%)":      "list only: pooled (items kept + corrected) / total original AI items — every record counts equally",
            "Avg Items Removed Rate (%)":  "list only: pooled items removed / total original AI items — every record counts equally",
            "Avg Items Added Rate (%)":    "list only: pooled items added / total original AI items — every record counts equally",
        }
        # Model name columns all share the same formula
        for col in df.columns:
            if col not in explanation:
                explanation[col] = "pooled: (total records where this field was edited / total records reviewed for this field) × 100"
        self._write_df(ws, df, freeze_col=3, explanation=explanation)
        self._auto_width(ws, min_w=12, max_w=35)

    def write_evaluator_behavior(self, df: pd.DataFrame) -> None:
        ws = self.wb.create_sheet("Evaluator Behavior")
        self._write_title(ws, "Evaluator Behavior — Edit Rate reflects overall strictness", len(df.columns))
        explanation = {
            "Evaluator":                   "",
            "Records Reviewed":            "total records reviewed across all collections and models",
            "Edit Rate (%)":               "pooled: (total records with ≥1 edit / total records reviewed) × 100 — every record counts equally",
            "Records Unchanged (%)":       "pooled: (total records with no edits / total records reviewed) × 100 — every record counts equally",
            "Text % Changed":              "pooled: total edit distance ÷ max(total original, total edited) character volume — every record counts equally",
            "Subject Acceptance (%)":      "pooled: (total subjects approved / total subjects suggested) × 100 — every record counts equally",
            "Heading Approval (%)":        "pooled: (total headings approved / total headings suggested) × 100 — every record counts equally",
            "Archivist Addition Rate (%)": "pooled: (records where ≥1 vocab term added / total records reviewed) × 100 — every record counts equally",
            "Avg Custom Terms Added":      "pooled: total vocab terms added from search results / total records reviewed",
            "Avg Custom Subjects Added":   "pooled: total subject terms written in manually / total records reviewed",
            "Notes Written (%)":           "pooled: (records where evaluator wrote a note / total records reviewed) × 100",
        }
        self._write_df(ws, df, freeze_col=2, explanation=explanation)
        self._auto_width(ws, min_w=12, max_w=35)

    def write_collection_difficulty(self, df: pd.DataFrame) -> None:
        ws = self.wb.create_sheet("Collection Difficulty")
        self._write_title(ws, "Collection Difficulty", len(df.columns))
        explanation = {
            "Collection":                "collection name",
            "Records Reviewed":          "total records reviewed for this collection across all models and evaluators",
            "Avg Edit Rate (%)":         "pooled: (total records with ≥1 edit / total records reviewed) × 100 — every record counts equally",
            "Avg Text % Changed":        "pooled: total edit distance ÷ max(total original, total edited) character volume — every record counts equally",
            "Chars Added / Record":      "pooled: total chars inserted ÷ total records reviewed — every record counts equally",
            "Chars Deleted / Record":    "pooled: total chars removed ÷ total records reviewed — every record counts equally",
            "Subject Acceptance (%)":    "pooled: (total subjects approved / total subjects suggested) × 100 — every record counts equally",
            "Heading Approval (%)":      "pooled: (total headings approved / total headings suggested) × 100 — every record counts equally",
            "Best Model":                "model with the lowest pooled edit rate for this collection",
            "Best Model Edit Rate (%)":  "pooled edit rate of the best-performing model for this collection",
        }
        self._write_df(ws, df, freeze_col=2, explanation=explanation)
        self._auto_width(ws, min_w=12, max_w=35)

    def write_evaluator_notes(self, df: pd.DataFrame) -> None:
        ws = self.wb.create_sheet("Evaluator Notes")
        n_cols = len(df.columns) if not df.empty else 2
        self._write_title(ws, "Evaluator Notes — One row per Collection × Model; one column per evaluator", n_cols)
        if not df.empty:
            explanation: dict = {"Collection": "", "Model": ""}
            for col in df.columns:
                if col not in explanation:
                    explanation[col] = "raw notes written by this evaluator during review (one entry per drawing)"
            self._write_df(ws, df, freeze_col=3, explanation=explanation)
            self._auto_width(ws, min_w=12, max_w=25)
            ws.column_dimensions["B"].width = 50
            # Widen evaluator note columns (everything after Collection and Model)
            for col_idx in range(3, len(df.columns) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 55
        else:
            ws.cell(row=2, column=1, value="No notes found.")

    def write_glossary(self) -> None:
        ws = self.wb.create_sheet("Metric Glossary")
        self._write_title(
            ws,
            "Metric Glossary — Plain-language definitions for every metric, in tab order",
            4,
        )
        headers = ["Tab", "Metric", "Direction", "Definition"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

        tab_order: list[str] = []
        for row_data in GLOSSARY:
            if row_data["Tab"] not in tab_order:
                tab_order.append(row_data["Tab"])

        for row_idx, row_data in enumerate(GLOSSARY, 3):
            fill = _GLOSS_FILLS[tab_order.index(row_data["Tab"]) % 2]
            for col_idx, key in enumerate(["Tab", "Metric", "Direction", "Definition"], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
                cell.alignment = TOP_WRAP
                cell.fill = fill

        ws.freeze_panes = ws.cell(row=3, column=1)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 150

    def save(self) -> None:
        self.wb.save(self.output_path)
        print(f"Saved: {self.output_path}")

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Compile and analyze evaluator edit reports across collections and models.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument(
        "--evaluations-dir",
        default=None,
        help="Path to the evaluations/ directory. Auto-detected if not specified.",
    )
    p.add_argument(
        "--output-dir",
        default=None,
        help="Directory to write the output Excel file. Defaults to evaluations-dir.",
    )
    return p


def main() -> int:
    args = build_argparser().parse_args()

    evaluations_dir = args.evaluations_dir
    if not evaluations_dir:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_folders = os.path.join(script_dir, "output_folders")
        evaluations_dir = _find_evaluations_dir(output_folders)
        if not evaluations_dir:
            print("Error: Could not auto-detect evaluations/ directory.")
            print("Use --evaluations-dir to specify its location.")
            return 1

    print("Edit Report Analysis")
    print("=" * 60)
    print(f"Evaluations directory: {evaluations_dir}")

    reports = discover_all_reports(evaluations_dir)
    if not reports:
        print("No edit_report_*.json files found. Check the directory structure.")
        return 1

    evaluators  = sorted({r["evaluator"]  for r in reports})
    collections = sorted({r["collection"] for r in reports})
    models      = sorted({r["model"]      for r in reports})

    print(
        f"Loaded {len(reports)} report(s) from "
        f"{len(evaluators)} evaluator(s), "
        f"{len(collections)} collection(s), "
        f"{len(models)} model(s)."
    )
    print(f"  Evaluators:  {', '.join(evaluators)}")
    print(f"  Collections: {', '.join(collections)}")
    print(f"  Models:      {', '.join(models)}")
    print()

    print("Computing analyses...")
    df_model_overview      = compute_model_overview(reports)
    df_task_breakdown      = compute_task_breakdown(reports)
    df_evaluator_behavior  = compute_evaluator_behavior(reports)
    df_collection_diff     = compute_collection_difficulty(reports)
    df_notes               = compile_evaluator_notes(reports)

    output_dir = args.output_dir or evaluations_dir
    os.makedirs(output_dir, exist_ok=True)
    today = date.today().strftime("%Y-%m-%d")
    output_path = os.path.join(output_dir, f"edit_report_analysis_{today}.xlsx")

    print("Writing Excel workbook...")
    writer = ExcelWriter(output_path)
    writer.write_model_overview(df_model_overview)
    writer.write_task_breakdown(df_task_breakdown)
    writer.write_evaluator_behavior(df_evaluator_behavior)
    writer.write_collection_difficulty(df_collection_diff)
    writer.write_evaluator_notes(df_notes)
    writer.write_glossary()
    writer.save()

    print("\nDone.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
