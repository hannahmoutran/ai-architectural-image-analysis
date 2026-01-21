#!/usr/bin/env python3
"""
Step 6: Integrate Archivist Edits into Workflow Outputs
========================================================

Processes archivist review decisions from exported JSON and updates all deliverable files.

This script:
1. Reads archivist decisions from JSON exports (created by Step 5 HTML review interface)
2. Backs up original files to an 'original-outputs' folder
3. Applies edits to the workflow JSON (drawings_workflow.json)
4. Updates the Excel deliverable (drawings_workflow.xlsx)
5. Adds an 'Edit History' sheet tracking all changes with statistics

Usage:
    python step-6-integrate-archivist-edits.py
    python run.py 6
"""

import os
import sys
import json
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# Add CODE directory to path for imports
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)

from shared_utilities import find_newest_folder


class ArchivistEditsIntegrator:
    """Integrates archivist edits from the HTML review interface back into workflow files."""

    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.folder_name = os.path.basename(folder_path)
        self.metadata_folder = os.path.join(folder_path, "metadata", "collection_metadata")
        self.review_folder = os.path.join(folder_path, "review")
        self.exports_folder = os.path.join(self.review_folder, "exports")
        self.original_outputs_folder = os.path.join(folder_path, "original-outputs")

        self.workflow_json_path = os.path.join(self.metadata_folder, "drawings_workflow.json")
        self.workflow_excel_path = os.path.join(self.metadata_folder, "drawings_workflow.xlsx")

        self.workflow_data = None
        self.decisions_data = None
        self.edit_history = []
        self.stats = {
            'total_records_in_export': 0,
            'records_with_edits': 0,
            'records_reviewed_only': 0,
            'total_field_edits': 0,
            'total_term_decisions': 0,
            'terms_approved': 0,
            'terms_rejected': 0,
            'custom_terms_added': 0,
            'custom_subjects_added': 0,
            'edits_by_field': {},
            'archivist_name': '',
            'export_timestamp': '',
            'integration_timestamp': ''
        }

    def find_latest_export(self):
        """Find the most recent archivist decisions export JSON."""
        if not os.path.exists(self.exports_folder):
            print(f"Error: Exports folder not found at {self.exports_folder}")
            print("Please run Step 5 and export decisions first.")
            return None

        json_files = [f for f in os.listdir(self.exports_folder) if f.endswith('.json')]

        if not json_files:
            print(f"Error: No JSON export files found in {self.exports_folder}")
            print("Please export your review decisions from the HTML interface first.")
            return None

        # Sort by modification time, newest first
        json_files.sort(key=lambda f: os.path.getmtime(os.path.join(self.exports_folder, f)), reverse=True)

        latest_file = json_files[0]
        return os.path.join(self.exports_folder, latest_file)

    def load_decisions(self, decisions_path):
        """Load archivist decisions from JSON export."""
        try:
            with open(decisions_path, 'r', encoding='utf-8') as f:
                self.decisions_data = json.load(f)

            # Support both old 'cataloger_name' and new 'archivist_name' keys
            self.stats['archivist_name'] = self.decisions_data.get('archivist_name',
                                            self.decisions_data.get('cataloger_name', 'Unknown'))
            self.stats['export_timestamp'] = self.decisions_data.get('export_timestamp', '')
            self.stats['total_records_in_export'] = len(self.decisions_data.get('decisions', []))

            print(f"Loaded {self.stats['total_records_in_export']} record decisions")
            print(f"Archivist: {self.stats['archivist_name']}")
            print(f"Export timestamp: {self.stats['export_timestamp']}")

            return True
        except Exception as e:
            print(f"Error loading decisions: {e}")
            return False

    def load_workflow_json(self):
        """Load the workflow JSON data."""
        if not os.path.exists(self.workflow_json_path):
            print(f"Error: Workflow JSON not found at {self.workflow_json_path}")
            return False

        try:
            with open(self.workflow_json_path, 'r', encoding='utf-8') as f:
                self.workflow_data = json.load(f)

            # Handle api_stats entry at the end
            if self.workflow_data and isinstance(self.workflow_data[-1], dict) and 'api_stats' in self.workflow_data[-1]:
                self.api_stats = self.workflow_data[-1]
                self.workflow_data = self.workflow_data[:-1]
            else:
                self.api_stats = None

            print(f"Loaded workflow JSON with {len(self.workflow_data)} records")
            return True
        except Exception as e:
            print(f"Error loading workflow JSON: {e}")
            return False

    def backup_original_files(self):
        """Backup original files to original-outputs folder."""
        os.makedirs(self.original_outputs_folder, exist_ok=True)

        files_to_backup = [
            (self.workflow_json_path, "drawings_workflow.json"),
            (self.workflow_excel_path, "drawings_workflow.xlsx")
        ]

        backed_up = []
        for src_path, filename in files_to_backup:
            if os.path.exists(src_path):
                dest_path = os.path.join(self.original_outputs_folder, filename)

                # Only backup if not already backed up (preserve truly original files)
                if not os.path.exists(dest_path):
                    shutil.copy2(src_path, dest_path)
                    backed_up.append(filename)
                    print(f"   Backed up: {filename}")
                else:
                    print(f"   Backup already exists: {filename} (preserving original)")

        return backed_up

    def apply_field_edit(self, record, field_name, edit_data, record_id):
        """Apply a single field edit to a record."""
        analysis = record.get('analysis', {})
        original_value = edit_data.get('original', '')
        new_value = edit_data.get('value', '')

        # Map field names between export and workflow JSON
        field_mapping = {
            'title': 'title',
            'genre': 'genre',
            'description': 'description',
            'ocr_text': 'ocr_text',
            'format_media': 'format_media',
            'date_on_drawing': 'date_on_drawing',
            'sheet_info': 'sheet_info',
            'content_warning': 'content_warning',
            'contributors': 'contributors',
            'named_entities': 'named_entities',
            'geographic_entities': 'geographic_entities',
            'subjects': 'subjects'
        }

        json_field = field_mapping.get(field_name, field_name)

        if json_field in analysis or json_field in field_mapping.values():
            analysis[json_field] = new_value
            record['analysis'] = analysis

            # Track the edit
            self.edit_history.append({
                'record_id': record_id,
                'field': field_name,
                'original_value': self._truncate_for_display(original_value),
                'new_value': self._truncate_for_display(new_value),
                'edit_type': 'field_edit'
            })

            # Update stats
            self.stats['total_field_edits'] += 1
            if field_name not in self.stats['edits_by_field']:
                self.stats['edits_by_field'][field_name] = 0
            self.stats['edits_by_field'][field_name] += 1

            return True
        return False

    def _truncate_for_display(self, value, max_length=100):
        """Truncate a value for display in edit history."""
        if value is None:
            return ""
        if isinstance(value, list):
            value = str(value)
        value = str(value)
        if len(value) > max_length:
            return value[:max_length] + "..."
        return value

    def apply_term_decisions(self, record, term_decisions, record_id):
        """Apply vocabulary term decisions to a record."""
        analysis = record.get('analysis', {})

        # Track term decisions
        for term_id, status in term_decisions.items():
            self.stats['total_term_decisions'] += 1
            if status == 'approved':
                self.stats['terms_approved'] += 1
            elif status == 'rejected':
                self.stats['terms_rejected'] += 1

            self.edit_history.append({
                'record_id': record_id,
                'field': 'vocabulary_term',
                'original_value': term_id,
                'new_value': status,
                'edit_type': 'term_decision'
            })

        # Store term decisions in the record for reference
        if 'archivist_term_decisions' not in analysis:
            analysis['archivist_term_decisions'] = {}
        analysis['archivist_term_decisions'] = term_decisions
        record['analysis'] = analysis

        return True

    def apply_custom_terms(self, record, custom_terms, custom_subjects, record_id):
        """Apply custom terms and subjects added by archivist."""
        analysis = record.get('analysis', {})

        # Add custom terms to a dedicated field
        if custom_terms:
            analysis['archivist_custom_terms'] = custom_terms
            self.stats['custom_terms_added'] += len(custom_terms)

            for term in custom_terms:
                self.edit_history.append({
                    'record_id': record_id,
                    'field': 'custom_term',
                    'original_value': '',
                    'new_value': f"{term.get('label', '')} ({term.get('source', '')})",
                    'edit_type': 'custom_term_added'
                })

        # Add custom subjects to the subjects list
        if custom_subjects:
            existing_subjects = analysis.get('subjects', [])
            if not isinstance(existing_subjects, list):
                existing_subjects = [existing_subjects] if existing_subjects else []

            for subj in custom_subjects:
                label = subj.get('label', '')
                if label and label not in existing_subjects:
                    existing_subjects.append(label)
                    self.stats['custom_subjects_added'] += 1

                    self.edit_history.append({
                        'record_id': record_id,
                        'field': 'custom_subject',
                        'original_value': '',
                        'new_value': label,
                        'edit_type': 'custom_subject_added'
                    })

            analysis['subjects'] = existing_subjects

        record['analysis'] = analysis
        return True

    def apply_all_edits(self):
        """Apply all archivist edits to the workflow data."""
        if not self.decisions_data or not self.workflow_data:
            return False

        decisions = self.decisions_data.get('decisions', [])

        for decision in decisions:
            record_id = decision.get('record_id')

            # Find the corresponding record in workflow data (1-indexed to 0-indexed)
            record_idx = record_id - 1
            if record_idx < 0 or record_idx >= len(self.workflow_data):
                print(f"   Warning: Record ID {record_id} not found in workflow data")
                continue

            record = self.workflow_data[record_idx]
            has_edits = False

            # Apply field edits
            edits = decision.get('edits', {})
            if edits:
                for field_name, edit_data in edits.items():
                    if edit_data.get('edited', False):
                        self.apply_field_edit(record, field_name, edit_data, record_id)
                        has_edits = True

            # Apply term decisions
            term_decisions = decision.get('term_decisions', {})
            if term_decisions:
                self.apply_term_decisions(record, term_decisions, record_id)
                has_edits = True

            # Apply custom terms and subjects
            custom_terms = decision.get('custom_terms', [])
            custom_subjects = decision.get('custom_subjects', [])
            if custom_terms or custom_subjects:
                self.apply_custom_terms(record, custom_terms, custom_subjects, record_id)
                has_edits = True

            # Add archivist metadata
            analysis = record.get('analysis', {})
            analysis['archivist_reviewed'] = decision.get('reviewed', False)
            # Support both old 'cataloger_notes' and new 'archivist_notes' keys
            analysis['archivist_notes'] = decision.get('archivist_notes',
                                           decision.get('cataloger_notes', ''))
            analysis['archivist_review_date'] = self.stats['export_timestamp']
            analysis['archivist_name'] = self.stats['archivist_name']
            record['analysis'] = analysis

            # Update stats
            if has_edits:
                self.stats['records_with_edits'] += 1
            elif decision.get('reviewed', False):
                self.stats['records_reviewed_only'] += 1

        return True

    def save_workflow_json(self):
        """Save the updated workflow JSON."""
        try:
            # Re-add api_stats if it existed
            output_data = self.workflow_data.copy()
            if self.api_stats:
                output_data.append(self.api_stats)

            with open(self.workflow_json_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, indent=2, ensure_ascii=False)

            print(f"   Saved updated workflow JSON")
            return True
        except Exception as e:
            print(f"Error saving workflow JSON: {e}")
            return False

    def update_excel_deliverable(self):
        """Update the Excel deliverable with archivist edits and add Edit History sheet."""
        if not os.path.exists(self.workflow_excel_path):
            print(f"   Warning: Excel file not found at {self.workflow_excel_path}")
            return False

        try:
            wb = load_workbook(self.workflow_excel_path)

            # Get or create the main analysis sheet
            if 'Analysis' in wb.sheetnames:
                ws = wb['Analysis']
            elif 'Sheet' in wb.sheetnames:
                ws = wb['Sheet']
            else:
                ws = wb.active

            # Update cells based on edits
            # The Excel columns typically are:
            # A: Folder, B: Page, C: Image, D: Title, E: Contributors, F: Genre,
            # G: OCR Text, H: Description, I: Format, J: Subjects, K: Date,
            # L: Sheet Info, M: Named Entities, N: Geographic Entities, O: Content Warning

            column_mapping = {
                'title': 4,  # D
                'contributors': 5,  # E
                'genre': 6,  # F
                'ocr_text': 7,  # G
                'description': 8,  # H
                'format_media': 9,  # I
                'subjects': 10,  # J
                'date_on_drawing': 11,  # K
                'sheet_info': 12,  # L
                'named_entities': 13,  # M
                'geographic_entities': 14,  # N
                'content_warning': 15  # O
            }

            # Apply edits to Excel
            for decision in self.decisions_data.get('decisions', []):
                record_id = decision.get('record_id')
                row_num = record_id + 1  # +1 for header row

                edits = decision.get('edits', {})
                for field_name, edit_data in edits.items():
                    if edit_data.get('edited', False) and field_name in column_mapping:
                        col_num = column_mapping[field_name]
                        new_value = edit_data.get('value', '')

                        # Format list values
                        if isinstance(new_value, list):
                            if field_name == 'contributors':
                                # Format contributors specially
                                contrib_strs = []
                                for c in new_value:
                                    if isinstance(c, dict):
                                        name = c.get('name', '')
                                        role = c.get('role', '')
                                        if role:
                                            contrib_strs.append(f"{name} ({role})")
                                        else:
                                            contrib_strs.append(name)
                                    else:
                                        contrib_strs.append(str(c))
                                new_value = '; '.join(contrib_strs)
                            else:
                                new_value = ', '.join(str(v) for v in new_value)

                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = new_value
                        # Highlight edited cells
                        cell.fill = PatternFill(start_color="FFFFD700", end_color="FFFFD700", fill_type="solid")

            # Add Edit History sheet
            self._add_edit_history_sheet(wb)

            wb.save(self.workflow_excel_path)
            print(f"   Saved updated Excel deliverable")
            return True

        except Exception as e:
            print(f"Error updating Excel: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _add_edit_history_sheet(self, wb):
        """Add an Edit History sheet to the workbook."""
        # Remove existing Edit History sheet if present
        if 'Edit History' in wb.sheetnames:
            del wb['Edit History']

        ws = wb.create_sheet('Edit History')

        # Styles
        header_fill = PatternFill(start_color="FF2C3E50", end_color="FF2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        stats_fill = PatternFill(start_color="FFE8F4F8", end_color="FFE8F4F8", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Summary Statistics Section
        ws['A1'] = "ARCHIVIST EDIT INTEGRATION SUMMARY"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')

        self.stats['integration_timestamp'] = datetime.now().isoformat()

        summary_data = [
            ("Archivist Name:", self.stats['archivist_name']),
            ("Export Timestamp:", self.stats['export_timestamp']),
            ("Integration Timestamp:", self.stats['integration_timestamp']),
            ("", ""),
            ("Total Records in Export:", self.stats['total_records_in_export']),
            ("Records with Edits:", self.stats['records_with_edits']),
            ("Records Reviewed Only (no edits):", self.stats['records_reviewed_only']),
            ("", ""),
            ("Total Field Edits:", self.stats['total_field_edits']),
            ("Total Term Decisions:", self.stats['total_term_decisions']),
            ("Terms Approved:", self.stats['terms_approved']),
            ("Terms Rejected:", self.stats['terms_rejected']),
            ("Custom Terms Added:", self.stats['custom_terms_added']),
            ("Custom Subjects Added:", self.stats['custom_subjects_added']),
        ]

        row = 3
        for label, value in summary_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            if label:
                ws.cell(row=row, column=1).fill = stats_fill
                ws.cell(row=row, column=2).fill = stats_fill
            row += 1

        # Edits by Field Section
        row += 1
        ws.cell(row=row, column=1, value="Edits by Field:").font = Font(bold=True)
        row += 1
        for field, count in sorted(self.stats['edits_by_field'].items()):
            ws.cell(row=row, column=1, value=f"  {field}:")
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Detailed Edit History Section
        row += 2
        ws.cell(row=row, column=1, value="DETAILED EDIT HISTORY").font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        # Headers
        headers = ["Record ID", "Field", "Edit Type", "Original Value", "New Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row += 1

        # Edit history rows
        for edit in self.edit_history:
            ws.cell(row=row, column=1, value=edit['record_id']).border = border
            ws.cell(row=row, column=2, value=edit['field']).border = border
            ws.cell(row=row, column=3, value=edit['edit_type']).border = border

            orig_cell = ws.cell(row=row, column=4, value=edit['original_value'])
            orig_cell.border = border
            orig_cell.alignment = Alignment(wrap_text=True, vertical='top')

            new_cell = ws.cell(row=row, column=5, value=edit['new_value'])
            new_cell.border = border
            new_cell.alignment = Alignment(wrap_text=True, vertical='top')

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40

    def print_summary(self):
        """Print a summary of the integration."""
        print("\n" + "=" * 60)
        print("INTEGRATION SUMMARY")
        print("=" * 60)
        print(f"Archivist: {self.stats['archivist_name']}")
        print(f"Total records processed: {self.stats['total_records_in_export']}")
        print(f"Records with edits: {self.stats['records_with_edits']}")
        print(f"Records reviewed only: {self.stats['records_reviewed_only']}")
        print(f"\nField edits: {self.stats['total_field_edits']}")
        if self.stats['edits_by_field']:
            for field, count in sorted(self.stats['edits_by_field'].items()):
                print(f"   {field}: {count}")
        print(f"\nTerm decisions: {self.stats['total_term_decisions']}")
        print(f"   Approved: {self.stats['terms_approved']}")
        print(f"   Rejected: {self.stats['terms_rejected']}")
        print(f"Custom terms added: {self.stats['custom_terms_added']}")
        print(f"Custom subjects added: {self.stats['custom_subjects_added']}")

    def run(self):
        """Main execution method."""
        print("\n" + "=" * 60)
        print("Step 6: Integrate Archivist Edits")
        print("=" * 60)

        print(f"\nUsing folder: {self.folder_name}")

        # Find latest export
        print("\n1. Finding archivist decisions export...")
        export_path = self.find_latest_export()
        if not export_path:
            return False
        print(f"   Found: {os.path.basename(export_path)}")

        # Load decisions
        print("\n2. Loading archivist decisions...")
        if not self.load_decisions(export_path):
            return False

        # Load workflow JSON
        print("\n3. Loading workflow data...")
        if not self.load_workflow_json():
            return False

        # Backup original files
        print("\n4. Backing up original files...")
        self.backup_original_files()

        # Apply edits
        print("\n5. Applying archivist edits...")
        if not self.apply_all_edits():
            return False
        print(f"   Applied edits to {self.stats['records_with_edits']} records")

        # Save updated JSON
        print("\n6. Saving updated workflow JSON...")
        if not self.save_workflow_json():
            return False

        # Update Excel
        print("\n7. Updating Excel deliverable...")
        self.update_excel_deliverable()

        # Print summary
        self.print_summary()

        print("\n" + "=" * 60)
        print("STEP 6 COMPLETE")
        print("=" * 60)
        print(f"\nUpdated files:")
        print(f"   {self.workflow_json_path}")
        print(f"   {self.workflow_excel_path}")
        print(f"\nOriginal files backed up to:")
        print(f"   {self.original_outputs_folder}")
        print("=" * 60)

        return True


def main():
    """Main entry point."""
    print("Step 6: Integrate Archivist Edits")
    print("=" * 60)

    # Find newest output folder
    base_output_dir = os.path.join(script_dir, "output_folders")
    folder_path = find_newest_folder(base_output_dir)

    if not folder_path:
        print(f"No output folders found in: {base_output_dir}")
        print("Please run Steps 1-5 first.")
        return 1

    print(f"Auto-selected newest folder: {os.path.basename(folder_path)}")

    # Confirm with user
    response = input("\nThis will modify deliverable files. Original files will be backed up. Continue? (yes/no): ").strip().lower()
    if response not in ['yes', 'y']:
        print("Operation cancelled.")
        return 0

    # Create integrator and run
    integrator = ArchivistEditsIntegrator(folder_path)
    success = integrator.run()

    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())
