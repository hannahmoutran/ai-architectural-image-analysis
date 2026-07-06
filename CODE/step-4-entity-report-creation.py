# Step 4: Entity Report Creation
# Create local authority file for named entities with fuzzy matching

import os
import json
import logging
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from collections import defaultdict
import re
from shared_utilities import find_newest_folder
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# Fuzzy matching imports - try rapidfuzz first (faster), fallback to fuzzywuzzy
try:
    from rapidfuzz import fuzz, process
    FUZZY_LIB = "rapidfuzz"
except ImportError:
    try:
        from fuzzywuzzy import fuzz, process
        FUZZY_LIB = "fuzzywuzzy"
    except ImportError:
        print("Warning: Neither rapidfuzz nor fuzzywuzzy found. Install one for fuzzy matching:")
        print("pip install rapidfuzz  # recommended")
        print("# or")
        print("pip install fuzzywuzzy python-levenshtein")
        FUZZY_LIB = None

class EntityAuthority:
    """Build authority records for named entities with fuzzy matching."""
    
    def __init__(self, folder_path: str, similarity_threshold: int = 85):
        self.folder_path = folder_path
        self.workflow_type = None
        self.json_data = None
        self.json_folder = None
        self.similarity_threshold = similarity_threshold
        self.entity_clusters = {}  # Maps normalized names to canonical forms
        
    def detect_workflow_type(self) -> bool:
        """Detect workflow type."""
        json_folder = os.path.join(self.folder_path, "metadata", "json")

        # Check for workflow JSON files in metadata/json folder
        if os.path.exists(os.path.join(json_folder, "text_workflow.json")):
            self.workflow_type = 'text'
            self.json_folder = json_folder
        elif os.path.exists(os.path.join(json_folder, "image_workflow.json")):
            self.workflow_type = 'image'
            self.json_folder = json_folder
        elif os.path.exists(os.path.join(json_folder, "drawings_workflow.json")):
            self.workflow_type = 'drawings'
            self.json_folder = json_folder
        else:
            return False
        return True

    def load_json_data(self) -> bool:
        """Load JSON data."""
        json_filename = f"{self.workflow_type}_workflow.json"
        json_path = os.path.join(self.json_folder, json_filename)
        
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                self.json_data = json.load(f)
            return True
        except Exception as e:
            logging.error(f"Error loading JSON data: {e}")
            return False
    
    def normalize_entity_name(self, name: str) -> str:
        """Basic normalization for entity names."""
        # Remove extra whitespace and punctuation
        normalized = re.sub(r'\s+', ' ', name.strip())
        
        # Handle common abbreviations and variations for better matching
        normalized = normalized.replace('&', 'and')
        normalized = re.sub(r'\bCo\.\b', 'Company', normalized, flags=re.IGNORECASE)
        normalized = re.sub(r'\bInc\.\b', 'Incorporated', normalized, flags=re.IGNORECASE)
        
        return normalized
    
    def find_similar_entity(self, entity_name: str, existing_entities: List[str]) -> Optional[str]:
        """Find if entity is similar to existing ones using fuzzy matching."""
        if not FUZZY_LIB or not existing_entities:
            return None
        
        normalized_name = self.normalize_entity_name(entity_name)
        normalized_existing = [self.normalize_entity_name(e) for e in existing_entities]
        
        # Find best match
        match = process.extractOne(
            normalized_name, 
            normalized_existing, 
            scorer=fuzz.ratio,
            score_cutoff=self.similarity_threshold
        )
        
        if match:
            # Return the original entity name that corresponds to the match
            match_index = normalized_existing.index(match[0])
            return existing_entities[match_index]
        
        return None
    
    def parse_typed_entity(self, entity_str: str) -> Tuple[str, str]:
        """Parse entity string to extract name and type.
        
        Input: 'John F. Staub (architect)' or 'Atlanta, Ga. (location)'
        Output: ('John F. Staub', 'architect')
        """
        # Look for pattern: 'Name (type)'
        match = re.match(r'^(.+?)\s*\(([^)]+)\)\s*$', entity_str.strip())
        
        if match:
            name = match.group(1).strip()
            entity_type = match.group(2).strip().lower()
            return name, entity_type
        else:
            # If no type indicator found, return original string and classify it
            name = entity_str.strip()
            entity_type = self.classify_entity_fallback(name)
            return name, entity_type
    
    def classify_entity_fallback(self, entity: str) -> str:
        """Fallback classification for entities without type indicators."""
        entity_lower = entity.lower()
        
        # Geographic locations (common suffixes)
        if any(suffix in entity for suffix in [', Tenn.', ', Md.', ', Va.', ', Ga.', ', Ala.', ', N. C.', ', Fla.', ', S. C.', ', Tex.', ', Miss.', ', Ky.', ', La.', ', Ark.', ', Mo.', ', Okla.', ', N. J.', ', N. Y.', ', Pa.', ', D. C.', ' (location)', ', Tx.', ', Calif.', ', Cal.']):
            return 'location'
        
        # Residences
        if 'residence' in entity_lower or 'house' in entity_lower:
            return 'building'
        
        # Organizations/Companies
        if any(term in entity_lower for term in ['company', 'co.', 'inc.', 'corp.', 'publishing', 'association', 'institute']):
            return 'organization'
        
        # Architectural firms (multiple names with &, commas)
        if any(pattern in entity for pattern in ['&', ' and ', ', ']):
            return 'firm'
        
        # Personal names (has initials or multiple capitalized words)
        if re.match(r'^[A-Z]\. ?[A-Z]\.? [A-Z]', entity):  # J. Staub, E. R. Denmark
            return 'person'
        elif re.match(r'^[A-Z][a-z]+ [A-Z]\. [A-Z]', entity):  # Laurence H. Fowler
            return 'person'
        elif len(entity.split()) >= 2 and all(word[0].isupper() for word in entity.split()[:2]):
            return 'person'
        
        return 'unknown'
    
    def cluster_similar_entities(self, raw_entities: List[str]) -> Dict[str, List[str]]:
        """Group similar entities using fuzzy matching."""
        if not FUZZY_LIB:
            # If no fuzzy matching library, treat each entity as unique
            return {entity: [entity] for entity in raw_entities}
        
        clusters = {}
        processed = set()
        
        print(f"Clustering {len(raw_entities)} entities with similarity threshold {self.similarity_threshold}%...")
        
        for entity in raw_entities:
            if entity in processed:
                continue
                
            # Find all similar entities
            similar_entities = [entity]
            processed.add(entity)
            
            for other_entity in raw_entities:
                if other_entity in processed:
                    continue
                    
                # Check similarity
                similarity = fuzz.ratio(
                    self.normalize_entity_name(entity),
                    self.normalize_entity_name(other_entity)
                )
                
                if similarity >= self.similarity_threshold:
                    similar_entities.append(other_entity)
                    processed.add(other_entity)
            
            # Use the most common or longest form as canonical
            canonical = max(similar_entities, key=lambda x: (raw_entities.count(x), len(x)))
            clusters[canonical] = similar_entities
        
        # Show clustering results
        merged_count = sum(len(cluster) - 1 for cluster in clusters.values() if len(cluster) > 1)
        if merged_count > 0:
            print(f"   Merged {merged_count} entity variations into {len(clusters)} unique entities")
            
            # Show some examples of merges
            merge_examples = [(canonical, cluster) for canonical, cluster in clusters.items() if len(cluster) > 1][:5]
            for canonical, cluster in merge_examples:
                print(f"   • '{canonical}' ← {cluster[1:]}")
        
        return clusters
    
    def extract_all_entities(self) -> Dict[str, Dict]:
        """Extract all named entities and analyze patterns with fuzzy matching."""
        raw_entity_data = []  # Store raw entities with their contexts
        
        # Skip API stats
        data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data
        
        # First pass: collect all raw entities
        for item in data_items:
            if 'analysis' in item:
                analysis = item['analysis']
                folder = item.get('folder', 'unknown')
                # Extract filename from image_path for instance reference
                image_path = item.get('image_path', '')
                if image_path:
                    filename = os.path.basename(image_path)
                    instance_ref = f"{folder}/{filename}"
                else:
                    # Fallback for text workflows that may not have image_path
                    page = item.get('page_number', 0)
                    instance_ref = f"{folder}/page{page}"
                
                # Extract entities
                entities = analysis.get('named_entities', [])
                if isinstance(entities, str):
                    entities = [e.strip() for e in entities.split(',') if e.strip()]

                # Expand entities that may be concatenated with " - " delimiter
                # (handles cases like "- Name1 (type) - Name2 (type) - Name3 (type)")
                expanded_entities = []
                for entity in entities:
                    # Remove leading "- " if present (bullet point format)
                    entity = entity.lstrip('- ').strip()
                    # Check if this contains multiple entities separated by " - "
                    # Pattern: look for " - " followed by a capital letter or quote (start of new entity)
                    if ' - ' in entity and re.search(r' - [A-Z"\']', entity):
                        # Split on " - " but be careful with patterns like "(type) - Name"
                        parts = re.split(r'\s+-\s+(?=[A-Z"\'])', entity)
                        expanded_entities.extend([p.strip() for p in parts if p.strip()])
                    else:
                        if entity:
                            expanded_entities.append(entity)
                entities = expanded_entities
                
                # Extract context for additional analysis
                # Support all workflow types: text (cleaned_text), image (text_transcription), drawings (description)
                text_content = (analysis.get('cleaned_text', '') or
                               analysis.get('text_transcription', '') or
                               analysis.get('description', ''))
                # Support toc_entry (text/image workflows) or title (drawings workflow)
                toc_entry = analysis.get('toc_entry', '') or analysis.get('title', '')
                context = f"{text_content} {toc_entry}".lower()
                
                for entity_str in entities:
                    if not entity_str or len(entity_str.strip()) < 2:
                        continue
                    
                    entity_name, entity_type = self.parse_typed_entity(entity_str)
                    if not entity_name:
                        continue
                    
                    raw_entity_data.append({
                        'original_str': entity_str.strip(),
                        'entity_name': entity_name,
                        'entity_type': entity_type,
                        'folder': folder,
                        'instance_ref': instance_ref,
                        'context': context
                    })
        
        # Get all unique entity names for clustering
        all_entity_names = [item['entity_name'] for item in raw_entity_data]
        entity_clusters = self.cluster_similar_entities(all_entity_names)
        
        # Second pass: group by canonical entity names
        entity_records = defaultdict(lambda: {
            'appearances': [],
            'contexts': [],
            'entity_type': 'unknown',
            'locations': set(),
            'time_periods': set(),
            'original_forms': set(),
            'clustered_variants': set()
        })
        
        for item in raw_entity_data:
            entity_name = item['entity_name']
            
            # Find canonical name for this entity
            canonical_name = None
            for canonical, cluster in entity_clusters.items():
                if entity_name in cluster:
                    canonical_name = canonical
                    break
            
            if not canonical_name:
                canonical_name = entity_name
            
            record = entity_records[canonical_name]
            record['appearances'].append(item['instance_ref'])
            record['contexts'].append(item['context'])
            record['original_forms'].add(item['original_str'])
            record['clustered_variants'].add(entity_name)
            
            # Set entity type (use the parsed type or keep existing if already set)
            if record['entity_type'] == 'unknown' or item['entity_type'] != 'unknown':
                record['entity_type'] = item['entity_type']
            
            # Extract year from folder name
            year_match = re.search(r'(\d{4})', item['folder'])
            if year_match:
                record['time_periods'].add(year_match.group(1))
        
        # Convert sets to lists for JSON serialization and calculate frequency
        for entity, record in entity_records.items():
            record['locations'] = list(record['locations'])
            record['time_periods'] = list(record['time_periods'])
            record['original_forms'] = list(record['original_forms'])
            record['clustered_variants'] = list(record['clustered_variants'])
            record['frequency'] = len(record['appearances'])
        
        return dict(entity_records)
    
    def create_authority_file(self, entity_records: Dict) -> bool:
        """Create comprehensive authority file."""
        try:
            authority_path = os.path.join(self.json_folder, "entity_authority.json")

            # Create structured authority data
            authority_data = {
                "metadata": {
                    "created": datetime.now().isoformat(),
                    "source": "Architectural Drawing Collection",
                    "total_entities": len(entity_records),
                    "extraction_method": "Named Entity Recognition with Type Classification",
                    "fuzzy_matching": {
                        "enabled": FUZZY_LIB is not None,
                        "library": FUZZY_LIB,
                        "similarity_threshold": self.similarity_threshold
                    },
                    "entity_types": list(set(record['entity_type'] for record in entity_records.values()))
                },
                "entities": entity_records,
                "statistics": self.generate_statistics(entity_records)
            }
            
            with open(authority_path, 'w', encoding='utf-8') as f:
                json.dump(authority_data, f, indent=2, ensure_ascii=False)
            
            print(f"Created entity authority file: {authority_path}")
            return True
            
        except Exception as e:
            logging.error(f"Error creating authority file: {e}")
            return False
    
    def generate_statistics(self, entity_records: Dict) -> Dict:
        """Generate statistics about entities."""
        stats = {
            'by_type': defaultdict(int),
            'by_frequency': defaultdict(int),
            'by_decade': defaultdict(int),
            'type_examples': defaultdict(list),
            'clustering_stats': {
                'entities_with_variants': 0,
                'total_variants_merged': 0
            }
        }
        
        for entity, record in entity_records.items():
            entity_type = record['entity_type']
            frequency = record['frequency']
            
            stats['by_type'][entity_type] += 1
            
            # Add examples for each type (up to 3)
            if len(stats['type_examples'][entity_type]) < 3:
                stats['type_examples'][entity_type].append(entity)
            
            # Clustering statistics
            if len(record['clustered_variants']) > 1:
                stats['clustering_stats']['entities_with_variants'] += 1
                stats['clustering_stats']['total_variants_merged'] += len(record['clustered_variants']) - 1
            
            # Frequency buckets
            if frequency == 1:
                stats['by_frequency']['single_mention'] += 1
            elif frequency <= 3:
                stats['by_frequency']['low_frequency'] += 1
            elif frequency <= 10:
                stats['by_frequency']['medium_frequency'] += 1
            else:
                stats['by_frequency']['high_frequency'] += 1
            
            # Decade analysis
            for year in record['time_periods']:
                decade = f"{year[:3]}0s"
                stats['by_decade'][decade] += 1
        
        # Convert defaultdicts to regular dicts
        return {
            'by_type': dict(stats['by_type']),
            'by_frequency': dict(stats['by_frequency']),
            'by_decade': dict(stats['by_decade']),
            'type_examples': dict(stats['type_examples']),
            'clustering_stats': stats['clustering_stats']
        }
    
    def create_human_readable_report(self, entity_records: Dict) -> bool:
        """Create human-readable authority report."""
        try:
            metadata_dir = os.path.join(self.folder_path, "metadata")
            report_path = os.path.join(metadata_dir, "entity_authority_report.txt")
            
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write("ENTITY AUTHORITY REPORT\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total Entities: {len(entity_records)}\n")
                f.write(f"Fuzzy Matching: {'Enabled' if FUZZY_LIB else 'Disabled'}")
                if FUZZY_LIB:
                    f.write(f" ({FUZZY_LIB}, threshold: {self.similarity_threshold}%)")
                f.write("\n\n")
                
                # Show statistics summary
                stats = self.generate_statistics(entity_records)
                
                if FUZZY_LIB and stats['clustering_stats']['entities_with_variants'] > 0:
                    f.write("FUZZY MATCHING RESULTS:\n")
                    f.write("-" * 25 + "\n")
                    f.write(f"  Entities with variants: {stats['clustering_stats']['entities_with_variants']}\n")
                    f.write(f"  Total variants merged: {stats['clustering_stats']['total_variants_merged']}\n\n")
                
                f.write("ENTITY TYPE BREAKDOWN:\n")
                f.write("-" * 25 + "\n")
                for entity_type, count in sorted(stats['by_type'].items(), key=lambda x: x[1], reverse=True):
                    f.write(f"  {entity_type.upper().replace('_', ' ')}: {count} entities\n")
                f.write("\n")
                
                # Group by type for detailed listing
                by_type = defaultdict(list)
                for entity, record in entity_records.items():
                    by_type[record['entity_type']].append((entity, record))
                
                for entity_type in sorted(by_type.keys()):
                    entities = by_type[entity_type]
                    f.write(f"{entity_type.upper().replace('_', ' ')} ({len(entities)} entities):\n")
                    f.write("-" * 40 + "\n")
                    
                    # Sort by frequency (most mentioned first)
                    entities.sort(key=lambda x: x[1]['frequency'], reverse=True)
                    
                    for entity, record in entities:
                        f.write(f"  {entity}\n")
                        f.write(f"    Frequency: {record['frequency']} appearances\n")
                        f.write(f"    Time periods: {', '.join(sorted(record['time_periods']))}\n")
                        
                        # Show clustered variants if any
                        if len(record['clustered_variants']) > 1:
                            variants = [v for v in record['clustered_variants'] if v != entity]
                            f.write(f"    Variants: {', '.join(variants)}\n")
                        
                        # Show different forms if multiple
                        if len(record['original_forms']) > 1:
                            f.write(f"    Forms: {', '.join(record['original_forms'])}\n")
                        
                        f.write(f"    Instances: {', '.join(record['appearances'][:5])}")
                        if len(record['appearances']) > 5:
                            f.write(f" ... (+{len(record['appearances'])-5} more)")
                        f.write("\n\n")
                
            print(f"Created human-readable report: {report_path}")
            return True
            
        except Exception as e:
            logging.error(f"Error creating report: {e}")
            return False
    
    def create_final_excel_deliverable(self, entity_records: Dict) -> bool:
        """Create final Excel workbook with 3 sheets (Edit Statistics added later by integrate-archivist-edits)."""
        try:
            metadata_dir = os.path.join(self.folder_path, "metadata")
            excel_path = os.path.join(metadata_dir, "final_deliverable.xlsx")

            wb = Workbook()

            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_alignment = Alignment(vertical="top", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # ==================== SHEET 1: Final Metadata ====================
            ws_metadata = wb.active
            ws_metadata.title = "Final Metadata"

            # Get data excluding API stats
            data_items = self.json_data[:-1] if self.json_data and 'api_stats' in self.json_data[-1] else self.json_data

            # Load vocabulary mapping to get selected subject headings
            vocab_mapping_path = os.path.join(self.json_folder, "vocabulary_mapping.json")
            selected_headings_by_page = {}  # key: (folder, page_number) -> list of selected terms

            if os.path.exists(vocab_mapping_path):
                with open(vocab_mapping_path, 'r', encoding='utf-8') as f:
                    vocab_data = json.load(f)

                for drawing in vocab_data.get('drawings', []):
                    folder = drawing.get('folder', '')
                    page_num = drawing.get('page_number', 0)
                    key = (folder, page_num)

                    selected_terms = []
                    vocab_results = drawing.get('vocabulary_search_results', {})
                    for subject, terms in vocab_results.items():
                        for term in terms:
                            if term.get('selected', False):
                                # Include label, URI, and vocabulary source
                                label = term.get('label', '')
                                uri = term.get('uri', '')
                                source = term.get('source', '')
                                selected_terms.append(f"{label} [{source}] ({uri})")

                    selected_headings_by_page[key] = selected_terms

            # Headers for metadata sheet - matching JSON field names
            metadata_headers = [
                "Folder", "Filename", "Title", "Contributors", "Genre",
                "Description", "Format/Media", "Date on Drawing", "Sheet Info",
                "Topics (AI)", "Subject Headings (Controlled Vocab)",
                "Named Entities", "Geographic Entities", "Content Warning"
            ]

            for col, header in enumerate(metadata_headers, 1):
                cell = ws_metadata.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Populate metadata rows
            for row_num, item in enumerate(data_items, 2):
                analysis = item.get('analysis', {})

                folder = item.get('folder', '')
                page_num = item.get('page_number', 0)
                image_path = item.get('image_path', '')
                filename = os.path.basename(image_path) if image_path else ''

                # Get contributors (list of dicts with name and role)
                contributors = analysis.get('contributors', [])
                if isinstance(contributors, list):
                    contrib_parts = []
                    for c in contributors:
                        if isinstance(c, dict):
                            name = c.get('name', '')
                            role = c.get('role', '')
                            if name and role:
                                contrib_parts.append(f"{name} ({role})")
                            elif name:
                                contrib_parts.append(name)
                        elif isinstance(c, str):
                            contrib_parts.append(c)
                    contributors_str = '; '.join(contrib_parts)
                else:
                    contributors_str = str(contributors) if contributors else ''

                # Get AI-generated topics
                subjects = analysis.get('topics', [])
                if isinstance(subjects, list):
                    subjects_str = '; '.join(subjects)
                else:
                    subjects_str = str(subjects) if subjects else ''

                # Get selected vocabulary subject headings from final_selected_terms
                final_terms = analysis.get('final_selected_terms', [])
                if final_terms:
                    headings_parts = []
                    for term in final_terms:
                        label = term.get('label', '')
                        uri = term.get('uri', '')
                        source = term.get('source', '')
                        if label:
                            headings_parts.append(f"{label} [{source}] ({uri})")
                    headings_str = '; '.join(headings_parts)
                else:
                    # Fallback to old method if final_selected_terms not present
                    key = (folder, page_num)
                    selected_headings = selected_headings_by_page.get(key, [])
                    headings_str = '; '.join(selected_headings) if selected_headings else ''

                # Handle named entities - may be list or string
                entities = analysis.get('named_entities', [])
                if isinstance(entities, list):
                    entities_str = '; '.join(entities)
                else:
                    entities_str = str(entities) if entities else ''

                # Handle geographic entities - may be list or string
                geo_entities = analysis.get('geographic_entities', [])
                if isinstance(geo_entities, list):
                    geo_entities_str = '; '.join(geo_entities)
                else:
                    geo_entities_str = str(geo_entities) if geo_entities else ''

                row_data = [
                    folder,
                    filename,
                    analysis.get('title', ''),
                    contributors_str,
                    analysis.get('genre', '') or analysis.get('document_type', ''),
                    analysis.get('description', ''),
                    analysis.get('format_media', ''),
                    analysis.get('date_on_drawing', '') or analysis.get('date', ''),
                    analysis.get('sheet_info', ''),
                    subjects_str,
                    headings_str,
                    entities_str,
                    geo_entities_str,
                    analysis.get('content_warning', '')
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws_metadata.cell(row=row_num, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

            # Set column widths for metadata sheet
            col_widths = [20, 25, 30, 35, 25, 60, 40, 15, 40, 40, 60, 40, 30, 15]
            for col, width in enumerate(col_widths, 1):
                ws_metadata.column_dimensions[ws_metadata.cell(row=1, column=col).column_letter].width = width

            # Freeze header row
            ws_metadata.freeze_panes = 'A2'

            # Save workbook
            wb.save(excel_path)
            print(f"Created final Excel deliverable: {excel_path}")
            return True

        except Exception as e:
            logging.error(f"Error creating Excel deliverable: {e}")
            import traceback
            traceback.print_exc()
            return False

    def run(self) -> bool:
        """Main execution method."""
        print(f"\nENTITY AUTHORITY BUILDER")
        print(f"Processing folder: {self.folder_path}")
        if FUZZY_LIB:
            print(f"Fuzzy matching: Enabled ({FUZZY_LIB}, threshold: {self.similarity_threshold}%)")
        else:
            print("Fuzzy matching: Disabled (install rapidfuzz or fuzzywuzzy)")
        print("-" * 50)
        
        if not self.detect_workflow_type():
            print("Could not detect workflow type")
            return False

        if not self.load_json_data():
            print("Could not load JSON data")
            return False
        
        # Extract all entities with fuzzy matching
        print("Extracting and analyzing typed named entities with fuzzy matching...")
        entity_records = self.extract_all_entities()
        
        if not entity_records:
            print("No named entities found")
            return False
        
        print(f"Found {len(entity_records)} unique entities after clustering")
        
        # Show type breakdown
        type_counts = defaultdict(int)
        for record in entity_records.values():
            type_counts[record['entity_type']] += 1
        
        for entity_type, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"   - {entity_type}: {count} entities")
        
        # Create authority files
        if not self.create_authority_file(entity_records):
            return False

        if not self.create_human_readable_report(entity_records):
            return False

        # Create final Excel deliverable with 4 sheets
        if not self.create_final_excel_deliverable(entity_records):
            return False

        print(f"\nSTEP 4 COMPLETE: Built entity authority in {os.path.basename(self.folder_path)}")
        print(f"Created: entity_authority.json, entity_authority_report.txt, final_deliverable.xlsx")

        return True
    
def main():
    
    # Default base directory for output folders
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_output_dir = os.path.join(script_dir, "output_folders")

    folder_path = find_newest_folder(base_output_dir)
    if not folder_path:
            print(f"No folders found in: {base_output_dir}")
            return 1
    print(f"Auto-selected newest folder: {os.path.basename(folder_path)}")

    builder = EntityAuthority(folder_path, similarity_threshold=85)
    success = builder.run()
    
    return 0 if success else 1

if __name__ == "__main__":
    exit(main())