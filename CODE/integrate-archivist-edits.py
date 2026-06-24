#!/usr/bin/env python3
"""
Integrate Archivist Edits into Workflow Outputs
================================================

Processes archivist review decisions from exported JSON and updates all deliverable files.

This script:
1. Reads archivist decisions from JSON exports (created by html-review.py)
2. Backs up original JSON files to an 'original-outputs' folder
3. Applies edits to the workflow JSON (metadata/json/drawings_workflow.json)
4. Handles cascade rejection logic (rejected subjects → reject derived headings)
5. Generates final_metadata.json with only approved/clean data
6. Creates edit_report.json (combined statistics and edit history)
7. Creates final_deliverable.xlsx with 2 sheets:
   - Final Metadata: All drawings with approved metadata
   - Edit Statistics: Summary of archivist edits and acceptance rates

Usage:
    python integrate-archivist-edits.py                           # Use latest export
    python integrate-archivist-edits.py --decisions path/to.json  # Use specific export
    python run.py 6

Analysis-Only Mode (for comparing multiple evaluators):
    python integrate-archivist-edits.py --evaluator "Alice" --decisions alice_edits.json
    python integrate-archivist-edits.py --evaluator "Bob" --decisions bob_edits.json

    When --evaluator is specified:
    - No workflow files are modified (read-only analysis)
    - Generates named report: edit_report_Alice_<folder>_<date>.json
    - Useful for comparing how different evaluators edited the same LLM output
"""

import os
import sys
import json
import shutil
import argparse
import re
from datetime import datetime
from difflib import SequenceMatcher
from rapidfuzz.distance import Levenshtein as LevenshteinDistance
from rapidfuzz import fuzz as rapidfuzz_fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# Add CODE directory to path for imports
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)

from shared_utilities import find_newest_folder


class ArchivistEditsIntegrator:
    """Integrates archivist edits from the HTML review interface back into workflow files."""

    def __init__(self, folder_path, evaluator_name=None, output_folder=None):
        self.folder_path = folder_path
        self.evaluator_name = evaluator_name
        self.analysis_only = evaluator_name is not None  # Auto-enable when evaluator specified
        self.folder_name = os.path.basename(folder_path)
        self.metadata_folder = os.path.join(folder_path, "metadata")
        self.review_folder = os.path.join(folder_path, "review")
        self.exports_folder = os.path.join(self.review_folder, "exports")
        self.original_outputs_folder = os.path.join(folder_path, "original-outputs")

        self.json_folder = os.path.join(self.metadata_folder, "json")
        self.workflow_json_path = os.path.join(self.json_folder, "drawings_workflow.json")

        # In evaluator mode, reports can be written to a separate output folder
        if output_folder and self.analysis_only:
            os.makedirs(output_folder, exist_ok=True)
            self.report_folder = output_folder
        else:
            self.report_folder = self.json_folder

        self.workflow_data = None
        self.decisions_data = None
        self.edit_history = []
        self.stats = {
            'total_records_in_export': 0,
            'records_with_edits': 0,
            'records_reviewed_only': 0,
            'total_field_edits': 0,
            'edits_by_field': {},
            'archivist_name': '',
            'export_timestamp': '',
            'integration_timestamp': '',
            # Detailed character-level metrics for text fields
            'text_field_metrics': {},  # {field_name: {chars_added, chars_deleted, edit_distance_sum, similarity_ratio_sum, token_sort_ratio_sum, edits_count}}
            # Subject metrics (the broad topics AI identified)
            'subjects_total': 0,
            'subjects_accepted': 0,
            'subjects_rejected': 0,
            'subjects_custom_added': 0,  # archivist-added subjects (not headings)
            # Subject heading metrics (controlled vocabulary terms)
            'ai_selected_total': 0,
            'ai_selected_approved': 0,
            'ai_selected_rejected': 0,
            'ai_selected_cascade_rejected': 0,
            'archivist_added_from_list': 0,
            'archivist_added_custom': 0,
            # Format/media term metrics (Getty AAT medium-selected-* terms)
            'medium_selected_total': 0,
            'medium_selected_approved': 0,
            'medium_selected_rejected': 0,
            # Per-record detailed metrics
            'per_record_metrics': [],
            'reviewed_only_metrics': [],
            # List field metrics (for contributors, entities, subjects)
            'list_field_metrics': {}
        }

    def _extract_models_from_logs(self):
        """Extract all models used from token usage logs in the logs folder."""
        logs_folder = os.path.join(self.folder_path, "logs")
        models = set()

        if not os.path.exists(logs_folder):
            return []

        for filename in os.listdir(logs_folder):
            # Only process token usage logs, skip step2 API logs (not LLM)
            if filename.endswith("_token_usage_log.txt") and "step2_vocab_api" not in filename:
                filepath = os.path.join(logs_folder, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        content = f.read()
                    match = re.search(r'Model used:\s*(.+?)(?:\n|$)', content)
                    if match:
                        models.add(match.group(1).strip())
                except Exception:
                    pass

        return sorted(list(models))

    def find_latest_export(self):
        """Find the most recent archivist decisions export JSON."""
        if not os.path.exists(self.exports_folder):
            print(f"Error: Exports folder not found at {self.exports_folder}")
            print("Please run Step 5 and export decisions first.")
            return None

        json_files = [f for f in os.listdir(self.exports_folder) if f.endswith('.json')]

        if not json_files:
            print(f"Error: No JSON export files found in {self.exports_folder}")
            print("Please export your review decisions from the HTML interface first.")
            return None

        # Sort by modification time, newest first
        json_files.sort(key=lambda f: os.path.getmtime(os.path.join(self.exports_folder, f)), reverse=True)

        latest_file = json_files[0]
        return os.path.join(self.exports_folder, latest_file)

    def load_decisions(self, decisions_path):
        """Load archivist decisions from JSON export."""
        try:
            with open(decisions_path, 'r', encoding='utf-8') as f:
                self.decisions_data = json.load(f)

            # Support both old 'cataloger_name' and new 'archivist_name' keys
            self.stats['archivist_name'] = self.decisions_data.get('archivist_name',
                                            self.decisions_data.get('cataloger_name', 'Unknown'))
            self.stats['export_timestamp'] = self.decisions_data.get('export_timestamp', '')
            self.stats['total_records_in_export'] = len(self.decisions_data.get('decisions', []))
            # Also get the total records from the original batch (may be larger if some weren't reviewed)
            self.stats['total_records_in_batch'] = self.decisions_data.get('total_records',
                                                    self.stats['total_records_in_export'])

            print(f"Loaded {self.stats['total_records_in_export']} record decisions")
            print(f"Archivist: {self.stats['archivist_name']}")
            print(f"Export timestamp: {self.stats['export_timestamp']}")

            return True
        except Exception as e:
            print(f"Error loading decisions: {e}")
            return False

    def load_workflow_json(self):
        """Load the workflow JSON data."""
        if not os.path.exists(self.workflow_json_path):
            print(f"Error: Workflow JSON not found at {self.workflow_json_path}")
            return False

        try:
            with open(self.workflow_json_path, 'r', encoding='utf-8') as f:
                self.workflow_data = json.load(f)

            # Handle api_stats entry at the end
            if self.workflow_data and isinstance(self.workflow_data[-1], dict) and 'api_stats' in self.workflow_data[-1]:
                self.api_stats = self.workflow_data[-1]
                self.workflow_data = self.workflow_data[:-1]
            else:
                self.api_stats = None

            print(f"Loaded workflow JSON with {len(self.workflow_data)} records")
            return True
        except Exception as e:
            print(f"Error loading workflow JSON: {e}")
            return False

    def backup_original_files(self):
        """Backup original JSON files to original-outputs folder."""
        os.makedirs(self.original_outputs_folder, exist_ok=True)

        files_to_backup = [
            (self.workflow_json_path, "drawings_workflow.json"),
        ]

        backed_up = []
        for src_path, filename in files_to_backup:
            if os.path.exists(src_path):
                dest_path = os.path.join(self.original_outputs_folder, filename)

                # Only backup if not already backed up (preserve truly original files)
                if not os.path.exists(dest_path):
                    shutil.copy2(src_path, dest_path)
                    backed_up.append(filename)
                    print(f"   Backed up: {filename}")
                else:
                    print(f"   Backup already exists: {filename} (preserving original)")

        return backed_up

    def apply_field_edit(self, record, field_name, edit_data, record_id):
        """Apply a single field edit to a record."""
        analysis = record.get('analysis', {})
        original_value = edit_data.get('original', '')
        new_value = edit_data.get('value', '')

        # Map field names between export and workflow JSON
        field_mapping = {
            'title': 'title',
            'genre': 'genre',
            'description': 'description',
            'format_media': 'format_media',
            'date_on_drawing': 'date_on_drawing',
            'sheet_info': 'sheet_info',
            'content_warning': 'content_warning',
            'contributors': 'contributors',
            'named_entities': 'named_entities',
            'geographic_entities': 'geographic_entities',
            'subjects': 'subjects'
        }

        # Fields that are text (for character-level diff)
        text_fields = {'title', 'genre', 'description', 'format_media', 'medium', 'support',
                       'date_on_drawing', 'sheet_info', 'content_warning'}
        # Fields that are lists
        list_fields = {'contributors', 'named_entities', 'geographic_entities'}

        json_field = field_mapping.get(field_name, field_name)

        if json_field in analysis or json_field in field_mapping.values():
            analysis[json_field] = new_value
            record['analysis'] = analysis

            # Calculate detailed metrics based on field type
            if field_name in text_fields:
                diff_metrics = self._calculate_text_diff_metrics(original_value, new_value)
                # Initialize field metrics if needed
                if field_name not in self.stats['text_field_metrics']:
                    self.stats['text_field_metrics'][field_name] = {
                        'edits_count': 0,
                        'reviews_count': 0,
                        'chars_added': 0,
                        'chars_deleted': 0,
                        'edit_distance_sum': 0,
                        'total_original_length': 0,
                        'total_new_length': 0,
                        'similarity_ratio_sum': 0,
                        'token_sort_ratio_sum': 0,
                    }
                # Accumulate metrics
                self.stats['text_field_metrics'][field_name]['edits_count'] += 1
                self.stats['text_field_metrics'][field_name]['reviews_count'] += 1
                self.stats['text_field_metrics'][field_name]['chars_added'] += diff_metrics['chars_added']
                self.stats['text_field_metrics'][field_name]['chars_deleted'] += diff_metrics['chars_deleted']
                self.stats['text_field_metrics'][field_name]['edit_distance_sum'] += diff_metrics['edit_distance']
                self.stats['text_field_metrics'][field_name]['total_original_length'] += diff_metrics['original_length']
                self.stats['text_field_metrics'][field_name]['total_new_length'] += diff_metrics['new_length']
                self.stats['text_field_metrics'][field_name]['similarity_ratio_sum'] += diff_metrics['similarity_ratio']
                self.stats['text_field_metrics'][field_name]['token_sort_ratio_sum'] += diff_metrics['token_sort_ratio']
            elif field_name in list_fields:
                diff_metrics = self._calculate_list_diff_metrics(original_value, new_value)
                # Initialize list field metrics if needed
                if field_name not in self.stats['list_field_metrics']:
                    self.stats['list_field_metrics'][field_name] = {
                        'edits_count': 0,
                        'reviews_count': 0,
                        'items_added': 0,
                        'items_removed': 0,
                        'items_kept': 0,
                        'items_corrected': 0,
                        'items_reviewed_original': 0,
                    }
                # Accumulate metrics
                self.stats['list_field_metrics'][field_name]['edits_count'] += 1
                self.stats['list_field_metrics'][field_name]['reviews_count'] += 1
                self.stats['list_field_metrics'][field_name]['items_added'] += diff_metrics['items_added']
                self.stats['list_field_metrics'][field_name]['items_removed'] += diff_metrics['items_removed']
                self.stats['list_field_metrics'][field_name]['items_kept'] += diff_metrics['items_kept']
                self.stats['list_field_metrics'][field_name]['items_corrected'] += diff_metrics['items_corrected']
                self.stats['list_field_metrics'][field_name]['items_reviewed_original'] += diff_metrics['original_count']
            else:
                diff_metrics = {}

            # Track the edit with detailed metrics
            self.edit_history.append({
                'record_id': record_id,
                'field': field_name,
                'original_value': original_value,
                'new_value': new_value,
                'edit_type': 'field_edit',
                'diff_metrics': diff_metrics
            })

            # Update stats
            self.stats['total_field_edits'] += 1
            if field_name not in self.stats['edits_by_field']:
                self.stats['edits_by_field'][field_name] = 0
            self.stats['edits_by_field'][field_name] += 1

            return True, diff_metrics
        return False, {}

    def _truncate_for_display(self, value, max_length=100):
        """Truncate a value for display in edit history."""
        if value is None:
            return ""
        if isinstance(value, list):
            value = str(value)
        value = str(value)
        if len(value) > max_length:
            return value[:max_length] + "..."
        return value

    def _resolve_term_label(self, term_id, analysis):
        """Resolve a term ID (e.g. 'subject-4') to a human-readable label."""
        try:
            if term_id.startswith('subject-'):
                idx = int(term_id.split('-', 1)[1])
                subjects = analysis.get('subjects', [])
                if 0 <= idx < len(subjects):
                    return subjects[idx]
            elif term_id.startswith('selected-'):
                idx = int(term_id.split('-', 1)[1])
                terms = analysis.get('final_selected_terms', [])
                if 0 <= idx < len(terms):
                    return terms[idx].get('label', term_id)
            elif term_id.startswith('medium-selected-'):
                idx = int(term_id.split('medium-selected-', 1)[1])
                terms = analysis.get('final_selected_medium_terms', [])
                if 0 <= idx < len(terms):
                    return terms[idx].get('label', term_id)
            elif term_id.startswith('medium-other-'):
                idx = int(term_id.split('medium-other-', 1)[1])
                medium_results = analysis.get('medium_vocabulary_search_results', {})
                other_terms = []
                selected_labels = {t.get('label', '').lower() for t in analysis.get('final_selected_medium_terms', [])}
                for candidates in medium_results.values():
                    for t in candidates:
                        if t.get('label', '').lower() not in selected_labels:
                            other_terms.append(t)
                if 0 <= idx < len(other_terms):
                    return other_terms[idx].get('label', term_id)
            elif term_id.startswith('other-'):
                # Format: other-{source_class}-{i}
                parts = term_id.split('-')
                if len(parts) >= 3:
                    idx = int(parts[-1])
                    source_class = '-'.join(parts[1:-1])
                    # Map short source class to full source name (matches _get_approved_other_terms)
                    _source_map = {'lcsh': 'LCSH', 'fast': 'FAST', 'aat': 'Getty AAT', 'tgn': 'Getty TGN'}
                    source_class_full = _source_map.get(source_class, source_class)
                    vocab_results = analysis.get('vocabulary_search_results', {})
                    selected_labels = {t.get('label', '').lower() for t in analysis.get('final_selected_terms', [])}
                    other_terms = []
                    for topic, candidates in vocab_results.items():
                        for t in candidates:
                            if t.get('label', '').lower() not in selected_labels and t.get('source', '') == source_class_full:
                                other_terms.append(t)
                    if 0 <= idx < len(other_terms):
                        return other_terms[idx].get('label', term_id)
        except (ValueError, IndexError, AttributeError):
            pass
        return term_id

    def _calculate_text_diff_metrics(self, original, new_value):
        """Calculate character-level difference metrics between two strings.

        Returns a dict with:
        - chars_added: characters in new_value not in original (pure insertions)
        - chars_deleted: characters in original not in new_value (pure deletions)
        - edit_distance: Levenshtein distance (minimum single-char edits to transform original → new)
        - pct_changed: edit_distance / max(original_length, new_length) * 100 — how different the result is
        - edit_effort_ratio: (chars_added + chars_deleted) / max(original_length, new_length) — how much work was done
        - original_length: length of original string
        - new_length: length of new string
        - similarity_ratio: character-level similarity 0-100 (inverse of edit distance)
        - token_sort_ratio: similarity 0-100 after sorting words alphabetically (catches reordering)
        """
        # Handle None values
        original = str(original) if original else ""
        new_value = str(new_value) if new_value else ""

        # Use SequenceMatcher for chars_added / chars_deleted
        matcher = SequenceMatcher(None, original, new_value)
        chars_added = 0
        chars_deleted = 0
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'delete':
                chars_deleted += (i2 - i1)
            elif tag == 'insert':
                chars_added += (j2 - j1)
            elif tag == 'replace':
                chars_deleted += (i2 - i1)
                chars_added += (j2 - j1)

        # Levenshtein distance: substitutions count as 1 (not 2), giving a clean "chars changed" number
        edit_distance = LevenshteinDistance.distance(original, new_value)
        original_length = len(original)
        new_length = len(new_value)
        pct_changed = round(edit_distance / max(original_length, new_length, 1) * 100, 1)
        edit_effort_ratio = round((chars_added + chars_deleted) / max(original_length, new_length, 1), 4)

        similarity_ratio = round(rapidfuzz_fuzz.ratio(original, new_value), 1)
        token_sort_ratio = round(rapidfuzz_fuzz.token_sort_ratio(original, new_value), 1)

        return {
            'chars_added': chars_added,
            'chars_deleted': chars_deleted,
            'original_length': original_length,
            'new_length': new_length,
            'edit_distance': edit_distance,
            'pct_changed': pct_changed,
            'edit_effort_ratio': edit_effort_ratio,
            'similarity_ratio': similarity_ratio,
            'token_sort_ratio': token_sort_ratio,
        }

    def _calculate_list_diff_metrics(self, original_list, new_list):
        """Calculate metrics for list-type fields (entities, contributors, etc.).

        Uses exact matching first, then fuzzy matching to identify items the archivist
        corrected (e.g. punctuation or spelling fixes) vs. truly added or removed.

        Returns a dict with:
        - items_kept: identical in both lists
        - items_corrected: similar enough to be considered the same entity (>= 85% match)
        - items_removed: in original but no close match in edited list
        - items_added: in edited list but no close match in original
        - corrected_pairs: list of (original, edited, score) tuples
        - added_items: list of truly new item values
        - removed_items: list of truly removed item values
        """
        FUZZY_THRESHOLD = 85

        # Normalize to lists
        if not isinstance(original_list, list):
            original_list = [original_list] if original_list else []
        if not isinstance(new_list, list):
            new_list = [new_list] if new_list else []

        def normalize_item(item):
            if isinstance(item, dict):
                return json.dumps(item, sort_keys=True)
            return str(item)

        original_norm = [normalize_item(i) for i in original_list if i]
        new_norm = [normalize_item(i) for i in new_list if i]

        original_set = set(original_norm)
        new_set = set(new_norm)

        # Exact matches
        kept = original_set & new_set
        unmatched_removed = list(original_set - new_set)
        unmatched_added = list(new_set - original_set)

        # Fuzzy matching on unmatched items
        corrected_pairs = []
        remaining_removed = list(unmatched_removed)
        remaining_added = list(unmatched_added)

        for orig_item in unmatched_removed:
            best_score = 0
            best_match = None
            for edited_item in remaining_added:
                score = rapidfuzz_fuzz.ratio(orig_item, edited_item)
                if score > best_score:
                    best_score = score
                    best_match = edited_item
            if best_score >= FUZZY_THRESHOLD and best_match is not None:
                corrected_pairs.append((orig_item, best_match, round(best_score, 1)))
                remaining_removed.remove(orig_item)
                remaining_added.remove(best_match)

        return {
            'items_kept': len(kept),
            'items_corrected': len(corrected_pairs),
            'items_removed': len(remaining_removed),
            'items_added': len(remaining_added),
            'original_count': len(original_set),
            'new_count': len(new_set),
            'corrected_pairs': corrected_pairs,
            'added_items': remaining_added,
            'removed_items': remaining_removed,
        }

    def apply_term_decisions(self, record, term_decisions, record_id):
        """Apply vocabulary term decisions to a record.

        Handles three types of decisions:
        - 'approved': Term is approved
        - 'rejected': Term is explicitly rejected
        - dict with 'status': 'cascade_rejected': Term auto-rejected due to parent subject rejection
        """
        analysis = record.get('analysis', {})

        # Track term decisions by type (based on term_id prefix)
        for term_id, decision in term_decisions.items():
            # Determine the status from either string or dict format
            if isinstance(decision, dict):
                status = decision.get('status', '')
                cascade_from = decision.get('cascadeFrom', '')
            else:
                status = decision
                cascade_from = ''

            # Categorize by term type and status
            if term_id.startswith('subject-'):
                # Subject decisions (the broad topics)
                if status == 'rejected':
                    self.stats['subjects_rejected'] += 1
                # Note: subjects_accepted is calculated later from total - rejected
            elif term_id.startswith('selected-'):
                # AI-selected subject headings
                if status == 'cascade_rejected':
                    self.stats['ai_selected_cascade_rejected'] += 1
                elif status == 'rejected':
                    self.stats['ai_selected_rejected'] += 1
                # Note: ai_selected_approved is calculated later
            elif term_id.startswith('other-'):
                # Archivist added from vocabulary list (not AI-selected)
                if status == 'approved':
                    self.stats['archivist_added_from_list'] += 1
            elif term_id.startswith('medium-selected-'):
                # AI-selected format/media terms (Getty AAT) — track rejections
                if status == 'rejected':
                    self.stats['medium_selected_rejected'] += 1
            elif term_id.startswith('medium-other-'):
                # Archivist opted in from format/media list
                if status == 'approved':
                    self.stats['archivist_added_from_list'] += 1

            # Log to edit history.
            # Skip default-approved terms that ended up approved — those are no-ops
            # (archivist rejected then restored, net result = no change from default).
            # Default-approved groups: subject-*, selected-*, medium-selected-*
            # Opt-in groups (other-*, medium-other-*) should always be logged when approved.
            DEFAULT_APPROVED_PREFIXES = ('subject-', 'selected-', 'medium-selected-')
            is_no_op_approval = (status == 'approved' and term_id.startswith(DEFAULT_APPROVED_PREFIXES))

            if not is_no_op_approval:
                term_label = self._resolve_term_label(term_id, analysis)
                if status == 'cascade_rejected':
                    self.edit_history.append({
                        'record_id': record_id,
                        'field': 'vocabulary_term',
                        'term_id': term_id,
                        'original_value': term_label,
                        'new_value': f'cascade_rejected (from: {self._resolve_term_label(cascade_from, analysis) if cascade_from else ""})',
                        'edit_type': 'term_cascade_rejected'
                    })
                else:
                    self.edit_history.append({
                        'record_id': record_id,
                        'field': 'vocabulary_term',
                        'term_id': term_id,
                        'original_value': term_label,
                        'new_value': status,
                        'edit_type': 'term_decision'
                    })

        # Store term decisions in the record for reference
        if 'archivist_term_decisions' not in analysis:
            analysis['archivist_term_decisions'] = {}
        analysis['archivist_term_decisions'] = term_decisions
        record['analysis'] = analysis

        return True

    def apply_custom_terms(self, record, custom_terms, custom_subjects, record_id):
        """Apply custom terms and subjects added by archivist."""
        analysis = record.get('analysis', {})

        # Add custom terms to a dedicated field
        if custom_terms:
            analysis['archivist_custom_terms'] = custom_terms
            self.stats['archivist_added_custom'] += len(custom_terms)

            for term in custom_terms:
                self.edit_history.append({
                    'record_id': record_id,
                    'field': 'custom_term',
                    'original_value': '',
                    'new_value': f"{term.get('label', '')} ({term.get('source', '')})",
                    'edit_type': 'custom_term_added'
                })

        # Add custom subjects to the subjects list
        if custom_subjects:
            existing_subjects = analysis.get('subjects', [])
            if not isinstance(existing_subjects, list):
                existing_subjects = [existing_subjects] if existing_subjects else []

            for subj in custom_subjects:
                label = subj.get('label', '')
                if label and label not in existing_subjects:
                    existing_subjects.append(label)
                    self.stats['subjects_custom_added'] += 1

                    self.edit_history.append({
                        'record_id': record_id,
                        'field': 'custom_subject',
                        'original_value': '',
                        'new_value': label,
                        'edit_type': 'custom_subject_added'
                    })

            analysis['subjects'] = existing_subjects

        record['analysis'] = analysis
        return True

    def apply_all_edits(self):
        """Apply all archivist edits to the workflow data."""
        if not self.decisions_data or not self.workflow_data:
            return False

        decisions = self.decisions_data.get('decisions', [])

        for decision in decisions:
            record_id = decision.get('record_id')

            # Find the corresponding record in workflow data (1-indexed to 0-indexed)
            record_idx = record_id - 1
            if record_idx < 0 or record_idx >= len(self.workflow_data):
                print(f"   Warning: Record ID {record_id} not found in workflow data")
                continue

            record = self.workflow_data[record_idx]
            has_edits = False

            # Track per-record metrics
            record_metrics = {
                'record_id': record_id,
                'archivist_notes': decision.get('archivist_notes',
                                   decision.get('cataloger_notes', '')),
                'field_edits': {},
                'term_decisions': {},
                'custom_additions': {}
            }

            # Apply field edits
            edits = decision.get('edits', {})
            if edits:
                for field_name, edit_data in edits.items():
                    if edit_data.get('edited', False):
                        success, diff_metrics = self.apply_field_edit(record, field_name, edit_data, record_id)
                        if success:
                            has_edits = True
                            record_metrics['field_edits'][field_name] = diff_metrics

            # Accumulate lengths for unedited text fields in this reviewed record.
            # This ensures pct_changed reflects "fraction of all reviewed text that needed
            # correction" rather than only measuring within the edited subset.
            _text_fields = {'title', 'genre', 'description', 'format_media', 'medium', 'support',
                            'date_on_drawing', 'sheet_info', 'content_warning'}
            _edited_text_fields = {
                fn for fn, ed in (edits or {}).items()
                if fn in _text_fields and ed.get('edited', False)
            }
            _analysis = record.get('analysis', {})
            for _field in _text_fields:
                if _field not in _edited_text_fields:
                    _val = _analysis.get(_field, '') or ''
                    if not isinstance(_val, str):
                        continue
                    _len = len(_val)
                    if _field not in self.stats['text_field_metrics']:
                        self.stats['text_field_metrics'][_field] = {
                            'edits_count': 0,
                            'reviews_count': 0,
                            'chars_added': 0,
                            'chars_deleted': 0,
                            'edit_distance_sum': 0,
                            'total_original_length': 0,
                            'total_new_length': 0,
                            'similarity_ratio_sum': 0,
                            'token_sort_ratio_sum': 0,
                        }
                    self.stats['text_field_metrics'][_field]['reviews_count'] += 1
                    self.stats['text_field_metrics'][_field]['total_original_length'] += _len
                    self.stats['text_field_metrics'][_field]['total_new_length'] += _len

            # Accumulate item counts for unedited list fields in this reviewed record.
            # Unedited = archivist approved all items, so original count = items_reviewed_original,
            # items_kept = original count, items_removed/added = 0.
            _list_fields = {'contributors', 'named_entities', 'geographic_entities'}
            _edited_list_fields = {
                fn for fn, ed in (edits or {}).items()
                if fn in _list_fields and ed.get('edited', False)
            }
            for _field in _list_fields:
                if _field not in _edited_list_fields:
                    _val = _analysis.get(_field, []) or []
                    if not isinstance(_val, list):
                        continue
                    _count = len(_val)
                    if _field not in self.stats['list_field_metrics']:
                        self.stats['list_field_metrics'][_field] = {
                            'edits_count': 0,
                            'reviews_count': 0,
                            'items_added': 0,
                            'items_removed': 0,
                            'items_kept': 0,
                            'items_corrected': 0,
                            'items_reviewed_original': 0,
                        }
                    self.stats['list_field_metrics'][_field]['reviews_count'] += 1
                    self.stats['list_field_metrics'][_field]['items_kept'] += _count
                    self.stats['list_field_metrics'][_field]['items_reviewed_original'] += _count

            # Apply term decisions
            term_decisions = decision.get('term_decisions', {})
            if term_decisions:
                self.apply_term_decisions(record, term_decisions, record_id)
                has_edits = True
            # Always calculate term metrics so totals are correct even with no decisions
            record_metrics['term_decisions'] = self._calculate_record_term_metrics(
                term_decisions, record
            )

            # Apply custom terms and subjects
            custom_terms = decision.get('custom_terms', [])
            custom_subjects = decision.get('custom_subjects', [])
            if custom_terms or custom_subjects:
                self.apply_custom_terms(record, custom_terms, custom_subjects, record_id)
                has_edits = True
                # Add custom additions to term_decisions metrics
                if 'term_decisions' not in record_metrics:
                    record_metrics['term_decisions'] = {}
                record_metrics['term_decisions']['archivist_added_custom'] = len(custom_terms)
                record_metrics['term_decisions']['subjects_custom_added'] = len(custom_subjects)

            # Add archivist metadata
            analysis = record.get('analysis', {})
            analysis['archivist_reviewed'] = decision.get('reviewed', False)
            # Support both old 'cataloger_notes' and new 'archivist_notes' keys
            analysis['archivist_notes'] = decision.get('archivist_notes',
                                           decision.get('cataloger_notes', ''))
            analysis['archivist_review_date'] = self.stats['export_timestamp']
            analysis['archivist_name'] = self.stats['archivist_name']
            record['analysis'] = analysis

            # Update stats
            if has_edits:
                self.stats['records_with_edits'] += 1
                self.stats['per_record_metrics'].append(record_metrics)
            elif decision.get('reviewed', False):
                self.stats['records_reviewed_only'] += 1
                self.stats['reviewed_only_metrics'].append({
                    'record_id': record_id,
                    'archivist_notes': decision.get('archivist_notes', '')
                })

        # Calculate aggregate subject/heading metrics across all records
        self._calculate_aggregate_subject_heading_metrics()

        return True

    def _calculate_record_term_metrics(self, term_decisions, record):
        """Calculate term metrics for a single record, organized by type."""
        analysis = record.get('analysis', {})

        # Get totals from the record data
        subjects = analysis.get('subjects', [])
        if not isinstance(subjects, list):
            subjects = [subjects] if subjects else []
        subjects_total = len(subjects)

        ai_selected = analysis.get('final_selected_terms', [])
        ai_selected_total = len(ai_selected)

        # Count decisions by type
        subjects_rejected = 0
        ai_rejected = 0
        ai_cascade_rejected = 0
        archivist_from_list = 0

        for term_id, decision in term_decisions.items():
            status = decision.get('status', decision) if isinstance(decision, dict) else decision

            if term_id.startswith('subject-'):
                if status == 'rejected':
                    subjects_rejected += 1
            elif term_id.startswith('selected-'):
                if status == 'cascade_rejected':
                    ai_cascade_rejected += 1
                elif status == 'rejected':
                    ai_rejected += 1
            elif term_id.startswith('other-'):
                if status == 'approved':
                    archivist_from_list += 1

        return {
            'subjects': {
                'total': subjects_total,
                'accepted': subjects_total - subjects_rejected,
                'rejected': subjects_rejected
            },
            'subject_headings': {
                'ai_selected': {
                    'total': ai_selected_total,
                    'approved': ai_selected_total - ai_rejected - ai_cascade_rejected,
                    'rejected': ai_rejected,
                    'cascade_rejected': ai_cascade_rejected
                },
                'archivist_added_from_list': archivist_from_list
            }
        }

    def _calculate_aggregate_subject_heading_metrics(self):
        """Calculate aggregate metrics for subjects and headings across all records."""
        for record in self.workflow_data:
            analysis = record.get('analysis', {})

            # Count subjects (the broad topics AI identified)
            original_subjects = analysis.get('subjects', [])
            if not isinstance(original_subjects, list):
                original_subjects = [original_subjects] if original_subjects else []
            self.stats['subjects_total'] += len(original_subjects)

            # Count AI-selected subject headings (controlled vocabulary terms)
            final_selected_terms = analysis.get('final_selected_terms', [])
            self.stats['ai_selected_total'] += len(final_selected_terms)

            # Count AI-selected format/media terms (Getty AAT)
            medium_selected_terms = analysis.get('final_selected_medium_terms', [])
            if isinstance(medium_selected_terms, list):
                self.stats['medium_selected_total'] += len(medium_selected_terms)

        # Calculate derived stats
        # subjects_accepted = total - rejected (rejected is counted in apply_term_decisions)
        self.stats['subjects_accepted'] = self.stats['subjects_total'] - self.stats['subjects_rejected']

        # ai_selected_approved = total - rejected - cascade_rejected
        self.stats['ai_selected_approved'] = (
            self.stats['ai_selected_total'] -
            self.stats['ai_selected_rejected'] -
            self.stats['ai_selected_cascade_rejected']
        )

        # medium_selected_approved = total - rejected
        self.stats['medium_selected_approved'] = (
            self.stats['medium_selected_total'] - self.stats['medium_selected_rejected']
        )

    def save_workflow_json(self):
        """Save the updated workflow JSON."""
        try:
            # Re-add api_stats if it existed
            output_data = self.workflow_data.copy()
            if self.api_stats:
                output_data.append(self.api_stats)

            with open(self.workflow_json_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, indent=2, ensure_ascii=False)

            print(f"   Saved updated workflow JSON")
            return True
        except Exception as e:
            print(f"Error saving workflow JSON: {e}")
            return False

    def create_final_excel_deliverable(self):
        """Create final Excel deliverable with 2 sheets: Final Metadata, Edit Statistics."""
        try:
            wb = Workbook()

            # Styles
            header_fill = PatternFill(start_color="FF2C3E50", end_color="FF2C3E50", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            stats_fill = PatternFill(start_color="FFE8F4F8", end_color="FFE8F4F8", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            wrap_alignment = Alignment(wrap_text=True, vertical='top')

            # ============ SHEET 1: Final Metadata ============
            ws1 = wb.active
            ws1.title = "Final Metadata"

            # Headers for Final Metadata - matching step-4 structure
            metadata_headers = [
                "Folder", "Filename", "Title", "Contributors", "Genre",
                "Description", "Format/Media", "Format/Media Terms (Getty AAT)", "Date on Drawing", "Sheet Info",
                "Subjects (AI)", "Subject Headings (Controlled Vocab)",
                "Named Entities", "Geographic Entities", "Content Warning",
                "Reviewed", "Archivist", "Review Date", "Notes"
            ]

            for col, header in enumerate(metadata_headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Load final metadata JSON for this sheet
            final_metadata_path = os.path.join(self.json_folder, "final_metadata.json")
            if os.path.exists(final_metadata_path):
                with open(final_metadata_path, 'r', encoding='utf-8') as f:
                    final_data = json.load(f)

                records = final_data.get('records', [])
                for row_idx, record in enumerate(records, 2):
                    metadata = record.get('metadata', {})
                    review_info = record.get('review_info', {})

                    # Get filename from image_path
                    image_path = record.get('image_path', '')
                    filename = os.path.basename(image_path) if image_path else ''

                    # Format contributors
                    contributors = metadata.get('contributors', [])
                    if isinstance(contributors, list):
                        contrib_strs = []
                        for c in contributors:
                            if isinstance(c, dict):
                                name = c.get('name', '')
                                role = c.get('role', '')
                                contrib_strs.append(f"{name} ({role})" if role else name)
                            else:
                                contrib_strs.append(str(c))
                        contributors_str = '; '.join(contrib_strs)
                    else:
                        contributors_str = str(contributors) if contributors else ''

                    # Format subject headings as "label [source] (uri)"
                    subject_headings = metadata.get('subject_headings', [])
                    heading_strs = []
                    for h in subject_headings:
                        if isinstance(h, dict):
                            label = h.get('label', '')
                            source = h.get('source', '')
                            uri = h.get('uri', '')
                            if label:
                                heading_strs.append(f"{label} [{source}] ({uri})")
                        else:
                            heading_strs.append(str(h))
                    headings_str = '; '.join(heading_strs)

                    # Format list fields
                    def format_list(lst):
                        if isinstance(lst, list):
                            return '; '.join(str(item) for item in lst if item)
                        return str(lst) if lst else ''

                    # Format medium terms as semicolon-separated labels with URIs
                    medium_terms_list = metadata.get('format_media_terms', [])
                    medium_terms_str = '; '.join(
                        f"{t.get('label', '')} ({t.get('uri', '')})" if t.get('uri') else t.get('label', '')
                        for t in medium_terms_list if t.get('label')
                    )

                    row_data = [
                        record.get('folder', ''),
                        filename,
                        metadata.get('title', ''),
                        contributors_str,
                        metadata.get('genre', ''),
                        metadata.get('description', ''),
                        metadata.get('format_media', ''),
                        medium_terms_str,
                        metadata.get('date_on_drawing', ''),
                        metadata.get('sheet_info', ''),
                        format_list(metadata.get('subjects', [])),
                        headings_str,
                        format_list(metadata.get('named_entities', [])),
                        format_list(metadata.get('geographic_entities', [])),
                        metadata.get('content_warning', ''),
                        'Yes' if review_info.get('reviewed') else 'No',
                        review_info.get('archivist_name', ''),
                        review_info.get('review_date', ''),
                        review_info.get('archivist_notes', '')
                    ]

                    for col, value in enumerate(row_data, 1):
                        cell = ws1.cell(row=row_idx, column=col, value=value)
                        cell.border = border
                        cell.alignment = wrap_alignment

            # Adjust column widths for Final Metadata
            col_widths = [20, 25, 30, 35, 25, 60, 40, 45, 15, 40, 40, 60, 40, 30, 15, 10, 20, 20, 30]
            for col_idx, width in enumerate(col_widths, 1):
                ws1.column_dimensions[ws1.cell(row=1, column=col_idx).column_letter].width = width

            # Freeze header row
            ws1.freeze_panes = 'A2'

            # ============ SHEET 2: Edit Statistics ============
            ws2 = wb.create_sheet("Edit Statistics")

            self.stats['integration_timestamp'] = datetime.now().isoformat()

            # Summary section at top
            ws2['A1'] = "ARCHIVIST EDIT INTEGRATION SUMMARY"
            ws2['A1'].font = Font(size=14, bold=True)
            ws2.merge_cells('A1:C1')

            # Extract models from logs
            models_used = self._extract_models_from_logs()

            summary_data = [
                ("Models Used:", ', '.join(models_used) if models_used else 'Unknown'),
                ("Archivist Name:", self.stats['archivist_name']),
                ("Export Timestamp:", self.stats['export_timestamp']),
                ("Integration Timestamp:", self.stats['integration_timestamp']),
                ("", ""),
                ("Total Records in Batch:", self.stats.get('total_records_in_batch', self.stats['total_records_in_export'])),
                ("Records Reviewed:", self.stats['total_records_in_export']),
                ("Records with Edits:", self.stats['records_with_edits']),
                ("Records Reviewed Only (no edits):", self.stats['records_reviewed_only']),
                ("", ""),
                ("Total Field Edits:", self.stats['total_field_edits']),
            ]

            row = 3
            for label, value in summary_data:
                cell1 = ws2.cell(row=row, column=1, value=label)
                cell1.font = Font(bold=True)
                ws2.cell(row=row, column=2, value=value)
                if label:
                    cell1.fill = stats_fill
                    ws2.cell(row=row, column=2).fill = stats_fill
                row += 1

            # Edits by Field Section
            row += 1
            ws2.cell(row=row, column=1, value="EDITS BY FIELD").font = Font(bold=True, size=12)
            row += 1
            for field, count in sorted(self.stats['edits_by_field'].items()):
                ws2.cell(row=row, column=1, value=f"  {field}:")
                ws2.cell(row=row, column=2, value=count)
                row += 1

            # Subject Acceptance Rates
            row += 1
            ws2.cell(row=row, column=1, value="SUBJECTS (AI-identified topics)").font = Font(bold=True, size=12)
            row += 1
            subjects_total = self.stats['subjects_total']
            subjects_accepted = self.stats['subjects_accepted']
            subjects_rejected = self.stats['subjects_rejected']
            acceptance_pct = (subjects_accepted / subjects_total * 100) if subjects_total > 0 else 100.0

            ws2.cell(row=row, column=1, value="  Total:").fill = stats_fill
            ws2.cell(row=row, column=2, value=subjects_total).fill = stats_fill
            row += 1
            ws2.cell(row=row, column=1, value="  Accepted:").fill = stats_fill
            ws2.cell(row=row, column=2, value=f"{subjects_accepted} ({acceptance_pct:.1f}%)").fill = stats_fill
            row += 1
            ws2.cell(row=row, column=1, value="  Rejected:").fill = stats_fill
            ws2.cell(row=row, column=2, value=subjects_rejected).fill = stats_fill
            row += 1
            if self.stats['subjects_custom_added'] > 0:
                ws2.cell(row=row, column=1, value="  Custom Added:").fill = stats_fill
                ws2.cell(row=row, column=2, value=self.stats['subjects_custom_added']).fill = stats_fill
                row += 1

            # Subject Heading Approval Rates
            row += 1
            ws2.cell(row=row, column=1, value="SUBJECT HEADINGS (controlled vocabulary)").font = Font(bold=True, size=12)
            row += 1

            ai_total = self.stats['ai_selected_total']
            ai_approved = self.stats['ai_selected_approved']
            ai_rejected = self.stats['ai_selected_rejected']
            ai_cascade = self.stats['ai_selected_cascade_rejected']
            ai_approval_pct = (ai_approved / ai_total * 100) if ai_total > 0 else 100.0

            ws2.cell(row=row, column=1, value="  AI-Selected Total:").fill = stats_fill
            ws2.cell(row=row, column=2, value=ai_total).fill = stats_fill
            row += 1
            ws2.cell(row=row, column=1, value="  AI-Selected Approved:").fill = stats_fill
            ws2.cell(row=row, column=2, value=f"{ai_approved} ({ai_approval_pct:.1f}%)").fill = stats_fill
            row += 1
            ws2.cell(row=row, column=1, value="  AI-Selected Rejected:").fill = stats_fill
            ws2.cell(row=row, column=2, value=ai_rejected).fill = stats_fill
            row += 1
            ws2.cell(row=row, column=1, value="  AI-Selected Cascade Rejected:").fill = stats_fill
            ws2.cell(row=row, column=2, value=ai_cascade).fill = stats_fill
            row += 1
            ws2.cell(row=row, column=1, value="  Archivist Added from List:").fill = stats_fill
            ws2.cell(row=row, column=2, value=self.stats['archivist_added_from_list']).fill = stats_fill
            row += 1
            ws2.cell(row=row, column=1, value="  Archivist Added Custom:").fill = stats_fill
            ws2.cell(row=row, column=2, value=self.stats['archivist_added_custom']).fill = stats_fill
            row += 1

            # Text field character metrics
            if self.stats['text_field_metrics']:
                row += 1
                ws2.cell(row=row, column=1, value="CHARACTER-LEVEL CHANGES").font = Font(bold=True, size=12)
                row += 1

                for field, metrics in sorted(self.stats['text_field_metrics'].items()):
                    edits = metrics['edits_count']
                    avg_pct = (metrics['edit_distance_sum'] / max(metrics['total_original_length'], metrics['total_new_length'], 1) * 100)
                    avg_sim = round(metrics['similarity_ratio_sum'] / edits, 1) if edits > 0 else 0
                    avg_tsr = round(metrics['token_sort_ratio_sum'] / edits, 1) if edits > 0 else 0
                    ws2.cell(row=row, column=1, value=f"  {field}:")
                    ws2.cell(row=row, column=2, value=f"{edits} edits, {metrics['edit_distance_sum']} chars changed (Levenshtein), {avg_pct:.1f}% avg change, edited field similarity: {avg_sim}/100, edited field token sort: {avg_tsr}/100")
                    row += 1

            # Adjust column widths for Edit Statistics
            ws2.column_dimensions['A'].width = 35
            ws2.column_dimensions['B'].width = 50
            ws2.column_dimensions['C'].width = 30

            # Save the workbook
            excel_path = os.path.join(self.metadata_folder, "final_deliverable.xlsx")
            wb.save(excel_path)
            print(f"   Created final_deliverable.xlsx with 2 sheets")
            return excel_path

        except Exception as e:
            print(f"Error creating final Excel deliverable: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _calculate_quality_scores(self):
        """Calculate aggregate quality scores for LLM comparison.

        Returns human-readable format with both fraction strings and percentages.
        """
        # Edit rate: proportion of records that needed edits
        total_records = self.stats['total_records_in_export']
        records_edited = self.stats['records_with_edits']
        edit_rate_pct = (records_edited / total_records * 100) if total_records > 0 else 0

        # Text field edit rate: proportion of individual text fields that were edited
        fields_reviewed = sum(m['reviews_count'] for m in self.stats['text_field_metrics'].values())
        fields_edited = sum(m['edits_count'] for m in self.stats['text_field_metrics'].values())
        text_field_edit_rate_pct = (fields_edited / fields_reviewed * 100) if fields_reviewed > 0 else 0

        # List field edit rate: proportion of individual list fields that were edited
        list_fields_reviewed = sum(m['reviews_count'] for m in self.stats['list_field_metrics'].values())
        list_fields_edited = sum(m['edits_count'] for m in self.stats['list_field_metrics'].values())
        list_field_edit_rate_pct = (list_fields_edited / list_fields_reviewed * 100) if list_fields_reviewed > 0 else 0

        # Subject acceptance rate
        subjects_total = self.stats['subjects_total']
        subjects_accepted = self.stats['subjects_accepted']
        subject_acceptance_pct = (subjects_accepted / subjects_total * 100) if subjects_total > 0 else 100.0

        # AI-selected heading approval rate
        ai_total = self.stats['ai_selected_total']
        ai_approved = self.stats['ai_selected_approved']
        ai_approval_pct = (ai_approved / ai_total * 100) if ai_total > 0 else 100.0

        # Overall text field pct_changed: edit_distance / max(orig, new) across all text fields
        total_edit_distance = sum(m['edit_distance_sum'] for m in self.stats['text_field_metrics'].values())
        total_orig_len = sum(m['total_original_length'] for m in self.stats['text_field_metrics'].values())
        total_new_len = sum(m['total_new_length'] for m in self.stats['text_field_metrics'].values())
        max_len = max(total_orig_len, total_new_len, 1)
        text_pct_changed = round(total_edit_distance / max_len * 100, 1)

        # List field item retention: of all reviewed AI-generated list items, what fraction survived
        list_items_kept = sum(m['items_kept'] for m in self.stats['list_field_metrics'].values())
        list_items_corrected = sum(m['items_corrected'] for m in self.stats['list_field_metrics'].values())
        list_items_removed = sum(m['items_removed'] for m in self.stats['list_field_metrics'].values())
        list_items_added = sum(m['items_added'] for m in self.stats['list_field_metrics'].values())
        list_items_original = sum(m['items_reviewed_original'] for m in self.stats['list_field_metrics'].values())
        list_net_change = list_items_added - list_items_removed
        list_retained = list_items_kept + list_items_corrected
        list_retention_pct = round(list_retained / list_items_original * 100, 1) if list_items_original > 0 else 100.0

        return {
            'edit_rate': {
                'display': f"{records_edited}/{total_records} items edited = {edit_rate_pct:.1f}%",
                'value': round(edit_rate_pct, 2)
            },
            'text_field_edit_rate': {
                'display': f"{fields_edited}/{fields_reviewed} fields edited = {text_field_edit_rate_pct:.1f}%",
                'value': round(text_field_edit_rate_pct, 2)
            },
            'list_field_edit_rate': {
                'display': f"{list_fields_edited}/{list_fields_reviewed} fields edited = {list_field_edit_rate_pct:.1f}%",
                'value': round(list_field_edit_rate_pct, 2)
            },
            'text_pct_changed': {
                'display': f"{total_edit_distance} edit_distance / {max_len} max_length = {text_pct_changed}%",
                'value': text_pct_changed
            },
            'list_item_retention_rate': {
                'display': f"{list_retained}/{list_items_original} retained ({list_items_kept} kept + {list_items_corrected} corrected) = {list_retention_pct}%",
                'value': list_retention_pct
            },
            'list_items_net_change': {
                'display': f"{list_items_kept} kept, {list_items_corrected} corrected, {list_items_removed} removed, {list_items_added} added (net {list_net_change:+d})",
                'value': list_net_change
            },
            'subject_acceptance_rate': {
                'display': f"{subjects_accepted}/{subjects_total} accepted = {subject_acceptance_pct:.1f}%",
                'value': round(subject_acceptance_pct, 2)
            },
            'ai_heading_approval_rate': {
                'display': f"{ai_approved}/{ai_total} approved = {ai_approval_pct:.2f}%",
                'value': round(ai_approval_pct, 2)
            },
            'format_media_approval_rate': {
                'display': f"{self.stats['medium_selected_approved']}/{self.stats['medium_selected_total']} approved = {(self.stats['medium_selected_approved'] / self.stats['medium_selected_total'] * 100) if self.stats['medium_selected_total'] > 0 else 100.0:.2f}%",
                'value': round((self.stats['medium_selected_approved'] / self.stats['medium_selected_total'] * 100) if self.stats['medium_selected_total'] > 0 else 100.0, 2)
            }
        }

    def _build_per_record_summary(self):
        """Build concise per-record metrics for the report with readable percentages."""
        records = []
        for record_metrics in self.stats['per_record_metrics']:
            # Get term decisions for this record
            term_decisions = record_metrics.get('term_decisions', {})

            # Calculate percentages for subjects
            subjects_info = term_decisions.get('subjects', {})
            subj_total = subjects_info.get('total', 0)
            subj_accepted = subjects_info.get('accepted', 0)
            subj_rejected = subjects_info.get('rejected', 0)
            subj_pct = (subj_accepted / subj_total * 100) if subj_total > 0 else 100.0

            # Calculate percentages for subject headings
            headings_info = term_decisions.get('subject_headings', {})
            ai_info = headings_info.get('ai_selected', {})
            ai_total = ai_info.get('total', 0)
            ai_approved = ai_info.get('approved', 0)
            ai_rejected = ai_info.get('rejected', 0)
            ai_cascade = ai_info.get('cascade_rejected', 0)
            ai_pct = (ai_approved / ai_total * 100) if ai_total > 0 else 100.0
            archivist_added = headings_info.get('archivist_added_from_list', 0)

            record = {
                'id': record_metrics['record_id'],
                'archivist_notes': record_metrics.get('archivist_notes', ''),
                'text_fields': {},
                'list_fields': {},
                'terms': {
                    'subjects': {
                        'summary': f"{subj_accepted}/{subj_total} accepted = {subj_pct:.0f}%",
                        'total': subj_total,
                        'accepted': subj_accepted,
                        'rejected': subj_rejected
                    },
                    'subject_headings': {
                        'ai_selected': {
                            'summary': f"{ai_approved}/{ai_total} approved = {ai_pct:.0f}%",
                            'total': ai_total,
                            'approved': ai_approved,
                            'rejected': ai_rejected,
                            'cascade_rejected': ai_cascade
                        },
                        'archivist_added_from_list': archivist_added
                    }
                }
            }

            # Process field edits
            for field_name, metrics in record_metrics.get('field_edits', {}).items():
                if 'edit_distance' in metrics:
                    # Text field
                    edit_distance = metrics.get('edit_distance', 0)
                    pct_changed = metrics.get('pct_changed', 0.0)
                    orig_len = metrics.get('original_length', 0)
                    new_len = metrics.get('new_length', 0)
                    chars_added = metrics.get('chars_added', 0)
                    chars_deleted = metrics.get('chars_deleted', 0)
                    record['text_fields'][field_name] = {
                        'chars_added': chars_added,
                        'chars_deleted': chars_deleted,
                        'original_length': orig_len,
                        'new_length': new_len,
                        'edit_distance': edit_distance,
                        'pct_changed': round(pct_changed, 1),
                        'edit_effort_ratio': round((chars_added + chars_deleted) / max(orig_len, new_len, 1), 4),
                        'similarity_ratio': metrics.get('similarity_ratio', 0),
                        'token_sort_ratio': metrics.get('token_sort_ratio', 0),
                    }
                elif 'items_added' in metrics:
                    # List field
                    record['list_fields'][field_name] = {
                        'kept': metrics.get('items_kept', 0),
                        'corrected': metrics.get('items_corrected', 0),
                        'removed': metrics.get('items_removed', 0),
                        'added': metrics.get('items_added', 0),
                    }

            records.append(record)

        # Include reviewed-only records (no changes made, all terms implicitly approved)
        for rm in self.stats['reviewed_only_metrics']:
            record_id = rm['record_id']
            record_idx = record_id - 1
            analysis = {}
            if 0 <= record_idx < len(self.workflow_data):
                analysis = self.workflow_data[record_idx].get('analysis', {})

            subjects = analysis.get('subjects', [])
            if not isinstance(subjects, list):
                subjects = [subjects] if subjects else []
            subj_total = len(subjects)

            ai_terms = analysis.get('final_selected_terms', [])
            ai_total = len(ai_terms) if isinstance(ai_terms, list) else 0

            records.append({
                'id': record_id,
                'reviewed_no_changes': True,
                'archivist_notes': rm.get('archivist_notes', ''),
                'text_fields': {},
                'list_fields': {},
                'terms': {
                    'subjects': {
                        'summary': f"{subj_total}/{subj_total} accepted = 100%",
                        'total': subj_total,
                        'accepted': subj_total,
                        'rejected': 0
                    },
                    'subject_headings': {
                        'ai_selected': {
                            'summary': f"{ai_total}/{ai_total} approved = 100%",
                            'total': ai_total,
                            'approved': ai_total,
                            'rejected': 0,
                            'cascade_rejected': 0
                        },
                        'archivist_added_from_list': 0
                    }
                }
            })

        records.sort(key=lambda r: r['id'])
        return records

    def generate_edit_statistics_report(self):
        """Generate a combined edit report with statistics and full edit history.

        Creates edit_report.json with:
        - batch_info: Models used, archivist, date, record counts, total field edits
        - average_quality_scores: Top-level metrics for LLM comparison
        - text_fields_summary: Per-field aggregated metrics
        - subjects / subject_headings: Term acceptance summaries
        - records: Per-record metrics with embedded edit history
        """
        self.stats['integration_timestamp'] = datetime.now().isoformat()

        # Extract models from logs
        models_used = self._extract_models_from_logs()

        # Calculate quality scores
        quality_scores = self._calculate_quality_scores()

        # Build text field summary with totals and percentages
        text_fields_summary = {}
        total_chars_added_all = 0
        total_chars_deleted_all = 0
        total_edit_distance_all_fields = 0
        total_edits_count = 0
        total_fields_reviewed = 0
        total_orig_len_all_fields = 0
        total_new_len_all_fields = 0

        for field_name, metrics in self.stats['text_field_metrics'].items():
            edits = metrics['edits_count']
            orig_len = metrics['total_original_length']
            new_len = metrics['total_new_length']
            edit_dist = metrics['edit_distance_sum']
            chars_added = metrics['chars_added']
            chars_deleted = metrics['chars_deleted']
            pct_changed = round(edit_dist / max(orig_len, new_len, 1) * 100, 1)
            edit_effort_ratio = round((chars_added + chars_deleted) / max(orig_len, new_len, 1), 4)

            avg_similarity = round(metrics['similarity_ratio_sum'] / edits, 1) if edits > 0 else 0
            avg_token_sort = round(metrics['token_sort_ratio_sum'] / edits, 1) if edits > 0 else 0

            text_fields_summary[field_name] = {
                'records_edited': edits,
                'records_reviewed': metrics['reviews_count'],
                'chars_added': chars_added,
                'chars_deleted': chars_deleted,
                'original_length': orig_len,
                'new_length': new_len,
                'edit_distance': edit_dist,
                'pct_changed': pct_changed,
                'edit_effort_ratio': edit_effort_ratio,
                'avg_edited_field_similarity_ratio': avg_similarity,
                'avg_edited_field_token_sort_ratio': avg_token_sort,
            }

            total_chars_added_all += chars_added
            total_chars_deleted_all += chars_deleted
            total_edit_distance_all_fields += edit_dist
            total_edits_count += edits
            total_fields_reviewed += metrics['reviews_count']
            total_orig_len_all_fields += orig_len
            total_new_len_all_fields += new_len

        overall_pct_changed = round(total_edit_distance_all_fields / max(total_orig_len_all_fields, total_new_len_all_fields, 1) * 100, 1) if total_edits_count > 0 else 0.0
        overall_edit_effort_ratio = round((total_chars_added_all + total_chars_deleted_all) / max(total_orig_len_all_fields, total_new_len_all_fields, 1), 4) if total_edits_count > 0 else 0.0

        field_edit_rate_pct = round(total_edits_count / total_fields_reviewed * 100, 1) if total_fields_reviewed > 0 else 0.0

        text_fields_totals = {
            'fields_edited': total_edits_count,
            'fields_reviewed': total_fields_reviewed,
            'text_field_edit_rate': f"{total_edits_count}/{total_fields_reviewed} fields edited = {field_edit_rate_pct}%",
            'chars_added': total_chars_added_all,
            'chars_deleted': total_chars_deleted_all,
            'original_length': total_orig_len_all_fields,
            'new_length': total_new_len_all_fields,
            'edit_distance': total_edit_distance_all_fields,
            'pct_changed': overall_pct_changed,
            'edit_effort_ratio': overall_edit_effort_ratio,
        }

        # Build list field summary
        list_fields_summary = {}
        for field_name, metrics in self.stats['list_field_metrics'].items():
            original_count = metrics['items_reviewed_original']
            final_count = original_count - metrics['items_removed'] + metrics['items_added']
            net = metrics['items_added'] - metrics['items_removed']
            retained = metrics['items_kept'] + metrics['items_corrected']
            retention_pct = round(retained / original_count * 100, 1) if original_count > 0 else 100.0
            list_fields_summary[field_name] = {
                'records_edited': metrics['edits_count'],
                'records_reviewed': metrics['reviews_count'],
                'original_count': original_count,
                'final_count': final_count,
                'items_kept': metrics['items_kept'],
                'items_corrected': metrics['items_corrected'],
                'items_removed': metrics['items_removed'],
                'items_added': metrics['items_added'],
                'net_change': net,
                'retention_rate': retention_pct,
            }

        # Subjects summary (broad topics AI identified)
        subjects_total = self.stats['subjects_total']
        subjects_accepted = self.stats['subjects_accepted']
        subjects_rejected = self.stats['subjects_rejected']
        acceptance_pct = (subjects_accepted / subjects_total * 100) if subjects_total > 0 else 100.0

        subjects_summary = {
            'summary': f"{subjects_accepted}/{subjects_total} accepted = {acceptance_pct:.1f}%",
            'total': subjects_total,
            'accepted': subjects_accepted,
            'rejected': subjects_rejected,
            'acceptance_rate': round(acceptance_pct, 2)
        }
        if self.stats['subjects_custom_added'] > 0:
            subjects_summary['custom_added'] = self.stats['subjects_custom_added']

        # Subject headings summary (controlled vocabulary terms)
        ai_total = self.stats['ai_selected_total']
        ai_approved = self.stats['ai_selected_approved']
        ai_rejected = self.stats['ai_selected_rejected']
        ai_cascade = self.stats['ai_selected_cascade_rejected']
        ai_approval_pct = (ai_approved / ai_total * 100) if ai_total > 0 else 100.0
        ai_rejection_pct = ((ai_rejected + ai_cascade) / ai_total * 100) if ai_total > 0 else 0

        archivist_from_list = self.stats['archivist_added_from_list']
        archivist_custom = self.stats['archivist_added_custom']

        subject_headings_summary = {
            'ai_selected': {
                'summary': f"{ai_approved}/{ai_total} approved = {ai_approval_pct:.1f}%",
                'total': ai_total,
                'approved': ai_approved,
                'rejected': ai_rejected,
                'cascade_rejected': ai_cascade,
                'approval_rate': round(ai_approval_pct, 2),
                'rejection_rate': round(ai_rejection_pct, 2)
            },
            'archivist_additions': {
                'from_vocabulary_list': archivist_from_list,
                'custom_terms': archivist_custom,
                'total_added': archivist_from_list + archivist_custom
            }
        }

        # Build batch_info with optional evaluator field
        batch_info = {
            'models_used': models_used,
            'archivist': self.stats['archivist_name'],
            'date': self.stats['integration_timestamp'][:10],
            'total_records_in_batch': self.stats['total_records_in_batch'],
            'records_reviewed': self.stats['total_records_in_export'],
            'records_edited': self.stats['records_with_edits'],
            'records_reviewed_only': self.stats['records_reviewed_only'],
            'total_field_edits': len(self.edit_history),
        }
        if self.evaluator_name:
            batch_info['evaluator'] = self.evaluator_name
            batch_info['image_folder'] = self.folder_name

        # Group edit history by record_id so each record's edits can be embedded
        edits_by_record = {}
        for edit in self.edit_history:
            rid = edit['record_id']
            edits_by_record.setdefault(rid, []).append(edit)

        # Build per-record summary and embed raw edit history in each record
        records = self._build_per_record_summary()
        for record in records:
            record['edits'] = edits_by_record.get(record['id'], [])

        # Build the combined report
        report = {
            'generated': self.stats['integration_timestamp'],
            'batch_info': batch_info,
            'average_quality_scores': quality_scores,
            'text_fields_summary': {
                'text_fields_totals': text_fields_totals,
                'text_fields': text_fields_summary,
                'list_fields': list_fields_summary
            },
            'subjects': subjects_summary,
            'subject_headings': subject_headings_summary,
            'records': records,
        }

        # Save report (with evaluator name, folder, and date in filename if specified)
        if self.evaluator_name:
            safe_name = re.sub(r'[^\w\-]', '_', self.evaluator_name)
            date_str = self.stats['integration_timestamp'][:10]
            report_filename = f"edit_report_{safe_name}_{self.folder_name}_{date_str}.json"
        else:
            report_filename = "edit_report.json"
        report_path = os.path.join(self.report_folder, report_filename)
        try:
            with open(report_path, 'w', encoding='utf-8') as f:
                json.dump(report, f, indent=2, ensure_ascii=False)
            print(f"   Generated {report_filename}")
            return report_path
        except Exception as e:
            print(f"   Error generating edit report: {e}")
            return None

    def print_summary(self):
        """Print a summary of the integration."""
        print("\n" + "=" * 60)
        print("INTEGRATION SUMMARY")
        print("=" * 60)
        print(f"Archivist: {self.stats['archivist_name']}")
        print(f"Total records processed: {self.stats['total_records_in_export']}")
        print(f"Records with edits: {self.stats['records_with_edits']}")
        print(f"Records reviewed only: {self.stats['records_reviewed_only']}")

        print(f"\n--- Field Edits ---")
        print(f"Total field edits: {self.stats['total_field_edits']}")
        if self.stats['edits_by_field']:
            for field, count in sorted(self.stats['edits_by_field'].items()):
                print(f"   {field}: {count}")

        # Text field character metrics
        if self.stats['text_field_metrics']:
            print(f"\n--- Character-Level Changes ---")
            total_added = 0
            total_deleted = 0
            for field, metrics in sorted(self.stats['text_field_metrics'].items()):
                edits = metrics['edits_count']
                avg_sim = round(metrics['similarity_ratio_sum'] / edits, 1) if edits > 0 else 0
                avg_tsr = round(metrics['token_sort_ratio_sum'] / edits, 1) if edits > 0 else 0
                print(f"   {field}:")
                print(f"      Edits: {edits}")
                print(f"      Characters added: {metrics['chars_added']}")
                print(f"      Characters deleted: {metrics['chars_deleted']}")
                print(f"      Edit distance: {metrics['edit_distance_sum']}")
                print(f"      Avg edited field similarity ratio: {avg_sim}/100")
                print(f"      Avg edited field token sort ratio: {avg_tsr}/100")
                total_added += metrics['chars_added']
                total_deleted += metrics['chars_deleted']
            print(f"   TOTALS:")
            print(f"      All characters added: {total_added}")
            print(f"      All characters deleted: {total_deleted}")
            print(f"      Net change: {total_added - total_deleted:+d}")

        # Subject metrics (broad topics)
        print(f"\n--- Subjects (AI-identified topics) ---")
        print(f"   Total: {self.stats['subjects_total']}")
        print(f"   Accepted: {self.stats['subjects_accepted']}")
        print(f"   Rejected: {self.stats['subjects_rejected']}")
        if self.stats['subjects_custom_added'] > 0:
            print(f"   Custom added: {self.stats['subjects_custom_added']}")

        # Subject heading metrics (controlled vocabulary terms)
        print(f"\n--- Subject Headings (controlled vocabulary) ---")
        print(f"   AI-selected:")
        print(f"      Total: {self.stats['ai_selected_total']}")
        print(f"      Approved: {self.stats['ai_selected_approved']}")
        print(f"      Rejected: {self.stats['ai_selected_rejected']}")
        print(f"      Cascade rejected: {self.stats['ai_selected_cascade_rejected']}")
        print(f"   Archivist added from list: {self.stats['archivist_added_from_list']}")
        print(f"   Archivist added custom: {self.stats['archivist_added_custom']}")

    def _get_approved_other_terms(self, analysis, term_decisions):
        """Get terms from vocabulary_search_results that were approved via 'other-*' decisions.

        The HTML review page presents "other" vocabulary terms organized by source
        (LCSH, FAST, Getty AAT, Getty TGN) with IDs like 'other-aat-5' or 'other-lcsh-3'.
        This method reconstructs that same ordering to look up the approved terms.

        Args:
            analysis: The record's analysis dict containing vocabulary_search_results
            term_decisions: Dict of term decisions including 'other-*' approvals

        Returns:
            List of term dicts with label, uri, source, and derived_from_subject
        """
        vocab_results = analysis.get('vocabulary_search_results', {})
        if not vocab_results:
            return []

        # Build set of already-selected term labels (same logic as html-review.py)
        final_terms = analysis.get('final_selected_terms', [])
        selected_labels = set()
        for term in final_terms:
            selected_labels.add(term.get('label', '').lower())

        # Group terms by source in the same order as html-review.py
        source_map = {
            'lcsh': 'LCSH',
            'fast': 'FAST',
            'aat': 'Getty AAT',
            'tgn': 'Getty TGN'
        }
        sources = {'LCSH': [], 'FAST': [], 'Getty AAT': [], 'Getty TGN': []}

        for orig_term, matches in vocab_results.items():
            for match in matches:
                source = match.get('source', 'Unknown')
                label = match.get('label', '')
                # Skip if this term was already selected
                if label.lower() in selected_labels:
                    continue
                if source in sources:
                    sources[source].append({
                        'orig_term': orig_term,
                        'label': label,
                        'uri': match.get('uri', ''),
                        'source': source
                    })

        # Find all approved 'other-*' terms
        approved_terms = []
        for term_id, decision in term_decisions.items():
            if not term_id.startswith('other-'):
                continue

            # Only process approved decisions
            if decision != 'approved':
                continue

            # Parse the term_id: 'other-{source_class}-{index}'
            parts = term_id.split('-')
            if len(parts) != 3:
                continue

            source_class = parts[1]  # e.g., 'aat', 'lcsh', 'fast', 'tgn'
            try:
                index = int(parts[2])
            except ValueError:
                continue

            # Map source_class back to full source name
            source_name = source_map.get(source_class)
            if not source_name:
                continue

            # Get the term at that index (limited to first 15, same as HTML)
            source_terms = sources.get(source_name, [])[:15]
            if index < len(source_terms):
                term = source_terms[index]
                approved_terms.append({
                    'label': term['label'],
                    'uri': term['uri'],
                    'source': term['source'],
                    'derived_from_subject': term.get('orig_term', '')
                })

        return approved_terms

    def _clean_workflow_lists_after_decisions(self):
        """Remove rejected terms from the raw lists in workflow data.

        Called after generate_final_metadata() (which reads decisions correctly) and
        before save_workflow_json(), so drawings_workflow.json reflects only approved terms.

        Also prunes archivist_term_decisions to remove applied rejection entries, keeping
        only opt-in approvals (other-*, medium-other-*) so generate_final_metadata() can
        still correctly include those terms on any future run.
        """
        for record in self.workflow_data:
            analysis = record.get('analysis', {})
            term_decisions = analysis.get('archivist_term_decisions', {})
            if not term_decisions:
                continue

            # Filter subjects (remove rejected topics)
            original_subjects = analysis.get('subjects', [])
            if not isinstance(original_subjects, list):
                original_subjects = [original_subjects] if original_subjects else []
            filtered_subjects = []
            for i, subject in enumerate(original_subjects):
                decision = term_decisions.get(f"subject-{i}")
                if decision is None or decision == 'approved':
                    filtered_subjects.append(subject)
            analysis['subjects'] = filtered_subjects

            # Filter final_selected_terms (remove rejected/cascade_rejected headings)
            final_terms = analysis.get('final_selected_terms', [])
            filtered_terms = []
            for i, term in enumerate(final_terms):
                decision = term_decisions.get(f"selected-{i}")
                if isinstance(decision, dict) and decision.get('status') == 'cascade_rejected':
                    continue
                if decision == 'rejected':
                    continue
                filtered_terms.append(term)
            analysis['final_selected_terms'] = filtered_terms

            # Filter final_selected_medium_terms (remove rejected medium/support headings)
            medium_terms = analysis.get('final_selected_medium_terms', [])
            filtered_medium = []
            for i, term in enumerate(medium_terms):
                decision = term_decisions.get(f"medium-selected-{i}")
                if decision == 'rejected':
                    continue
                filtered_medium.append(term)
            analysis['final_selected_medium_terms'] = filtered_medium

            # Prune archivist_term_decisions: keep only opt-in approvals (other-*, medium-other-*)
            # so generate_final_metadata() can still add those terms on future runs.
            # Remove rejection decisions (subject-*, selected-*, medium-selected-*) since
            # those are now encoded in the filtered lists above.
            pruned_decisions = {
                tid: dec for tid, dec in term_decisions.items()
                if tid.startswith('other-') or tid.startswith('medium-other-')
            }
            analysis['archivist_term_decisions'] = pruned_decisions

            record['analysis'] = analysis

    def generate_final_metadata(self):
        """Generate final_metadata.json with only approved/clean data.

        This creates a clean JSON file containing only:
        - Edited/final field values
        - Only approved subjects (not rejected)
        - Only approved subject headings (not rejected or cascade-rejected)
        - Custom terms/subjects added by archivist
        - No intermediate data (raw_response, vocabulary_search_results, etc.)
        """
        if not self.workflow_data:
            print("   Warning: No workflow data loaded")
            return False

        final_records = []

        for record in self.workflow_data:
            analysis = record.get('analysis', {})
            term_decisions = analysis.get('archivist_term_decisions', {})

            # Get the list of original subjects
            original_subjects = analysis.get('subjects', [])
            if not isinstance(original_subjects, list):
                original_subjects = [original_subjects] if original_subjects else []

            # Filter subjects: keep only approved ones
            approved_subjects = []
            for i, subject in enumerate(original_subjects):
                term_id = f"subject-{i}"
                decision = term_decisions.get(term_id)
                # Keep if no decision (default approved) or explicitly approved
                if decision is None or decision == 'approved':
                    approved_subjects.append(subject)

            # Filter subject headings: keep only approved ones
            final_selected_terms = analysis.get('final_selected_terms', [])
            approved_headings = []
            for i, term in enumerate(final_selected_terms):
                term_id = f"selected-{i}"
                decision = term_decisions.get(term_id)

                # Check if decision is cascade_rejected (object format)
                if isinstance(decision, dict) and decision.get('status') == 'cascade_rejected':
                    continue  # Skip cascade-rejected
                elif decision == 'rejected':
                    continue  # Skip explicitly rejected

                # Keep approved or no decision (default approved)
                approved_headings.append({
                    'label': term.get('label', ''),
                    'uri': term.get('uri', ''),
                    'source': term.get('source', ''),
                    'derived_from_subject': term.get('derived_from_subject', '')
                })

            # Add any custom terms added by archivist
            custom_terms = analysis.get('archivist_custom_terms', [])
            for term in custom_terms:
                approved_headings.append({
                    'label': term.get('label', ''),
                    'uri': term.get('uri', ''),
                    'source': term.get('source', 'Manual'),
                    'derived_from_subject': ''
                })

            # Add "other-*" terms that were approved from the vocabulary list
            # These are terms from vocabulary_search_results that weren't AI-selected
            # but were manually approved by the archivist
            other_terms = self._get_approved_other_terms(analysis, term_decisions)
            for term in other_terms:
                approved_headings.append({
                    'label': term.get('label', ''),
                    'uri': term.get('uri', ''),
                    'source': term.get('source', ''),
                    'derived_from_subject': term.get('derived_from_subject', '')
                })

            # Filter medium/support vocabulary terms: keep only approved ones
            final_selected_medium_terms = analysis.get('final_selected_medium_terms', [])
            approved_medium_terms = []
            for i, term in enumerate(final_selected_medium_terms):
                term_id = f"medium-selected-{i}"
                decision = term_decisions.get(term_id)
                if decision == 'rejected':
                    continue  # Skip explicitly rejected
                approved_medium_terms.append({
                    'label': term.get('label', ''),
                    'uri': term.get('uri', ''),
                    'source': term.get('source', ''),
                    'derived_from_subject': term.get('derived_from_subject', '')
                })

            # Add "medium-other-*" terms that were approved from the medium vocabulary list
            medium_vocab_results = analysis.get('medium_vocabulary_search_results', {})
            if medium_vocab_results:
                selected_medium_labels = set(t.get('label', '').lower() for t in final_selected_medium_terms)
                other_medium_list = []
                for topic, terms in medium_vocab_results.items():
                    for term in terms:
                        label = term.get('label', '')
                        if label.lower() not in selected_medium_labels:
                            other_medium_list.append({'topic': topic, 'label': label,
                                                      'uri': term.get('uri', ''), 'source': term.get('source', 'Getty AAT')})
                for i, term in enumerate(other_medium_list):
                    term_id = f"medium-other-{i}"
                    decision = term_decisions.get(term_id)
                    if decision == 'approved':
                        approved_medium_terms.append({
                            'label': term['label'],
                            'uri': term['uri'],
                            'source': term['source'],
                            'derived_from_subject': term.get('topic', '')
                        })

            # Build clean record
            clean_record = {
                'folder': record.get('folder', ''),
                'page_number': record.get('page_number', 0),
                'image_path': record.get('image_path', ''),
                'metadata': {
                    'title': analysis.get('title', ''),
                    'contributors': analysis.get('contributors', []),
                    'genre': analysis.get('genre', ''),
                    'description': analysis.get('description', ''),
                    'format_media': analysis.get('format_media', ''),
                    'medium': analysis.get('medium', ''),
                    'support': analysis.get('support', ''),
                    'format_media_terms': approved_medium_terms,
                    'date_on_drawing': analysis.get('date_on_drawing', ''),
                    'sheet_info': analysis.get('sheet_info', ''),
                    'named_entities': analysis.get('named_entities', []),
                    'geographic_entities': analysis.get('geographic_entities', []),
                    'content_warning': analysis.get('content_warning', ''),
                    'subjects': approved_subjects,
                    'subject_headings': approved_headings
                },
                'review_info': {
                    'reviewed': analysis.get('archivist_reviewed', False),
                    'archivist_name': analysis.get('archivist_name', ''),
                    'review_date': analysis.get('archivist_review_date', ''),
                    'archivist_notes': analysis.get('archivist_notes', '')
                }
            }

            final_records.append(clean_record)

        # Write final metadata
        final_metadata_path = os.path.join(self.json_folder, "final_metadata.json")
        try:
            with open(final_metadata_path, 'w', encoding='utf-8') as f:
                json.dump({
                    'generated_timestamp': datetime.now().isoformat(),
                    'archivist_name': self.stats['archivist_name'],
                    'total_records': len(final_records),
                    'records': final_records
                }, f, indent=2, ensure_ascii=False)

            print(f"   Generated final_metadata.json with {len(final_records)} records")
            return True
        except Exception as e:
            print(f"   Error generating final_metadata.json: {e}")
            return False

    def run(self, decisions_path=None):
        """Main execution method.

        Args:
            decisions_path: Optional path to a specific decisions JSON file.
                          If not provided, uses the latest export.
        """
        print("\n" + "=" * 60)
        if self.analysis_only:
            print(f"Analyze Archivist Edits (Evaluator: {self.evaluator_name})")
        else:
            print("Integrate Archivist Edits")
        print("=" * 60)

        print(f"\nUsing folder: {self.folder_name}")
        if self.analysis_only:
            print(f"Mode: Analysis only (no files will be modified)")

        # Find or use specified export
        print("\n1. Finding archivist decisions export...")
        if decisions_path:
            if not os.path.exists(decisions_path):
                print(f"   Error: Specified decisions file not found: {decisions_path}")
                return False
            export_path = decisions_path
            print(f"   Using specified file: {os.path.basename(export_path)}")
        else:
            export_path = self.find_latest_export()
            if not export_path:
                return False
            print(f"   Found latest: {os.path.basename(export_path)}")

        # Load decisions
        print("\n2. Loading archivist decisions...")
        if not self.load_decisions(export_path):
            return False

        # Load workflow JSON
        print("\n3. Loading workflow data...")
        if not self.load_workflow_json():
            return False

        # Backup original files (skip in analysis-only mode)
        if self.analysis_only:
            print("\n4. Skipping backup (analysis-only mode)...")
        else:
            print("\n4. Backing up original files...")
            self.backup_original_files()

        # Apply edits
        print("\n5. Applying archivist edits...")
        if not self.apply_all_edits():
            return False
        print(f"   Applied edits to {self.stats['records_with_edits']} records")

        # Generate final metadata (skip in analysis-only mode)
        # Must run before save_workflow_json so it can read the full raw lists + decisions.
        if self.analysis_only:
            print("\n6. Skipping final_metadata.json (analysis-only mode)...")
        else:
            print("\n6. Generating final_metadata.json...")
            self.generate_final_metadata()

        # Remove rejected terms from workflow lists now that final_metadata is written.
        # Also prune applied rejection decisions from archivist_term_decisions.
        if not self.analysis_only:
            self._clean_workflow_lists_after_decisions()

        # Save updated workflow JSON (skip in analysis-only mode)
        if self.analysis_only:
            print("\n7. Skipping workflow JSON save (analysis-only mode)...")
        else:
            print("\n7. Saving updated workflow JSON...")
            if not self.save_workflow_json():
                return False

        # Generate combined edit report
        print("\n8. Generating report...")
        report_path = self.generate_edit_statistics_report()

        # Create final Excel deliverable (skip in analysis-only mode)
        if self.analysis_only:
            print("\n9. Skipping Excel deliverable (analysis-only mode)...")
            excel_path = None
        else:
            print("\n9. Creating final Excel deliverable...")
            excel_path = self.create_final_excel_deliverable()

        # Print summary
        self.print_summary()

        # Build output path (with evaluator name, folder, and date in filename if specified)
        if self.evaluator_name:
            safe_name = re.sub(r'[^\w\-]', '_', self.evaluator_name)
            date_str = self.stats['integration_timestamp'][:10]
            edit_report_path = os.path.join(self.report_folder, f"edit_report_{safe_name}_{self.folder_name}_{date_str}.json")
        else:
            edit_report_path = os.path.join(self.report_folder, "edit_report.json")

        print("\n" + "=" * 60)
        if self.analysis_only:
            print(f"ANALYSIS COMPLETE (Evaluator: {self.evaluator_name})")
            print("=" * 60)
            print(f"\nReport generated:")
            print(f"   {edit_report_path}")
            print(f"\nNo files were modified (analysis-only mode).")
        else:
            final_metadata_path = os.path.join(self.json_folder, "final_metadata.json")
            final_excel_path = os.path.join(self.metadata_folder, "final_deliverable.xlsx")
            print("INTEGRATION COMPLETE")
            print("=" * 60)
            print(f"\nUpdated files:")
            print(f"   {self.workflow_json_path}")
            print(f"   {final_metadata_path}")
            print(f"\nReport generated:")
            print(f"   {edit_report_path}")
            print(f"\nFinal deliverable:")
            print(f"   {final_excel_path}")
            print(f"\nOriginal files backed up to:")
            print(f"   {self.original_outputs_folder}")
        print("=" * 60)

        return True


def list_available_folders(base_dir):
    """List all available output folders."""
    if not os.path.exists(base_dir):
        return []

    folders = []
    for item in os.listdir(base_dir):
        item_path = os.path.join(base_dir, item)
        if os.path.isdir(item_path) and item.startswith("ArchImagesAI_"):
            folders.append((item, item_path))

    # Sort by modification time, newest first
    folders.sort(key=lambda x: os.path.getmtime(x[1]), reverse=True)
    return folders


def list_available_exports(exports_folder):
    """List all available JSON exports in the exports folder."""
    if not os.path.exists(exports_folder):
        return []

    exports = []
    for item in os.listdir(exports_folder):
        if item.endswith('.json'):
            item_path = os.path.join(exports_folder, item)
            exports.append((item, item_path))

    # Sort by modification time, newest first
    exports.sort(key=lambda x: os.path.getmtime(x[1]), reverse=True)
    return exports


def prompt_for_folder(base_dir):
    """Prompt user to select an output folder."""
    folders = list_available_folders(base_dir)

    if not folders:
        print(f"No output folders found in: {base_dir}")
        print("Please run Steps 1-4 first, then html-review.py.")
        return None

    print("\nAvailable output folders:")
    print("-" * 60)
    for i, (name, path) in enumerate(folders, 1):
        mtime = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M")
        marker = " (newest)" if i == 1 else ""
        print(f"  {i}. {name}{marker}")
        print(f"      Modified: {mtime}")

    print("-" * 60)
    print(f"  Enter 1-{len(folders)} to select a folder")
    print(f"  Or press Enter to use the newest folder (1)")
    print(f"  Or type a full path to a folder")

    while True:
        choice = input("\nSelect output folder: ").strip()

        # Default to newest
        if choice == "":
            return folders[0][1]

        # Check if it's a number
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(folders):
                return folders[idx][1]
            else:
                print(f"Invalid choice. Please enter 1-{len(folders)}.")
                continue

        # Check if it's a path
        if os.path.isdir(choice):
            return choice

        print(f"Invalid input. Please enter a number or valid path.")


def prompt_for_decisions(folder_path):
    """Prompt user to select an archivist decisions JSON file."""
    exports_folder = os.path.join(folder_path, "review", "exports")
    exports = list_available_exports(exports_folder)

    if not exports:
        print(f"\nNo JSON export files found in: {exports_folder}")
        print("Please export your review decisions from the HTML interface first.")
        print("\nYou can also enter a full path to a decisions JSON file.")

        while True:
            path = input("\nPath to decisions JSON (or 'q' to quit): ").strip()
            if path.lower() == 'q':
                return None
            if os.path.isfile(path) and path.endswith('.json'):
                return path
            print("Invalid path. Please enter a valid JSON file path.")

    print("\nAvailable archivist decisions exports:")
    print("-" * 60)
    for i, (name, path) in enumerate(exports, 1):
        mtime = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M")
        marker = " (newest)" if i == 1 else ""
        print(f"  {i}. {name}{marker}")
        print(f"      Modified: {mtime}")

    print("-" * 60)
    print(f"  Enter 1-{len(exports)} to select an export")
    print(f"  Or press Enter to use the newest export (1)")
    print(f"  Or type a full path to a JSON file")

    while True:
        choice = input("\nSelect decisions export: ").strip()

        # Default to newest
        if choice == "":
            return exports[0][1]

        # Check if it's a number
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(exports):
                return exports[idx][1]
            else:
                print(f"Invalid choice. Please enter 1-{len(exports)}.")
                continue

        # Check if it's a path
        if os.path.isfile(choice) and choice.endswith('.json'):
            return choice

        print(f"Invalid input. Please enter a number or valid JSON path.")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Integrate archivist edits into workflow outputs and generate final metadata.',
        epilog="""
Examples:
  python integrate-archivist-edits.py                           # Interactive prompts
  python integrate-archivist-edits.py --decisions path/to.json  # Use specific export
  python integrate-archivist-edits.py --folder /path/to/output  # Specify output folder

Analysis-only mode (for comparing multiple evaluators):
  python integrate-archivist-edits.py -e "Alice" -d alice_edits.json -f /path/to/pipeline/output -o /path/to/evaluator/reports
  python integrate-archivist-edits.py -e "Bob" -d bob_edits.json -f /path/to/pipeline/output -o /path/to/evaluator/reports

  --folder (-f): the original pipeline output folder (where drawings_workflow.json lives)
  --output (-o): where evaluator reports are saved (created if needed; defaults to pipeline folder's metadata/json/)
  This generates edit_report_Alice_<folder>_<date>.json and edit_report_Bob_<folder>_<date>.json
  without modifying any workflow files, allowing side-by-side comparison.
        """
    )
    parser.add_argument('--decisions', '-d',
                        help='Path to a specific archivist decisions JSON file')
    parser.add_argument('--folder', '-f',
                        help='Path to the output folder (defaults to interactive prompt)')
    parser.add_argument('--evaluator', '-e',
                        help='Evaluator name for analysis-only mode. When set, generates '
                             'named statistics report without modifying workflow files. '
                             'Useful for comparing multiple evaluators on the same output.')
    parser.add_argument('--output', '-o',
                        help='Output folder for evaluator reports (evaluator mode only). '
                             'Defaults to the pipeline folder\'s metadata/json/ directory.')
    parser.add_argument('--yes', '-y', action='store_true',
                        help='Skip confirmation prompt')

    args = parser.parse_args()

    print("Integrate Archivist Edits")
    print("=" * 60)

    base_output_dir = os.path.join(script_dir, "output_folders")

    # Get output folder - either from args or interactive prompt
    if args.folder:
        folder_path = args.folder
        if not os.path.exists(folder_path):
            print(f"Error: Specified folder not found: {folder_path}")
            return 1
        print(f"Using folder: {os.path.basename(folder_path)}")
    else:
        folder_path = prompt_for_folder(base_output_dir)
        if not folder_path:
            return 1
        print(f"\nSelected folder: {os.path.basename(folder_path)}")

    # Get decisions file - either from args or interactive prompt
    if args.decisions:
        decisions_path = args.decisions
        if not os.path.exists(decisions_path):
            print(f"Error: Specified decisions file not found: {decisions_path}")
            return 1
        print(f"Using decisions: {os.path.basename(decisions_path)}")
    else:
        decisions_path = prompt_for_decisions(folder_path)
        if not decisions_path:
            print("Operation cancelled.")
            return 0
        print(f"\nSelected decisions: {os.path.basename(decisions_path)}")

    # Confirm with user unless --yes flag or analysis-only mode
    if not args.yes and not args.evaluator:
        print("\n" + "-" * 60)
        print("Summary:")
        print(f"  Output folder: {os.path.basename(folder_path)}")
        print(f"  Decisions file: {os.path.basename(decisions_path)}")
        print("-" * 60)
        response = input("\nThis will modify deliverable files. Original files will be backed up. Continue? (yes/no): ").strip().lower()
        if response not in ['yes', 'y']:
            print("Operation cancelled.")
            return 0
    elif args.evaluator:
        print("\n" + "-" * 60)
        print("Summary (Analysis-Only Mode):")
        print(f"  Output folder: {os.path.basename(folder_path)}")
        print(f"  Decisions file: {os.path.basename(decisions_path)}")
        print(f"  Evaluator: {args.evaluator}")
        print("  (No files will be modified)")
        print("-" * 60)

    # Create integrator and run
    integrator = ArchivistEditsIntegrator(folder_path, evaluator_name=args.evaluator,
                                          output_folder=args.output)
    success = integrator.run(decisions_path=decisions_path)

    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())
