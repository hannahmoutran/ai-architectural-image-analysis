# Extracts metadata from architectural drawings using Anthropic's Claude model
import os
import json
import base64
import logging
from datetime import datetime
import anthropic
import tenacity
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
import time
from prompts import ArchitecturalDrawingPrompts
from shared_utilities import APIStats, postprocess_api_response, parse_json_response_enhanced

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Import custom modules
from model_pricing import calculate_cost
from token_logging import create_token_usage_log, log_individual_response
from batch_processor import BatchProcessor

# Suppress verbose HTTP logging
logging.getLogger("anthropic").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)

client = anthropic.Anthropic(api_key=os.getenv('CLAUDE_API_KEY'))
DEFAULT_MODEL = "claude-sonnet-4-5-20250929"

api_stats = APIStats()

def parse_json_response(raw_response):
    """JSON parsing with trailing comma handling."""
    return parse_json_response_enhanced(raw_response)


def parse_key_value_response(raw_response: str) -> tuple[dict, str]:
    """
    Parse key-value format response from LLM into a dictionary.

    Expected format (supports both plain and markdown-formatted):
    Title: First Floor Plan - Smith Residence
    **Title:** First Floor Plan - Smith Residence
    Contributors: John Smith (Architect); ABC Engineering (Structural Engineer)
    Genre: floor plan
    ...

    Returns:
        Tuple of (parsed_dict, error_message). Error is None on success.
    """
    if not raw_response or not raw_response.strip():
        return None, "Empty response"

    result = {}

    # Known field names and their dictionary keys (display name -> key)
    field_mapping = {
        'title': 'title',
        'contributors': 'contributors',
        'genre': 'genre',
        'ocr text': 'ocrText',
        'ocrtext': 'ocrText',
        'description': 'description',
        'format media': 'formatMedia',
        'formatmedia': 'formatMedia',
        'subjects': 'subjects',
        'date on drawing': 'dateOnDrawing',
        'dateondrawing': 'dateOnDrawing',
        'sheet info': 'sheetInfo',
        'sheetinfo': 'sheetInfo',
        'named entities': 'namedEntities',
        'namedentities': 'namedEntities',
        'geographic entities': 'geographicEntities',
        'geographicentities': 'geographicEntities',
        'content warning': 'contentWarning',
        'contentwarning': 'contentWarning'
    }

    current_key = None
    current_value_lines = []

    for line in raw_response.split('\n'):
        # Check if this line starts a new field
        found_field = False

        # Strip markdown formatting (**, *, #) from the line for field detection
        cleaned_line = line.strip()
        # Remove leading markdown headers (# ## ### etc.)
        cleaned_line = re.sub(r'^#+\s*', '', cleaned_line)
        # Remove bold/italic markers around the field name
        cleaned_line = re.sub(r'^\*+\s*', '', cleaned_line)
        cleaned_line = re.sub(r'\*+', '', cleaned_line)
        cleaned_line_lower = cleaned_line.lower()

        for display_name, dict_key in field_mapping.items():
            if cleaned_line_lower.startswith(display_name + ':'):
                # Save previous field if exists
                if current_key:
                    value = ' '.join(current_value_lines).strip()
                    result[current_key] = _parse_field_value(current_key, value)

                # Start new field
                current_key = dict_key
                # Get value after the colon (use cleaned line to preserve case of value)
                colon_pos = cleaned_line.lower().find(display_name + ':')
                if colon_pos != -1:
                    value_start = colon_pos + len(display_name) + 1
                    current_value_lines = [cleaned_line[value_start:].strip()]
                found_field = True
                break

        # If no new field found and we have a current key, this is a continuation
        # Skip lines that are just markdown headers or empty
        if not found_field and current_key and line.strip():
            stripped = line.strip()
            # Skip markdown header lines (e.g., "# Metadata Extraction")
            if not stripped.startswith('#'):
                current_value_lines.append(stripped)

    # Save the last field
    if current_key:
        value = ' '.join(current_value_lines).strip()
        result[current_key] = _parse_field_value(current_key, value)

    if not result:
        return None, "No valid fields found in response"

    return result, None


def _parse_field_value(field_key: str, value: str):
    """
    Parse a field value, converting list fields appropriately.

    Args:
        field_key: The dictionary key for the field
        value: The raw string value

    Returns:
        Parsed value (string or list depending on field type)
    """
    # Fields that should be parsed as lists (semicolon-separated)
    list_fields = ['contributors', 'subjects', 'namedEntities', 'geographicEntities']

    if field_key in list_fields:
        # Split by semicolon and clean up
        items = [item.strip() for item in value.split(';') if item.strip()]
        # For contributors, convert to list of dicts if possible
        if field_key == 'contributors':
            parsed_contributors = []
            for item in items:
                # Try to parse "Name (Role)" format
                match = re.match(r'^(.+?)\s*\(([^)]+)\)\s*$', item)
                if match:
                    parsed_contributors.append({
                        'name': match.group(1).strip(),
                        'role': match.group(2).strip()
                    })
                else:
                    parsed_contributors.append({'name': item, 'role': ''})
            return parsed_contributors
        return items

    return value


def parse_key_value_text(text: str) -> dict:
    """
    Parse key-value text format into a dictionary.

    Expected format:
    Creator: Gordon, James Riely, 1863-1937
    Title: James Riely Gordon collection
    Dates: 1890-1937
    Abstract: James Riely Gordon (1863-1937), an architect who practiced...

    Keys are case-insensitive and mapped to lowercase.
    """
    result = {}
    current_key = None
    current_value_lines = []

    # Known field names (case-insensitive)
    known_fields = ['creator', 'title', 'dates', 'abstract', 'extent', 'repository',
                    'identification', 'language']

    for line in text.split('\n'):
        # Check if this line starts a new field
        found_field = False
        for field in known_fields:
            if line.lower().startswith(field + ':'):
                # Save previous field if exists
                if current_key:
                    result[current_key] = ' '.join(current_value_lines).strip()

                # Start new field
                current_key = field.lower()
                # Get value after the colon
                value_start = len(field) + 1  # +1 for the colon
                current_value_lines = [line[value_start:].strip()]
                found_field = True
                break

        # If no new field found and we have a current key, this is a continuation
        if not found_field and current_key and line.strip():
            current_value_lines.append(line.strip())

    # Save the last field
    if current_key:
        result[current_key] = ' '.join(current_value_lines).strip()

    return result


def load_collection_context(input_folder: str = None) -> str:
    """
    Load collection context from text files in the input folder.

    Expected text format (key-value pairs):
    Creator: Gordon, James Riely, 1863-1937
    Title: James Riely Gordon collection
    Dates: 1890-1937
    Abstract: James Riely Gordon (1863-1937), an architect who practiced...
    Extent: 13 linear feet, 6,500 drawings (approximately)
    Repository: Alexander Architectural Archives, University of Texas Libraries
    """
    context_data = {}

    if not input_folder or not os.path.isdir(input_folder):
        return ""

    # Find and load any .txt files in the folder
    txt_files = [f for f in os.listdir(input_folder)
                  if f.lower().endswith('.txt') and os.path.isfile(os.path.join(input_folder, f))]

    for txt_file in sorted(txt_files):
        txt_path = os.path.join(input_folder, txt_file)
        try:
            with open(txt_path, 'r', encoding='utf-8') as f:
                file_content = f.read()
            file_data = parse_key_value_text(file_content)
            # Merge data (later files override earlier ones for duplicate keys)
            context_data.update(file_data)
            logging.info(f"Loaded collection context from: {txt_path}")
        except Exception as e:
            logging.warning(f"Could not load text from {txt_path}: {e}")

    if not context_data:
        return ""

    return ArchitecturalDrawingPrompts.create_collection_context(
        creator=context_data.get('creator', ''),
        title=context_data.get('title', ''),
        dates=context_data.get('dates', ''),
        abstract=context_data.get('abstract', ''),
        extent=context_data.get('extent', ''),
        repository=context_data.get('repository', '')
    )


def get_image_media_type(img_path: str) -> str:
    """Get the media type for an image based on its extension."""
    ext = os.path.splitext(img_path)[1].lower()
    media_types = {
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
        '.webp': 'image/webp'
    }
    # TIFF files will be converted to JPEG, so return jpeg type
    if ext in ['.tif', '.tiff']:
        return 'image/jpeg'
    return media_types.get(ext, 'image/jpeg')


def prepare_image_for_api(image_path: str) -> tuple[str, str]:
    """
    Prepare an image for the Claude API.
    Converts unsupported formats (like TIFF) to JPEG.

    Returns:
        Tuple of (base64_encoded_data, media_type)
    """
    ext = os.path.splitext(image_path)[1].lower()

    # Claude API supports: jpeg, png, gif, webp
    # TIFF needs to be converted
    if ext in ['.tif', '.tiff']:
        img = PILImage.open(image_path)
        # Convert to RGB if necessary (TIFF can be RGBA or other modes)
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGB')
        elif img.mode != 'RGB':
            img = img.convert('RGB')

        # Save to bytes buffer as JPEG
        buffer = BytesIO()
        img.save(buffer, format='JPEG', quality=95)
        buffer.seek(0)
        base64_image = base64.b64encode(buffer.read()).decode('utf-8')
        return base64_image, 'image/jpeg'
    else:
        # For supported formats, read directly
        with open(image_path, "rb") as image_file:
            base64_image = base64.b64encode(image_file.read()).decode('utf-8')
        return base64_image, get_image_media_type(image_path)


def collect_all_images(input_folder):
    """Collect all images to process.
    Accepts either:
      1) input_folder/Issue*/page*.jpg   (subfolders with page-numbered files)
      2) input_folder/page*.jpg          (images directly in the folder)
    Files without a 'pageN' pattern are allowed and will be ordered by filename.
    """
    all_images = []

    def safe_page_num(fname: str) -> int:
        m = re.search(r'page(\d+)', fname, re.IGNORECASE)
        return int(m.group(1)) if m else 0

    # 1) Images directly in input_folder
    direct_images = [f for f in os.listdir(input_folder)
                     if os.path.isfile(os.path.join(input_folder, f))
                     and f.lower().endswith(('.jpg', '.jpeg', '.png', '.tif', '.tiff'))]
    if direct_images:
        direct_images_sorted = sorted(direct_images, key=lambda x: (safe_page_num(x), x.lower()))
        folder_label = os.path.basename(os.path.normpath(input_folder)) or "root"
        for idx, img_file in enumerate(direct_images_sorted, start=1):
            page_number = safe_page_num(img_file) or idx
            img_path = os.path.join(input_folder, img_file)
            all_images.append((folder_label, page_number, img_path))

    # 2) Images inside subfolders of input_folder
    for folder_name in sorted(os.listdir(input_folder), key=lambda x: x.lower()):
        folder_path = os.path.join(input_folder, folder_name)
        if os.path.isdir(folder_path):
            image_files = [f for f in os.listdir(folder_path)
                           if f.lower().endswith(('.jpg', '.jpeg', '.png', '.tif', '.tiff'))]
            image_files_sorted = sorted(image_files, key=lambda x: (safe_page_num(x), x.lower()))
            for idx, img_file in enumerate(image_files_sorted, start=1):
                page_number = safe_page_num(img_file) or idx
                img_path = os.path.join(folder_path, img_file)
                all_images.append((folder_name, page_number, img_path))

    return all_images


@tenacity.retry(
    wait=tenacity.wait_exponential(multiplier=1, min=4, max=10),
    stop=tenacity.stop_after_attempt(3),
    retry=tenacity.retry_if_exception_type(Exception)
)
def process_image(image_path, model_name=DEFAULT_MODEL, collection_context=""):
    """Process a single architectural drawing image and return the parsed response."""
    # Prepare image (converts TIFF to JPEG if needed)
    base64_image, media_type = prepare_image_for_api(image_path)

    # Get the architectural drawing prompt with collection context
    prompt = ArchitecturalDrawingPrompts.get_architectural_drawing_prompt(collection_context)

    api_stats.total_requests += 1
    start_time = time.time()

    response = client.messages.create(
        model=model_name,
        max_tokens=3000,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": base64_image
                    }
                },
                {"type": "text", "text": prompt}
            ]
        }]
    )

    processing_time = time.time() - start_time
    api_stats.processing_times.append(processing_time)

    api_stats.total_input_tokens += response.usage.input_tokens
    api_stats.total_output_tokens += response.usage.output_tokens

    raw_response = response.content[0].text.strip()

    # Use key-value parsing function for the new format
    parsed_data, error = parse_key_value_response(raw_response)

    if parsed_data:
        # Ensure required fields exist with defaults
        required_fields = ['title', 'contributors', 'genre', 'ocrText',
                          'description','formatMedia', 'subjects', 'dateOnDrawing', 'sheetInfo',
                          'namedEntities', 'geographicEntities', 'contentWarning']

        for field in required_fields:
            if field not in parsed_data:
                if field in ['contributors', 'subjects', 'namedEntities', 'geographicEntities']:
                    parsed_data[field] = []
                else:
                    parsed_data[field] = ""

        # Post-process the response
        parsed_data = postprocess_api_response(parsed_data)

        return parsed_data, raw_response, response.usage, processing_time
    else:
        logging.error(f"Key-value parsing failed for {image_path}: {error}\nRaw response: {raw_response}")
        raise Exception(f"Key-value parsing failed: {error}")


def format_contributors(contributors):
    """Format contributors list for Excel display."""
    if not contributors:
        return ""

    formatted = []
    for contrib in contributors:
        if isinstance(contrib, dict):
            name = contrib.get('name', '')
            role = contrib.get('role', '')
            if name and role:
                formatted.append(f"{name} ({role})")
            elif name:
                formatted.append(name)
        elif isinstance(contrib, str):
            formatted.append(contrib)

    return "; ".join(formatted)


def process_folder_with_batch(input_folder, output_dir, model_name=DEFAULT_MODEL, collection_context=""):
    """Process folder - Claude doesn't support batch API in the same way, so we use individual processing."""

    # Create logs folder
    logs_folder_path = os.path.join(output_dir, "logs")
    if not os.path.exists(logs_folder_path):
        os.makedirs(logs_folder_path)

    # Collect all images
    all_images = collect_all_images(input_folder)
    total_items = len(all_images)

    # Claude API doesn't have the same batch processing as OpenAI, so we always use individual processing
    print(f"\nFound {total_items} images to process")
    print(f"Processing mode: INDIVIDUAL")
    print(f"Model: {model_name}")

    wb = Workbook()
    analysis_sheet = wb.active
    analysis_sheet.title = "Analysis"

    # Set up headers for architectural drawings
    analysis_headers = [
        'Folder', 'Drawing Number', 'Image Path',
        'Title', 'Contributors', 'Genre', 'OCR Text',
        'Description', 'Format/Media', 'Subjects',
        'Date on Drawing', 'Sheet Info',
        'Named Entities', 'Geographic Entities', 'Content Warning'
    ]
    analysis_sheet.append(analysis_headers)
    analysis_sheet.freeze_panes = 'A2'

    # Set column widths
    column_widths = [15, 12, 30, 40, 50, 25, 50, 50, 30, 40, 15, 20, 40, 30, 30]
    for i, width in enumerate(column_widths):
        analysis_sheet.column_dimensions[analysis_sheet.cell(row=1, column=i+1).column_letter].width = width

    # Create raw responses sheet
    raw_sheet = wb.create_sheet("Raw Responses")
    raw_headers = ['Folder', 'Drawing Number', 'API Response']
    raw_sheet.append(raw_headers)
    raw_sheet.freeze_panes = 'A2'

    for i, width in enumerate([15, 12, 120]):
        raw_sheet.column_dimensions[raw_sheet.cell(row=1, column=i+1).column_letter].width = width

    # Create issues sheet
    issues_sheet = wb.create_sheet("Issues")
    issues_sheet.append(["Image Path", "Error"])

    max_row_height = 409
    analysis_sheet.row_dimensions[1].height = 15
    raw_sheet.row_dimensions[1].height = 15

    all_results = []

    # Use individual processing
    return process_folder_individual(all_images, wb, analysis_sheet, raw_sheet, issues_sheet,
                                     logs_folder_path, model_name, max_row_height, all_results,
                                     collection_context)


def create_error_response(raw_response, error):
    """Create a standard error response dictionary."""
    return {
        "title": f"Error: {error}",
        "contributors": [],
        "genre": "",
        "description": raw_response,
        "formatMedia": "",
        "subjects": [],
        "dateOnDrawing": "",
        "sheetInfo": "",
        "namedEntities": [],
        "geographicEntities": [],
        "contentWarning": "None"
    }


def add_to_sheets(analysis_sheet, raw_sheet, folder_name, page_number, img_path,
                  response_data, raw_response, max_row_height):
    """Add data to analysis and raw sheets."""
    # Format contributors for display
    contributors_str = format_contributors(response_data.get('contributors', []))

    # Format subjects for display
    subjects = response_data.get('subjects', [])
    if isinstance(subjects, list):
        subjects_str = ', '.join(subjects)
    else:
        subjects_str = str(subjects)

    analysis_row = [
        folder_name,
        page_number,
        img_path,
        response_data.get('title', ''),
        contributors_str,
        response_data.get('genre', ''),
        response_data.get('ocrText', ''),
        response_data.get('description', ''),
        response_data.get('formatMedia', ''),
        subjects_str,
        response_data.get('dateOnDrawing', ''),
        response_data.get('sheetInfo', ''),
        ', '.join(response_data.get('namedEntities', [])) if isinstance(response_data.get('namedEntities'), list) else response_data.get('namedEntities', ''),
        ', '.join(response_data.get('geographicEntities', [])) if isinstance(response_data.get('geographicEntities'), list) else response_data.get('geographicEntities', ''),
        response_data.get('contentWarning', 'None')
    ]
    analysis_sheet.append(analysis_row)

    # Add thumbnail image
    try:
        img = PILImage.open(img_path)
        img.thumbnail((200, 200))
        output = BytesIO()
        img.save(output, format='JPEG')
        output.seek(0)
        img_excel = XLImage(output)
        img_excel.anchor = analysis_sheet.cell(row=analysis_sheet.max_row, column=3).coordinate
        analysis_sheet.add_image(img_excel)
    except Exception as e:
        logging.warning(f"Could not add thumbnail for {img_path}: {e}")

    analysis_sheet.row_dimensions[analysis_sheet.max_row].height = max_row_height

    # Set alignment
    current_row = analysis_sheet.max_row
    for col_idx, cell in enumerate(analysis_sheet[current_row], 1):
        if col_idx == 3:
            cell.alignment = Alignment(vertical='bottom', wrap_text=True)
        else:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Add to raw sheet
    raw_row = [folder_name, page_number, raw_response]
    raw_sheet.append(raw_row)
    raw_sheet.row_dimensions[raw_sheet.max_row].height = max_row_height

    for cell in raw_sheet[raw_sheet.max_row]:
        cell.alignment = Alignment(vertical='top', wrap_text=True)


def create_result_entry(folder_name, page_number, img_path, response_data, raw_response):
    """Create a result entry dictionary for JSON output."""
    return {
        'folder': folder_name,
        'page_number': page_number,
        'image_path': img_path,
        'analysis': {
            'title': response_data.get('title', ''),
            'contributors': response_data.get('contributors', []),
            'genre': response_data.get('genre', ''),
            'ocr_text': response_data.get('ocrText', ''),
            'description': response_data.get('description', ''),
            'format_media': response_data.get('formatMedia', ''),
            'subjects': response_data.get('subjects', []),
            'date_on_drawing': response_data.get('dateOnDrawing', ''),
            'sheet_info': response_data.get('sheetInfo', ''),
            'named_entities': response_data.get('namedEntities', []),
            'geographic_entities': response_data.get('geographicEntities', []),
            'content_warning': response_data.get('contentWarning', 'None'),
            'raw_response': raw_response
        }
    }


def process_folder_individual(all_images, wb, analysis_sheet, raw_sheet, issues_sheet,
                              logs_folder_path, model_name, max_row_height, all_results,
                              collection_context=""):
    """Process using individual API calls."""
    items_with_issues = 0
    total_processing_time = 0

    for i, (folder_name, page_number, img_path) in enumerate(all_images):
        row_number = i + 2
        filename = os.path.basename(img_path)

        print(f"\nProcessing drawing {i+1}/{len(all_images)}")
        print(f"   File: {filename}")

        try:
            response_data, raw_response, usage, processing_time = process_image(
                img_path, model_name, collection_context
            )
            total_processing_time += processing_time

            log_individual_response(
                logs_folder_path=logs_folder_path,
                script_name="architectural_drawings_metadata",
                row_number=row_number,
                barcode=f"{folder_name}_drawing{page_number}",
                response_text=raw_response,
                model_name=model_name,
                prompt_tokens=usage.input_tokens if usage else 0,
                completion_tokens=usage.output_tokens if usage else 0,
                processing_time=processing_time
            )

            print(f"   Processed successfully - Tokens: {(usage.input_tokens + usage.output_tokens) if usage else 0:,}")

        except Exception as e:
            logging.error(f"Error processing {img_path}: {str(e)}")
            items_with_issues += 1
            raw_response = f"Processing error: {str(e)}"
            response_data = create_error_response(raw_response, str(e))
            usage = None

            issues_sheet.append([img_path, str(e)])

            log_individual_response(
                logs_folder_path=logs_folder_path,
                script_name="architectural_drawings_metadata",
                row_number=row_number,
                barcode=f"{folder_name}_drawing{page_number}",
                response_text=raw_response,
                model_name=model_name,
                prompt_tokens=0,
                completion_tokens=0,
                processing_time=0
            )

            print(f"   Processing failed: {str(e)}")

        # Add to sheets
        add_to_sheets(analysis_sheet, raw_sheet, folder_name, page_number,
                     img_path, response_data, raw_response, max_row_height)

        # Add to results
        all_results.append(create_result_entry(folder_name, page_number,
                                               img_path, response_data, raw_response))

        # Add delay between requests
        time.sleep(1)

    return (wb, all_results, api_stats, len(all_images), items_with_issues, total_processing_time,
           api_stats.total_input_tokens, api_stats.total_output_tokens, False)


def main():
    # Check for config override via environment variable
    model_name = os.getenv('CONFIG_MODEL', DEFAULT_MODEL)

    # Start timing the entire script execution
    script_start_time = time.time()

    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Check for config override via environment variable
    input_folder_name = os.getenv('CONFIG_IMAGE_FOLDER', "alfred-zucker")
    input_folder = os.path.join(script_dir, "image_folders", input_folder_name)

    # Load collection context from any JSON files in the input folder
    collection_context = load_collection_context(input_folder=input_folder)

    # Optional visibility for troubleshooting
    if not os.path.exists(input_folder):
        print(f"Input folder not found: {os.path.abspath(input_folder)}")
        return 1

    print(f"\nARCHITECTURAL DRAWINGS METADATA EXTRACTION (Claude)\n")
    if collection_context:
        print(f"Collection context: {input_folder_name}")
    print(f"Using input folder: image_folders/{input_folder_name}")


    # Create dynamic output folder name
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H-%M-%S")

    # Create folder name
    folder_name = f"ArchImagesAI_{input_folder_name}_{model_name}_{current_date}_Time_{current_time}"

    # Create the full output directory path
    base_output_dir = os.path.join(script_dir, "output_folders")
    output_dir = os.path.join(base_output_dir, folder_name)

    # Create the directory
    os.makedirs(output_dir, exist_ok=True)

    # Create metadata folder structure
    metadata_folder = os.path.join(output_dir, "metadata")
    os.makedirs(metadata_folder, exist_ok=True)

    print(f"Output directory: output_folders/{folder_name}")

    # Process folder
    (wb, all_results, api_stats, total_items, items_with_issues, total_processing_time,
     total_prompt_tokens, total_completion_tokens, was_batch_processed) = process_folder_with_batch(
        input_folder, output_dir, model_name, collection_context
    )

    # Add API Stats sheet
    api_summary = {
        "total_requests": api_stats.total_requests,
        "total_input_tokens": total_prompt_tokens,
        "total_output_tokens": total_completion_tokens,
        "total_tokens": total_prompt_tokens + total_completion_tokens,
        "processing_mode": "BATCH" if was_batch_processed else "INDIVIDUAL"
    }

    all_results.append({"api_stats": api_summary})

    stats_sheet = wb.create_sheet("API Stats")
    stats_sheet.append(["Metric", "Value"])
    for key, value in api_summary.items():
        stats_sheet.append([key, value])

    # Save files
    excel_path = os.path.join(metadata_folder, "drawings_workflow.xlsx")
    json_path = os.path.join(metadata_folder, "drawings_workflow.json")

    wb.save(excel_path)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False)


    # Calculate script metrics
    script_duration = time.time() - script_start_time

    # Calculate actual cost
    estimated_cost = calculate_cost(
        model_name=model_name,
        prompt_tokens=total_prompt_tokens,
        completion_tokens=total_completion_tokens,
        is_batch=was_batch_processed
    )

    # Create logs folder
    logs_folder_path = os.path.join(output_dir, "logs")
    if not os.path.exists(logs_folder_path):
        os.makedirs(logs_folder_path)

    # Create standardized token usage log
    create_token_usage_log(
        logs_folder_path=logs_folder_path,
        script_name="architectural_drawings_metadata",
        model_name=model_name,
        total_items=total_items,
        items_with_issues=items_with_issues,
        total_time=total_processing_time,
        total_prompt_tokens=total_prompt_tokens,
        total_completion_tokens=total_completion_tokens,
        additional_metrics={
            "Total script execution time": f"{script_duration:.2f}s",
            "Processing time percentage": f"{(total_processing_time/script_duration)*100:.1f}%" if script_duration > 0 else "0%",
            "Items successfully processed": total_items - items_with_issues,
            "Processing mode": "BATCH" if was_batch_processed else "INDIVIDUAL",
            "Cost": f"${estimated_cost:.4f}",
            "Average tokens per item": f"{(total_prompt_tokens + total_completion_tokens)/total_items:.0f}" if total_items > 0 else "0"
        }
    )

    # Final summary - terminal output
    print(f"\nSTEP 1 COMPLETE")
    print(f"Successfully processed: {total_items - items_with_issues}/{total_items} drawings")
    print(f"Items with issues: {items_with_issues}")
    print(f"Total script time: {script_duration:.1f}s ({script_duration/60:.1f} minutes)")
    print(f"Tokens: {total_prompt_tokens + total_completion_tokens:,} (Input: {total_prompt_tokens:,}, Output: {total_completion_tokens:,})")
    print(f"Cost estimate: ${estimated_cost:.4f}")


if __name__ == "__main__":
    exit(main())
